#!/usr/bin/env python3
"""Find SpreadsheetBench rendered images with unusually large blank regions."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from PIL import Image, ImageStat
from tqdm import tqdm

Image.MAX_IMAGE_PIXELS = None


SCRIPT_DIR = Path(__file__).resolve().parent
SPREADSHEETBENCH_DIR = SCRIPT_DIR.parent
REPO_ROOT = SPREADSHEETBENCH_DIR.parent.parent.resolve()
SPLITS = {
    "verified_400": {
        "root": SPREADSHEETBENCH_DIR / "spreadsheetbench_verified_400",
        "suffixes": ("init",),
    },
    "all_912": {
        "root": SPREADSHEETBENCH_DIR / "all_data_912_v0.1",
        "suffixes": ("init", "input"),
    },
}


def repo_relative(path: Path | str) -> str:
    path_obj = Path(path)
    try:
        return path_obj.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return str(path_obj)


def resolve_repo_path(path: Path | str) -> Path:
    path_obj = Path(path)
    if path_obj.is_absolute():
        return path_obj
    return REPO_ROOT / path_obj


def parse_image_name(path: Path) -> dict[str, Any]:
    stem = path.stem
    sheet_index = None
    match = re.search(r"___(\d+)$", stem)
    if match:
        sheet_index = int(match.group(1))
        stem = stem[: match.start()]

    file_match = re.match(r"(?P<case>\d+)_(?P<id>.+)_(?P<suffix>init|input)$", stem)
    if not file_match:
        return {"file_stem": stem, "sheet_index": sheet_index}

    parsed = file_match.groupdict()
    parsed["file_stem"] = stem
    parsed["sheet_index"] = sheet_index
    return parsed


def source_workbook(split_root: Path, parsed: dict[str, Any]) -> Path | None:
    item_id = parsed.get("id")
    file_stem = parsed.get("file_stem")
    if not item_id or not file_stem:
        return None
    path = split_root / "spreadsheet" / item_id / f"{file_stem}.xlsx"
    return path if path.exists() else None


def image_stats(path: Path, white_threshold: int) -> dict[str, Any]:
    image = Image.open(path)
    width, height = image.size
    sample = image.convert("RGB")
    scale = 1.0
    max_sample_pixels = 25_000_000
    if width * height > max_sample_pixels:
        scale = (max_sample_pixels / (width * height)) ** 0.5
        sample_size = (max(1, int(width * scale)), max(1, int(height * scale)))
        sample = sample.resize(sample_size, Image.Resampling.BOX)

    gray = sample.convert("L")
    mask = gray.point(lambda pixel: 0 if pixel >= white_threshold else 255, mode="1")
    bbox = mask.getbbox()
    stat = ImageStat.Stat(sample)

    if bbox is None:
        content_ratio = 0.0
        margins = [width, height, width, height]
    else:
        left, top, right, bottom = bbox
        sample_width, sample_height = sample.size
        content_ratio = ((right - left) * (bottom - top)) / (sample_width * sample_height)
        margins = [
            round(left / scale),
            round(top / scale),
            round((sample_width - right) / scale),
            round((sample_height - bottom) / scale),
        ]

    extrema = stat.extrema
    is_exact_white = all(channel_min == 255 and channel_max == 255 for channel_min, channel_max in extrema)
    return {
        "width": width,
        "height": height,
        "area": width * height,
        "content_ratio": round(content_ratio, 6),
        "margins": margins,
        "max_margin_ratio": round(max(margins) / max(width, height), 6),
        "mean_rgb": [round(value, 2) for value in stat.mean],
        "is_exact_white": is_exact_white,
    }


def sheet_stats(workbook_path: Path | None, sheet_index: int | None) -> dict[str, Any]:
    if workbook_path is None:
        return {}

    wb = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        worksheets = wb.worksheets
        if sheet_index is None:
            sheet_index = 1
        if sheet_index < 1 or sheet_index > len(worksheets):
            return {"sheet_index": sheet_index, "sheet_error": "sheet index out of range"}

        ws = worksheets[sheet_index - 1]
        value_cells = 0
        styled_cells = 0
        min_row = min_col = max_row = max_col = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    value_cells += 1
                    min_row = cell.row if min_row is None else min(min_row, cell.row)
                    max_row = cell.row if max_row is None else max(max_row, cell.row)
                    min_col = cell.column if min_col is None else min(min_col, cell.column)
                    max_col = cell.column if max_col is None else max(max_col, cell.column)
                if cell.has_style:
                    styled_cells += 1

        return {
            "sheet_index": sheet_index,
            "sheet_name": ws.title,
            "sheet_state": ws.sheet_state,
            "sheet_dimension": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "value_cells": value_cells,
            "styled_cells": styled_cells,
            "value_bounds": [min_row, min_col, max_row, max_col],
            "merged_ranges": len(ws.merged_cells.ranges),
        }
    finally:
        wb.close()


def reason(record: dict[str, Any]) -> str:
    if record.get("is_exact_white"):
        if record.get("value_cells") == 0:
            return "blank worksheet exported as a full white PDF page"
        return "rendered page is all white although workbook has values; likely hidden/off-page content or print/export issue"
    if record.get("content_ratio", 1.0) <= 0.25 and record.get("value_cells") == 0:
        return "worksheet has no values; remaining visible content is mostly page/background"
    if record.get("area", 0) >= 8_000_000:
        if record.get("max_row", 0) >= 1000 or record.get("max_column", 0) >= 26:
            return "large worksheet extent from styles/used range creates a large export canvas"
        return "large real table or print area creates a large export canvas"
    if record.get("max_margin_ratio", 0) >= 0.25:
        return "large white margin around detected content"
    return "candidate by image-size/content-ratio threshold"


def collect(split_names: list[str], white_threshold: int, image_folder: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for split in split_names:
        split_root = SPLITS[split]["root"]
        image_root = split_root / image_folder
        paths = sorted(image_root.glob("*.png"))
        for image_path in tqdm(paths, desc=f"Scanning {split}", unit="image"):
            parsed = parse_image_name(image_path)
            workbook_path = source_workbook(split_root, parsed)
            stats = image_stats(image_path, white_threshold)
            record = {
                "split": split,
                "image_path": repo_relative(image_path),
                "workbook_path": repo_relative(workbook_path) if workbook_path else None,
                **parsed,
                **stats,
            }
            records.append(record)
    return records


def add_sheet_context(records: list[dict[str, Any]]) -> None:
    for record in tqdm(records, desc="Reading candidate workbooks", unit="image"):
        workbook_path = resolve_repo_path(record["workbook_path"]) if record.get("workbook_path") else None
        sheet = sheet_stats(workbook_path, record.get("sheet_index"))
        record.update(sheet)
        record["reason"] = reason(record)


def is_candidate(record: dict[str, Any], min_area: int, max_content_ratio: float, max_margin_ratio: float) -> bool:
    if record.get("is_exact_white"):
        return True
    if record.get("area", 0) >= min_area and record.get("content_ratio", 1.0) <= max_content_ratio:
        return True
    if record.get("max_margin_ratio", 0.0) >= max_margin_ratio:
        return True
    return False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Check SpreadsheetBench PNGs for large blank/whitespace areas.")
    parser.add_argument("--splits", default="verified_400,all_912")
    parser.add_argument("--image-folder", default="image", help="Folder under each split root containing PNGs.")
    parser.add_argument("--white-threshold", type=int, default=245)
    parser.add_argument("--min-area", type=int, default=500_000)
    parser.add_argument("--max-content-ratio", type=float, default=0.25)
    parser.add_argument("--max-margin-ratio", type=float, default=0.25)
    parser.add_argument("--top", type=int, default=50)
    parser.add_argument("--output", default=repo_relative(SCRIPT_DIR / "large_whitespace_report.json"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    split_names = [name.strip() for name in args.splits.split(",") if name.strip()]
    records = collect(split_names, args.white_threshold, args.image_folder)
    candidates = [
        record
        for record in records
        if is_candidate(record, args.min_area, args.max_content_ratio, args.max_margin_ratio)
    ]
    candidates.sort(
        key=lambda record: (
            not record.get("is_exact_white", False),
            record.get("content_ratio", 1.0),
            -record.get("area", 0),
        )
    )
    largest_canvas = sorted(records, key=lambda record: record.get("area", 0), reverse=True)[: args.top]
    top_candidates = candidates[: args.top]
    context_records = {record["image_path"]: record for record in top_candidates + largest_canvas}
    add_sheet_context(list(context_records.values()))

    report = {
        "total_images": len(records),
        "candidate_count": len(candidates),
        "blank_image_count": sum(1 for record in records if record.get("is_exact_white")),
        "largest_canvas": largest_canvas,
        "candidates": top_candidates,
    }
    output_path = resolve_repo_path(args.output)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"total_images={report['total_images']}")
    print(f"candidate_count={report['candidate_count']}")
    print(f"blank_image_count={report['blank_image_count']}")
    print(f"report={repo_relative(output_path)}")
    for record in candidates[: min(args.top, 20)]:
        print(
            f"{record['image_path']} "
            f"size={record['width']}x{record['height']} "
            f"content_ratio={record['content_ratio']} "
            f"sheet={record.get('sheet_name')} "
            f"values={record.get('value_cells')} "
            f"styles={record.get('styled_cells')} "
            f"reason={record.get('reason')}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
