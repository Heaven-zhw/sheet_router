#!/usr/bin/env python3
"""Render SpreadsheetBench workbooks as RealHiTBench-style table images.

This renderer keeps the original XLSX package intact as much as possible.  For
each worksheet, it creates a temporary copy of the workbook where only that
worksheet is visible, sets a print area, asks LibreOffice to export to PDF, and
rasterizes the result to PNG.  The output has no Excel row/column headers, while
cell fills, borders, fonts, merged cells, and many workbook-level styles remain
those seen by LibreOffice.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image
from tqdm import tqdm


SCRIPT_DIR = Path(__file__).resolve().parent
SPREADSHEETBENCH_DIR = SCRIPT_DIR.parent
REPO_DIR = SPREADSHEETBENCH_DIR.parent.parent
REPO_ROOT = REPO_DIR.resolve()

if str(REPO_DIR) not in sys.path:
    sys.path.insert(0, str(REPO_DIR))

from core.excel2image_linux import _find_binary, _run_command, _trim_whitespace  # noqa: E402


DEFAULT_SPLITS = ("verified_400", "all_912")
SPLITS = {
    "verified_400": {
        "root": SPREADSHEETBENCH_DIR / "spreadsheetbench_verified_400",
        "preferred_suffixes": ("init",),
    },
    "all_912": {
        "root": SPREADSHEETBENCH_DIR / "all_data_912_v0.1",
        "preferred_suffixes": ("init", "input"),
    },
}

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def repo_relative(path: Path | str) -> str:
    path_obj = Path(path)
    try:
        return path_obj.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return str(path_obj)


def resolve_repo_path(path: Path | str) -> Path:
    path_obj = Path(path)
    if path_obj.is_absolute():
        return path_obj
    return REPO_ROOT / path_obj


@dataclass(frozen=True)
class WorkbookJob:
    split: str
    item_id: str
    source_path: Path
    output_dir: Path


@dataclass(frozen=True)
class WorksheetInfo:
    index: int
    name: str
    xml_path: str
    print_area: str
    orientation: str


def parse_ids(raw_ids: str | None) -> set[str] | None:
    if not raw_ids:
        return None
    return {item.strip() for item in raw_ids.split(",") if item.strip()}


def natural_key(path: Path) -> tuple:
    parts = re.split(r"(\d+)", path.name)
    return tuple(int(part) if part.isdigit() else part for part in parts)


def load_dataset(dataset_path: Path) -> list[dict]:
    with dataset_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, list):
        raise TypeError(f"Expected a JSON list in {dataset_path}, got {type(payload).__name__}")
    return payload


def find_source_workbooks(item: dict, split_root: Path, preferred_suffixes: Iterable[str]) -> list[Path]:
    item_id = str(item["id"])
    rel_dir = item.get("spreadsheet_path") or f"spreadsheet/{item_id}"
    workbook_dir = split_root / rel_dir

    for suffix in preferred_suffixes:
        candidates = sorted(workbook_dir.glob(f"*_{suffix}.xlsx"), key=natural_key)
        if candidates:
            return candidates
    return []


def build_jobs(
    split_names: list[str],
    selected_ids: set[str] | None,
    limit: int | None,
    output_folder: str,
) -> list[WorkbookJob]:
    jobs: list[WorkbookJob] = []
    for split in split_names:
        config = SPLITS[split]
        split_root = config["root"]
        output_dir = split_root / output_folder
        dataset = load_dataset(split_root / "dataset.json")

        for item in dataset:
            item_id = str(item["id"])
            if selected_ids is not None and item_id not in selected_ids:
                continue

            for source_path in find_source_workbooks(item, split_root, config["preferred_suffixes"]):
                jobs.append(
                    WorkbookJob(
                        split=split,
                        item_id=item_id,
                        source_path=source_path,
                        output_dir=output_dir,
                    )
                )

            if limit is not None and len(jobs) >= limit:
                return jobs[:limit]
    return jobs


def has_meaningful_style(cell) -> bool:
    fill = cell.fill
    if fill and fill.fill_type:
        return True

    border = cell.border
    if border:
        for side in (border.left, border.right, border.top, border.bottom):
            if side and side.style:
                return True

    return False


def used_bounds(ws) -> tuple[int, int, int, int]:
    min_row = min_col = None
    max_row = max_col = None

    def include(row: int, col: int) -> None:
        nonlocal min_row, min_col, max_row, max_col
        min_row = row if min_row is None else min(min_row, row)
        max_row = row if max_row is None else max(max_row, row)
        min_col = col if min_col is None else min(min_col, col)
        max_col = col if max_col is None else max(max_col, col)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, "") or has_meaningful_style(cell):
                include(cell.row, cell.column)

    for merged_range in ws.merged_cells.ranges:
        include(merged_range.min_row, merged_range.min_col)
        include(merged_range.max_row, merged_range.max_col)

    if min_row is None:
        return (1, 1, 1, 1)
    return (min_row, min_col, max_row, max_col)


def absolute_range(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    start = f"${get_column_letter(min_col)}${min_row}"
    end = f"${get_column_letter(max_col)}${max_row}"
    return f"{start}:{end}"


def output_paths_for_workbook(job: WorkbookJob, sheet_count: int) -> list[Path]:
    stem = job.source_path.stem
    if sheet_count == 1:
        return [job.output_dir / f"{stem}.png"]
    return [job.output_dir / f"{stem}___{idx}.png" for idx in range(1, sheet_count + 1)]


def quote_sheet_name(name: str) -> str:
    escaped = name.replace("'", "''")
    return f"'{escaped}'"


def xml_tag(name: str) -> str:
    return f"{{{NS_MAIN}}}{name}"


def rel_tag(name: str) -> str:
    return f"{{{NS_PKG_REL}}}{name}"


def resolve_workbook_sheet_paths(src_path: Path) -> list[str]:
    with zipfile.ZipFile(src_path) as zin:
        workbook_xml = ET.fromstring(zin.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

    rid_to_target = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_xml.findall(rel_tag("Relationship"))
    }
    paths = []
    for sheet in workbook_xml.findall(f"{xml_tag('sheets')}/{xml_tag('sheet')}"):
        rid = sheet.attrib.get(f"{{{NS_REL}}}id")
        target = rid_to_target.get(rid, "")
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        paths.append(target)
    return paths


def collect_worksheets(src_path: Path, include_hidden: bool) -> list[WorksheetInfo]:
    sheet_paths = resolve_workbook_sheet_paths(src_path)
    wb = openpyxl.load_workbook(src_path, data_only=False, read_only=False)
    try:
        worksheets = [
            (idx, ws)
            for idx, ws in enumerate(wb.worksheets)
            if include_hidden or ws.sheet_state == "visible"
        ]
        if not worksheets:
            worksheets = list(enumerate(wb.worksheets))

        infos = []
        for idx, ws in worksheets:
            bounds = used_bounds(ws)
            min_row, min_col, max_row, max_col = bounds
            row_count = max_row - min_row + 1
            col_count = max_col - min_col + 1
            infos.append(
                WorksheetInfo(
                    index=idx,
                    name=ws.title,
                    xml_path=sheet_paths[idx],
                    print_area=absolute_range(bounds),
                    orientation="landscape" if col_count > row_count else "portrait",
                )
            )
        return infos
    finally:
        wb.close()


def ensure_child(parent: ET.Element, tag: str, after_tags: tuple[str, ...] = ()) -> ET.Element:
    child = parent.find(xml_tag(tag))
    if child is not None:
        return child

    child = ET.Element(xml_tag(tag))
    insert_at = 0
    for idx, existing in enumerate(list(parent)):
        local_name = existing.tag.rsplit("}", 1)[-1]
        if local_name in after_tags:
            insert_at = idx + 1
    parent.insert(insert_at, child)
    return child


def patch_workbook_xml(raw: bytes, target: WorksheetInfo) -> bytes:
    root = ET.fromstring(raw)
    sheets_el = root.find(xml_tag("sheets"))
    if sheets_el is None:
        raise RuntimeError("workbook.xml has no sheets element")

    for idx, sheet in enumerate(sheets_el.findall(xml_tag("sheet"))):
        if idx == target.index:
            sheet.attrib.pop("state", None)
        else:
            sheet.set("state", "hidden")

    book_views = root.find(xml_tag("bookViews"))
    if book_views is not None:
        workbook_view = book_views.find(xml_tag("workbookView"))
        if workbook_view is not None:
            workbook_view.set("activeTab", str(target.index))

    defined_names = root.find(xml_tag("definedNames"))
    if defined_names is None:
        defined_names = ET.Element(xml_tag("definedNames"))
        root.append(defined_names)

    for defined_name in list(defined_names.findall(xml_tag("definedName"))):
        if (
            defined_name.attrib.get("name") == "_xlnm.Print_Area"
            and defined_name.attrib.get("localSheetId") == str(target.index)
        ):
            defined_names.remove(defined_name)

    print_area = ET.Element(
        xml_tag("definedName"),
        {"name": "_xlnm.Print_Area", "localSheetId": str(target.index)},
    )
    print_area.text = f"{quote_sheet_name(target.name)}!{target.print_area}"
    defined_names.append(print_area)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_sheet_xml(raw: bytes, target: WorksheetInfo, print_gridlines: bool) -> bytes:
    root = ET.fromstring(raw)

    sheet_pr = ensure_child(root, "sheetPr")
    page_setup_pr = ensure_child(sheet_pr, "pageSetUpPr")
    page_setup_pr.set("fitToPage", "1")
    page_setup_pr.set("autoPageBreaks", "1")

    sheet_views = root.find(xml_tag("sheetViews"))
    if sheet_views is not None:
        for sheet_view in sheet_views.findall(xml_tag("sheetView")):
            sheet_view.set("showGridLines", "1" if print_gridlines else "0")

    print_options = ensure_child(root, "printOptions", after_tags=("sheetPr", "dimension", "sheetViews"))
    print_options.set("gridLines", "1" if print_gridlines else "0")
    print_options.set("gridLinesSet", "1" if print_gridlines else "0")

    page_margins = ensure_child(
        root,
        "pageMargins",
        after_tags=("sheetData", "mergeCells", "conditionalFormatting", "dataValidations", "hyperlinks"),
    )
    page_margins.set("left", "0.05")
    page_margins.set("right", "0.05")
    page_margins.set("top", "0.05")
    page_margins.set("bottom", "0.05")
    page_margins.set("header", "0")
    page_margins.set("footer", "0")

    page_setup = ensure_child(root, "pageSetup", after_tags=("pageMargins",))
    page_setup.set("orientation", target.orientation)
    page_setup.set("fitToWidth", "1")
    page_setup.set("fitToHeight", "0")
    page_setup.attrib.pop("scale", None)
    page_setup.attrib.pop("usePrinterDefaults", None)

    for tag in ("rowBreaks", "colBreaks", "headerFooter"):
        child = root.find(xml_tag(tag))
        if child is not None:
            root.remove(child)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def make_single_sheet_workbook(
    src_path: Path,
    target: WorksheetInfo,
    staged_path: Path,
    print_gridlines: bool,
) -> None:
    with zipfile.ZipFile(src_path, "r") as zin, zipfile.ZipFile(staged_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            raw = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                raw = patch_workbook_xml(raw, target)
            elif info.filename == target.xml_path:
                raw = patch_sheet_xml(raw, target, print_gridlines)
            zout.writestr(info, raw)


def convert_workbook_to_pdf(staged_path: Path, raw_dir: Path, timeout: int) -> Path:
    soffice_bin = _find_binary("SOFFICE_BIN", "libreoffice")
    profile_dir = raw_dir / "libreoffice-profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    profile_uri = profile_dir.resolve().as_uri()

    _run_command(
        [
            soffice_bin,
            f"-env:UserInstallation={profile_uri}",
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            str(raw_dir),
            str(staged_path),
        ],
        timeout=timeout,
    )

    pdf_path = raw_dir / f"{staged_path.stem}.pdf"
    if not pdf_path.exists():
        raise RuntimeError(f"LibreOffice did not produce the expected PDF: {pdf_path}")
    return pdf_path


def rasterize_pdf(pdf_path: Path, raw_dir: Path, dpi: int, timeout: int) -> list[Path]:
    pdftoppm_bin = _find_binary("PDFTOPPM_BIN", "pdftoppm")
    prefix = raw_dir / pdf_path.stem
    _run_command([pdftoppm_bin, "-png", "-r", str(dpi), str(pdf_path), str(prefix)], timeout=timeout)

    single = raw_dir / f"{pdf_path.stem}.png"
    if single.exists():
        return [single]
    return sorted(raw_dir.glob(f"{pdf_path.stem}-*.png"), key=natural_key)


def save_pages_as_table_image(raw_paths: list[Path], out_path: Path, tmpdir: Path) -> int:
    if not raw_paths:
        raise RuntimeError(f"No rasterized PDF pages were produced for {out_path}")

    if len(raw_paths) == 1:
        _trim_whitespace(str(raw_paths[0]), str(out_path), padding=8)
        return 1

    trimmed_paths = []
    for idx, raw_path in enumerate(raw_paths, start=1):
        trimmed_path = tmpdir / f"trimmed_page_{idx}.png"
        _trim_whitespace(str(raw_path), str(trimmed_path), padding=4)
        trimmed_paths.append(trimmed_path)

    images = [Image.open(path).convert("RGB") for path in trimmed_paths]
    try:
        width = max(image.width for image in images)
        height = sum(image.height for image in images)
        stitched = Image.new("RGB", (width, height), "white")
        top = 0
        for image in images:
            stitched.paste(image, (0, top))
            top += image.height
        stitched.save(out_path)
    finally:
        for image in images:
            image.close()

    return len(raw_paths)


def render_workbook(
    job: WorkbookJob,
    dpi: int,
    overwrite: bool,
    include_hidden: bool,
    print_gridlines: bool,
    timeout: int,
) -> dict:
    worksheets = collect_worksheets(job.source_path, include_hidden)
    output_paths = output_paths_for_workbook(job, len(worksheets))
    if not overwrite and all(path.exists() for path in output_paths):
        return {
            "split": job.split,
            "id": job.item_id,
            "source": repo_relative(job.source_path),
            "images": [repo_relative(path) for path in output_paths],
            "sheets": [sheet.name for sheet in worksheets],
            "status": "skipped",
        }

    job.output_dir.mkdir(parents=True, exist_ok=True)
    rendered_paths = [path for path in output_paths if path.exists() and not overwrite]
    page_counts = {}

    for sheet_number, (worksheet, out_path) in enumerate(zip(worksheets, output_paths), start=1):
        if out_path.exists() and not overwrite:
            continue

        with tempfile.TemporaryDirectory(prefix="spreadsheetbench_table_image_") as tmpdir_name:
            tmpdir = Path(tmpdir_name)
            staged_path = tmpdir / f"sheet_{sheet_number}.xlsx"
            make_single_sheet_workbook(job.source_path, worksheet, staged_path, print_gridlines)
            pdf_path = convert_workbook_to_pdf(staged_path, tmpdir, timeout)
            raw_paths = rasterize_pdf(pdf_path, tmpdir, dpi=dpi, timeout=timeout)
            page_counts[worksheet.name] = save_pages_as_table_image(raw_paths, out_path, tmpdir)
            rendered_paths.append(out_path)

    return {
        "split": job.split,
        "id": job.item_id,
        "source": repo_relative(job.source_path),
        "images": [repo_relative(path) for path in rendered_paths],
        "sheets": [sheet.name for sheet in worksheets],
        "pdf_pages_per_sheet": page_counts,
        "status": "rendered",
    }


def check_external_tools() -> None:
    missing = [name for name in ("libreoffice", "pdftoppm") if shutil.which(name) is None]
    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(
            f"Missing external command(s): {joined}. "
            "Install LibreOffice Calc and poppler-utils, or set SOFFICE_BIN/PDFTOPPM_BIN."
        )


def save_manifest(records: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Render SpreadsheetBench XLSX files to RealHiTBench-style table PNG images."
    )
    parser.add_argument(
        "--splits",
        default=",".join(DEFAULT_SPLITS),
        help=f"Comma-separated split names. Available: {', '.join(SPLITS)}",
    )
    parser.add_argument("--ids", help="Comma-separated dataset ids to render.")
    parser.add_argument("--limit", type=int, help="Maximum number of workbook files to process.")
    parser.add_argument("--dpi", type=int, default=300, help="Rasterization DPI for PNG output.")
    parser.add_argument("--workers", type=int, default=1, help="Number of workbook render jobs to run in parallel.")
    parser.add_argument("--output-folder", default="image", help="Folder under each split root for PNG output.")
    parser.add_argument("--overwrite", action="store_true", help="Regenerate existing PNG files.")
    parser.add_argument("--include-hidden", action="store_true", help="Render hidden worksheets too.")
    parser.add_argument(
        "--no-gridlines",
        action="store_true",
        help="Do not print light worksheet gridlines for cells without explicit borders.",
    )
    parser.add_argument("--timeout", type=int, default=300, help="Timeout in seconds for LibreOffice/pdftoppm.")
    parser.add_argument("--dry-run", action="store_true", help="Print planned workbooks without rendering.")
    parser.add_argument(
        "--manifest",
        default=repo_relative(SCRIPT_DIR / "render_table_image_manifest.json"),
        help="Path to write a JSON manifest.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    split_names = [name.strip() for name in args.splits.split(",") if name.strip()]
    unknown = sorted(set(split_names) - set(SPLITS))
    if unknown:
        raise ValueError(f"Unknown split(s): {', '.join(unknown)}")

    selected_ids = parse_ids(args.ids)
    limit = args.limit if args.limit and args.limit > 0 else None
    jobs = build_jobs(split_names, selected_ids, limit, args.output_folder)

    print(f"Found {len(jobs)} workbook job(s).")
    missing_ids = selected_ids - {job.item_id for job in jobs} if selected_ids else set()
    if missing_ids:
        print(f"Warning: no workbook found for id(s): {', '.join(sorted(missing_ids))}")

    if args.dry_run:
        for job in jobs[:50]:
            print(f"[{job.split}] {job.item_id}: {repo_relative(job.source_path)} -> {repo_relative(job.output_dir)}")
        if len(jobs) > 50:
            print(f"... {len(jobs) - 50} more")
        return 0

    check_external_tools()
    records: list[dict] = []
    failures: list[dict] = []

    def run_one(job: WorkbookJob) -> dict:
        try:
            return render_workbook(
                job,
                dpi=args.dpi,
                overwrite=args.overwrite,
                include_hidden=args.include_hidden,
                print_gridlines=not args.no_gridlines,
                timeout=args.timeout,
            )
        except Exception as exc:
            return {
                "split": job.split,
                "id": job.item_id,
                "source": repo_relative(job.source_path),
                "status": "failed",
                "error": str(exc),
            }

    workers = max(1, args.workers)
    if workers == 1:
        iterator = (run_one(job) for job in jobs)
        for record in tqdm(iterator, total=len(jobs), desc="Rendering table images", unit="workbook"):
            records.append(record)
            if record.get("status") == "failed":
                failures.append(record)
                tqdm.write(f"FAILED {record.get('source')}: {record.get('error')}")
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_job = {executor.submit(run_one, job): job for job in jobs}
            for future in tqdm(
                as_completed(future_to_job),
                total=len(future_to_job),
                desc="Rendering table images",
                unit="workbook",
            ):
                record = future.result()
                records.append(record)
                if record.get("status") == "failed":
                    failures.append(record)
                    tqdm.write(f"FAILED {record.get('source')}: {record.get('error')}")

    manifest_path = resolve_repo_path(args.manifest)
    save_manifest(records, manifest_path)

    rendered = sum(1 for record in records if record.get("status") == "rendered")
    skipped = sum(1 for record in records if record.get("status") == "skipped")
    print(f"Done. rendered={rendered}, skipped={skipped}, failed={len(failures)}")
    print(f"Manifest: {repo_relative(manifest_path)}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
