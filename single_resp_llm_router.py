import argparse
import copy
import json
import os
import re
import shutil
import traceback
import warnings
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import suppress
from dataclasses import asdict
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import openpyxl
import requests
from openpyxl.utils import column_index_from_string, get_column_letter
from tqdm import tqdm

from core.routing import RouteDecision, SpreadsheetProfiler, unique_keep_order
from core.solver.realhit_cot import RealHiTCoTSolver
from core.solver.spreadsheet_pot import SPREADSHEET_DATA_SPLITS, SpreadSheetPoTSolver
from core.utils import load_jsonl, save_jsonl

import realhit_cot as realhit_entry
import spreadsheet_pot as spreadsheet_entry
from single_resp_router import (
    build_error_result,
    build_eval_entry,
    cleanup_intermediate_files,
    output_filenames,
    report_scores,
    run_solver,
)


repo_dir = os.path.abspath(os.path.dirname(__file__))


SINGLE_FORMATS = (
    "markdown",
    "official_latex",
    "latex",
    "html",
    "csv",
    "dataframe",
    "json_rows",
    "json_cells",
    "image",
    "excel_1_image",
    "default_image",
)
TEXT_FORMATS = {"markdown", "official_latex", "latex", "html", "csv", "dataframe", "json_rows", "json_cells"}
IMAGE_FORMATS = {"image", "excel_1_image", "default_image"}
QA_PRIORITY = (
    "latex",
    "json_cells",
    "markdown",
    "json_rows",
    "html",
    "dataframe",
    "csv",
    "image",
    "excel_1_image",
    "default_image",
)
MANIPULATION_PRIORITY = (
    "json_cells",
    "json_rows",
    "markdown",
    "latex",
    "html",
    "dataframe",
    "csv",
    "excel_1_image",
    "image",
    "default_image",
)
QA_BASE_SCORES = {
    "latex": 3.0,
    "json_cells": 2.6,
    "markdown": 2.3,
    "json_rows": 2.0,
    "html": 1.5,
    "dataframe": 1.2,
    "csv": 1.1,
    "image": 0.5,
    "excel_1_image": 0.4,
    "default_image": 0.2,
}
MANIPULATION_BASE_SCORES = {
    "json_cells": 4.0,
    "json_rows": 3.0,
    "markdown": 2.2,
    "latex": 2.0,
    "html": 1.8,
    "dataframe": 1.5,
    "csv": 1.4,
    "excel_1_image": 0.9,
    "image": 0.7,
    "default_image": 0.4,
}

VISUAL_TERMS = (
    "color",
    "colour",
    "background",
    "fill",
    "font",
    "bold",
    "border",
    "highlight",
    "shade",
    "conditional formatting",
    "format reference",
    "formatting reference",
    "same format",
    "keep format",
    "keep formatting",
    "preserve format",
    "preserve formatting",
    "cell format",
    "number format",
    "row height",
    "column width",
    "chart",
    "picture",
    "screenshot",
)
FORMAT_FALSE_POSITIVES = (
    "answer format",
    "output format",
    "final answer format",
    "format your answer",
    "format the answer",
)
ROW_RECORD_TERMS = (
    "match",
    "matching",
    "duplicate",
    "lookup",
    "filter",
    "sort",
    "group",
    "group by",
    "sum by",
    "for each row",
    "each row",
    "append row",
    "add row",
    "subtotal",
    "total row",
)
NUMERIC_TERMS = (
    "max",
    "min",
    "sum",
    "total",
    "average",
    "mean",
    "median",
    "count",
    "rate",
    "percentage",
    "percent",
    "exceed",
    "below",
    "above",
    "between",
    "top",
    "bottom",
    "highest",
    "lowest",
    "largest",
    "smallest",
    "greater",
    "less",
    "calculate",
)
STRUCTURE_TERMS = (
    "header",
    "row header",
    "column header",
    "multi-level",
    "multilevel",
    "hierarchy",
    "nested",
    "sub-table",
    "subtable",
    "above",
    "below",
    "left",
    "right",
    "section",
    "category",
    "merge",
    "merged",
)
COORDINATE_PATTERNS = (
    r"\b[A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?\b",
    r"\brow\s+\d+\b",
    r"\bcolumn\s+[A-Z]{1,3}\b",
)

ROUTER_SYSTEM_PROMPT = """You are a model-independent spreadsheet representation router.

Your job is to tag what evidence a downstream spreadsheet solver needs.
Do not choose the table format directly. Do not answer the spreadsheet question. Do not write code.

The tags must depend only on the spreadsheet file summary and the user query.
Never use the downstream solver model name, model family, or per-model performance.

Be conservative about visual/style tags: words like "answer format" or "output format" do not mean spreadsheet visual style.
Return only valid JSON."""

ROUTER_USER_PROMPT_TEMPLATE = """# Task
Tag the evidence needs for this spreadsheet task.

# Synthetic Examples
Example A:
Query: "Which product has the largest Q3 margin under the Retail group?"
Workbook clue: merged group headers and quarter subcolumns.
Output tags: complex_headers=true, numeric_reasoning=true, exact_coordinates=false, row_records=false, visual_style=false.

Example B:
Query: "Filter orders after 2024-03-01 by customer ID, sort them by date, and add a subtotal row."
Workbook clue: dense table with clear row records.
Output tags: row_records=true, numeric_reasoning=true, exact_coordinates=false, complex_headers=false, visual_style=false.

Example C:
Query: "Fill the target block using the sample block as a formatting reference and keep the same borders and colors."
Workbook clue: styled cells and target range.
Output tags: visual_style=true, format_reference=true, exact_coordinates=true.

Example D:
Query: "What value appears at cell F12?"
Workbook clue: sparse sheet with many blank cells.
Output tags: exact_coordinates=true, sparse_or_large=true.

# Input
{routing_input_json}

# Tag Meanings
- visual_style: the solver needs colors, fills, borders, bold/italic, row height, column width, charts, images, or screenshot layout.
- format_reference: the query asks to preserve/copy/use visual formatting from a sample range.
- exact_coordinates: exact cells, ranges, output positions, or sheet names are central.
- row_records: the task is like filtering, matching, grouping, sorting, deduplicating, appending rows, or aggregating by keys.
- complex_headers: merged headers, multi-level headers, nested sections, row/column header hierarchy, or cross-header lookup matter.
- sparse_or_large: the workbook is sparse, wide, long, or likely to be hard to serialize densely.
- multi_sheet: reasoning across multiple worksheets matters.
- numeric_reasoning: arithmetic, max/min, count, comparison, percentage, or totals matter.

# Output JSON Schema
{{
  "needs": {{
    "visual_style": false,
    "format_reference": false,
    "exact_coordinates": false,
    "row_records": false,
    "complex_headers": false,
    "sparse_or_large": false,
    "multi_sheet": false,
    "numeric_reasoning": false
  }},
  "default_risk": "none | low | medium | high",
  "evidence": ["short evidence 1", "short evidence 2"]
}}"""

ROUTER_REPAIR_PROMPT_TEMPLATE = """Your previous response was not a valid evidence tagging JSON.

Format error:
{error}

Return only one valid JSON object matching this schema:
{{
  "needs": {{
    "visual_style": false,
    "format_reference": false,
    "exact_coordinates": false,
    "row_records": false,
    "complex_headers": false,
    "sparse_or_large": false,
    "multi_sheet": false,
    "numeric_reasoning": false
  }},
  "default_risk": "none",
  "evidence": ["short evidence"]
}}

Do not answer the spreadsheet question. Do not use Markdown fences."""


def safe_name(value: Optional[str]) -> str:
    return (value or "default_model").replace("/", "_").replace("\\", "_")


def estimate_tokens_from_chars(num_chars: int) -> int:
    return max(1, int(num_chars / 3.5))


def text_contains_any(text: str, terms: Sequence[str]) -> bool:
    lowered = (text or "").lower()
    return any(term in lowered for term in terms)


def first_available(order: Sequence[str], available: Set[str]) -> Optional[str]:
    for value in order:
        if value in available:
            return value
    return None


def task_mode_for_dataset(dataset: str) -> str:
    return "qa_cot" if dataset == "realhitbench" else "manipulation_pot"


def solver_mode_for_dataset(dataset: str) -> str:
    return "cot_qa" if dataset == "realhitbench" else "pot_code"


def priority_for_task(task_mode: str) -> Tuple[str, ...]:
    return QA_PRIORITY if task_mode == "qa_cot" else MANIPULATION_PRIORITY


def default_for_task(task_mode: str) -> str:
    return "latex" if task_mode == "qa_cot" else "json_cells"


def image_priority(dataset: str) -> Tuple[str, ...]:
    if dataset == "realhitbench":
        return ("image", "excel_1_image", "default_image")
    return ("excel_1_image", "image", "default_image")


def normalize_chat_url(url: Optional[str]) -> str:
    base_url = (url or "localhost:8000").rstrip("/")
    if not base_url.startswith(("http://", "https://")):
        base_url = f"http://{base_url}"
    if base_url.endswith("/chat/completions"):
        return base_url
    if not base_url.endswith("/v1"):
        base_url = f"{base_url}/v1"
    return f"{base_url}/chat/completions"


def auth_headers(api_key: Optional[str]) -> Optional[Dict[str, str]]:
    key = api_key or os.environ.get("API_KEY") or os.environ.get("OPENAI_API_KEY")
    if not key:
        return None
    return {"Authorization": f"Bearer {key}"}


def chat_completion_content(
    *,
    url: Optional[str],
    api_key: Optional[str],
    messages: List[Dict[str, Any]],
    model_name: Optional[str],
    temperature: float,
    top_p: float,
    max_tokens: int,
    timeout: int,
) -> str:
    model_key = (model_name or "").lower()
    payload = {
        "messages": messages,
        "model": model_name or "model",
        "max_tokens": max_tokens,
        "temperature": temperature,
        "top_p": top_p,
        "skip_special_tokens": True if model_key.startswith("gemma-3") else False,
        "spaces_between_special_tokens": False,
        "chat_template_kwargs": {"enable_thinking": False},
    }
    last_error = None
    for _ in range(3):
        try:
            response = requests.post(
                url=normalize_chat_url(url),
                json=payload,
                headers=auth_headers(api_key),
                verify=False,
                timeout=timeout,
            )
            data = response.json()
            if "choices" in data and data["choices"]:
                return data["choices"][0].get("message", {}).get("content") or ""
            if "error" in data:
                raise RuntimeError(str(data["error"]))
            raise RuntimeError(f"Unexpected chat response keys: {sorted(data.keys())}")
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Router LLM call failed: {last_error}")


def strip_markdown_fence(text: str) -> str:
    text = (text or "").strip()
    fence = re.fullmatch(r"```(?:json)?\s*(.*?)\s*```", text, flags=re.S | re.I)
    if fence:
        return fence.group(1).strip()
    return text


def remove_thinking_blocks(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text or "", flags=re.S | re.I).strip()


def json_candidates(text: str) -> List[str]:
    cleaned = strip_markdown_fence(remove_thinking_blocks(text))
    candidates = [cleaned]
    json_fence = re.search(r"```json\s*(.*?)\s*```", text or "", flags=re.S | re.I)
    if json_fence:
        candidates.append(json_fence.group(1).strip())
    generic_fence = re.search(r"```\s*(.*?)\s*```", text or "", flags=re.S)
    if generic_fence:
        candidates.append(generic_fence.group(1).strip())
    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start != -1 and end != -1 and end > start:
        candidates.append(cleaned[start : end + 1])

    out = []
    seen = set()
    for candidate in candidates:
        candidate = candidate.strip()
        if candidate and candidate not in seen:
            out.append(candidate)
            seen.add(candidate)
    return out


def parse_router_json(text: str, available_formats: Set[str]) -> Dict[str, Any]:
    errors = []
    payload = None
    for candidate in json_candidates(text):
        try:
            payload = json.loads(candidate)
            break
        except Exception as exc:
            errors.append(str(exc))
    if payload is None:
        raise ValueError("Router response is not parseable JSON: " + "; ".join(errors[-3:]))
    if not isinstance(payload, dict):
        raise ValueError("Router response must be a JSON object.")

    needs = payload.get("needs") or payload.get("query_tags") or {}
    if not isinstance(needs, dict):
        raise ValueError("needs must be an object.")

    aliases = {
        "visual_style": ("visual_style", "visual_style_required"),
        "format_reference": ("format_reference",),
        "exact_coordinates": ("exact_coordinates", "coordinate_required"),
        "row_records": ("row_records", "row_record_operation"),
        "complex_headers": ("complex_headers", "complex_structure_required"),
        "sparse_or_large": ("sparse_or_large", "large_or_sparse_table"),
        "multi_sheet": ("multi_sheet",),
        "numeric_reasoning": ("numeric_reasoning",),
    }
    normalized_needs = {}
    for key, names in aliases.items():
        normalized_needs[key] = any(bool(needs.get(name, False)) for name in names)
    payload["needs"] = normalized_needs

    default_risk = str(payload.get("default_risk", "none")).strip().lower()
    if default_risk not in {"none", "low", "medium", "high"}:
        default_risk = "none"
    payload["default_risk"] = default_risk

    evidence = payload.get("evidence") or payload.get("reason") or []
    if isinstance(evidence, str):
        evidence = [evidence]
    if not isinstance(evidence, list):
        evidence = []
    payload["evidence"] = [str(item).strip() for item in evidence if str(item).strip()]
    return payload


def cell_to_json_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        text = value
    elif hasattr(value, "isoformat"):
        with suppress(Exception):
            return value.isoformat()
        text = str(value)
    else:
        text = str(value)
    if isinstance(text, str) and len(text) > 80:
        return text[:77] + "..."
    return text


def parse_a1_range(range_text: str) -> Tuple[int, int, int, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_text or "")
    if not match:
        return 1, 1, 1, 1
    col1, row1, col2, row2 = match.groups()
    return int(row1), column_index_from_string(col1), int(row2), column_index_from_string(col2)


def preview_workbook(
    xlsx_path: str,
    sheet_summaries: List[Dict[str, Any]],
    max_rows: int,
    max_cols: int,
) -> Dict[str, Any]:
    previews: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(xlsx_path):
        return previews

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    except Exception:
        return previews

    try:
        summary_by_name = {item.get("sheet_name"): item for item in sheet_summaries}
        for ws in wb.worksheets:
            summary = summary_by_name.get(ws.title) or {}
            min_row, min_col, max_row_idx, max_col_idx = parse_a1_range(summary.get("used_range", "A1:A1"))
            row_end = min(max_row_idx, min_row + max_rows - 1)
            col_end = min(max_col_idx, min_col + max_cols - 1)

            rows = []
            for row in ws.iter_rows(min_row=min_row, max_row=row_end, min_col=min_col, max_col=col_end):
                rows.append([cell_to_json_value(cell.value) for cell in row])

            merged_ranges = [str(item) for item in list(ws.merged_cells.ranges)[:20]]
            previews[ws.title] = {
                "header_preview": rows[0] if rows else [],
                "first_rows_preview": rows,
                "merged_ranges": merged_ranges,
            }
    finally:
        wb.close()
    return previews


def sheet_style_summary(sheet: Dict[str, Any]) -> List[str]:
    out = []
    if int(sheet.get("distinct_fill_colors") or 0) > 0:
        out.append("fill_color")
    if int(sheet.get("bold_cells") or 0) > 0:
        out.append("bold_font")
    if int(sheet.get("bordered_cells") or 0) > 0:
        out.append("border")
    if int(sheet.get("hidden_rows") or 0) > 0:
        out.append("hidden_rows")
    if int(sheet.get("hidden_cols") or 0) > 0:
        out.append("hidden_cols")
    if int(sheet.get("charts") or 0) > 0:
        out.append("chart")
    if int(sheet.get("embedded_images") or 0) > 0:
        out.append("embedded_image")
    if int(sheet.get("merged_ranges") or 0) > 0:
        out.append("merged_cells")
    return out


def build_available_formats(profile: Dict[str, Any]) -> List[str]:
    available = set(profile.get("available_text_formats") or []) | set(profile.get("available_image_formats") or [])
    return [fmt for fmt in SINGLE_FORMATS if fmt in available]


def estimate_format_tokens(profile: Dict[str, Any], available_formats: Sequence[str]) -> Dict[str, int]:
    text_chars = int(profile.get("total_text_chars") or 0)
    used_cells = int(profile.get("total_used_cells") or 0)
    nonempty = int(profile.get("total_nonempty_cells") or 0)
    rows = int(profile.get("max_rows") or 0) * max(1, int(profile.get("num_sheets") or 1))
    merged = int(profile.get("num_merged_ranges") or 0)

    char_estimates = {
        "csv": text_chars + used_cells * 2 + rows * 8,
        "markdown": text_chars + used_cells * 5 + rows * 12,
        "dataframe": text_chars + used_cells * 4 + rows * 12,
        "html": text_chars + used_cells * 18 + rows * 20,
        "json_rows": text_chars + used_cells * 8 + rows * 26 + merged * 30,
        "json_cells": text_chars + nonempty * 55 + merged * 30,
        "official_latex": text_chars + used_cells * 4 + merged * 35,
        "latex": text_chars + used_cells * 4 + merged * 35,
    }
    out = {}
    for fmt in available_formats:
        if fmt in TEXT_FORMATS:
            out[fmt] = estimate_tokens_from_chars(max(1, char_estimates.get(fmt, text_chars)))
    return out


def build_workbook_profile_for_router(
    profile: Dict[str, Any],
    available_formats: Sequence[str],
    preview_rows: int,
    preview_cols: int,
) -> Dict[str, Any]:
    sheet_summaries = profile.get("sheet_summaries") or []
    previews = preview_workbook(profile.get("xlsx_path", ""), sheet_summaries, preview_rows, preview_cols)

    sheets = []
    for item in sheet_summaries:
        preview = previews.get(item.get("sheet_name"), {})
        style_summary = sheet_style_summary(item)
        sheets.append(
            {
                "name": item.get("sheet_name"),
                "used_range": item.get("used_range"),
                "rows": item.get("used_rows"),
                "cols": item.get("used_cols"),
                "non_empty_cells": item.get("nonempty_cells"),
                "density": round(
                    (item.get("nonempty_cells") or 0) / max(1, item.get("used_cells") or 1),
                    4,
                ),
                "merged_ranges": preview.get("merged_ranges", []),
                "has_style_signal": bool(style_summary),
                "style_signal_summary": style_summary,
                "header_preview": preview.get("header_preview", []),
                "first_rows_preview": preview.get("first_rows_preview", []),
            }
        )

    return {
        "num_sheets": profile.get("num_sheets"),
        "sheets": sheets,
        "total_used_cells": profile.get("total_used_cells"),
        "total_non_empty_cells": profile.get("total_nonempty_cells"),
        "density": profile.get("nonempty_ratio"),
        "has_style_signal": bool(
            profile.get("has_background_color")
            or profile.get("has_charts_or_images")
            or int(profile.get("num_bold_cells") or 0) > 0
            or int(profile.get("num_bordered_cells") or 0) > 0
            or profile.get("has_hidden_rows_or_cols")
        ),
        "has_merged_cells": bool(profile.get("merged_cell_signal")),
        "max_rows": profile.get("max_rows"),
        "max_cols": profile.get("max_cols"),
        "estimated_tokens": estimate_format_tokens(profile, available_formats),
        "image_counts": profile.get("image_counts") or {},
    }


def query_likely_visual_style(query: str) -> bool:
    lowered = (query or "").lower()
    if not text_contains_any(query, VISUAL_TERMS):
        return False
    if text_contains_any(query, FORMAT_FALSE_POSITIVES) and not any(
        term in lowered
        for term in (
            "color",
            "colour",
            "background",
            "fill",
            "font",
            "border",
            "bold",
            "format reference",
            "formatting reference",
            "keep formatting",
            "preserve formatting",
        )
    ):
        return False
    return True


def will_truncate(fmt: str, routing_input: Dict[str, Any], margin: float = 0.7) -> bool:
    if fmt not in TEXT_FORMATS:
        return False
    budget = int(routing_input.get("token_budget") or 0)
    if budget <= 0:
        return False
    estimate = int((routing_input.get("workbook_profile") or {}).get("estimated_tokens", {}).get(fmt) or 0)
    return bool(estimate and estimate > budget * margin)


def all_text_formats_too_long(routing_input: Dict[str, Any], available: Set[str], margin: float = 0.7) -> bool:
    text_available = [fmt for fmt in available if fmt in TEXT_FORMATS]
    if not text_available:
        return True
    return all(will_truncate(fmt, routing_input, margin=margin) for fmt in text_available)


def compact_text_fallback(task_mode: str, routing_input: Dict[str, Any], available: Set[str]) -> Optional[str]:
    if task_mode == "qa_cot":
        order = ("json_cells", "latex", "csv", "json_rows", "markdown", "dataframe", "html")
    else:
        order = ("json_cells", "json_rows", "csv", "markdown", "latex", "dataframe", "html")
    for fmt in order:
        if fmt in available and not will_truncate(fmt, routing_input):
            return fmt
    return first_available(order, available)


def select_distinct_fallback(selected: str, task_mode: str, available: Set[str]) -> Optional[str]:
    for fmt in priority_for_task(task_mode):
        if fmt in available and fmt != selected:
            return fmt
    return None


def matches_any(text: str, patterns: Sequence[str]) -> bool:
    return any(re.search(pattern, text or "", flags=re.I) for pattern in patterns)


def program_need_tags(routing_input: Dict[str, Any]) -> Dict[str, bool]:
    query = routing_input.get("query", "")
    profile = routing_input.get("workbook_profile") or {}
    density = float(profile.get("density") or 0.0)
    max_rows = int(profile.get("max_rows") or 0)
    max_cols = int(profile.get("max_cols") or 0)
    task_mode = routing_input.get("task_mode", "qa_cot")
    default = default_for_task(task_mode)

    format_reference = text_contains_any(
        query,
        ("format reference", "formatting reference", "same format", "keep formatting", "preserve formatting"),
    )
    instruction_type = str(routing_input.get("instruction_type") or "")
    metadata_coordinate = bool(routing_input.get("answer_position") or routing_input.get("answer_sheet"))
    metadata_coordinate = metadata_coordinate and "Cell-Level" in instruction_type
    query_coordinate = matches_any(query, COORDINATE_PATTERNS)
    if routing_input.get("task_mode") == "manipulation_pot" and "Cell-Level" not in instruction_type:
        query_coordinate = False
    return {
        "visual_style": query_likely_visual_style(query),
        "format_reference": format_reference,
        "exact_coordinates": metadata_coordinate or query_coordinate,
        "row_records": text_contains_any(query, ROW_RECORD_TERMS),
        "complex_headers": text_contains_any(query, STRUCTURE_TERMS) or bool(profile.get("has_merged_cells")),
        "sparse_or_large": density < 0.35 or max_rows >= 80 or max_cols >= 12 or will_truncate(default, routing_input),
        "multi_sheet": int(profile.get("num_sheets") or 0) >= 2 or "sheet" in query.lower() or "worksheet" in query.lower(),
        "numeric_reasoning": text_contains_any(query, NUMERIC_TERMS),
    }


def merged_need_tags(llm_decision: Dict[str, Any], routing_input: Dict[str, Any]) -> Dict[str, bool]:
    llm_needs = llm_decision.get("needs") or {}
    program_needs = program_need_tags(routing_input)
    needs = {key: bool(llm_needs.get(key, False) or program_needs.get(key, False)) for key in program_needs}

    # Keep visual routing query-driven. Workbook styles alone are too common to imply image need.
    needs["visual_style"] = bool(
        (llm_needs.get("visual_style") and query_likely_visual_style(routing_input.get("query", "")))
        or program_needs["visual_style"]
        or needs["format_reference"]
    )
    return needs


def score_formats(needs: Dict[str, bool], routing_input: Dict[str, Any]) -> Dict[str, float]:
    task_mode = routing_input["task_mode"]
    dataset = routing_input["dataset"]
    profile = routing_input.get("workbook_profile") or {}
    scores = copy.deepcopy(QA_BASE_SCORES if task_mode == "qa_cot" else MANIPULATION_BASE_SCORES)

    if needs["visual_style"] or needs["format_reference"]:
        bonus = 5.0 if needs["format_reference"] else 4.0
        weights = {"image": 1.0, "excel_1_image": 0.8, "default_image": 0.4}
        if dataset == "spreadsheetbench":
            weights = {"excel_1_image": 1.0, "image": 0.8, "default_image": 0.4}
        for fmt, weight in weights.items():
            scores[fmt] += bonus * weight
        for fmt in TEXT_FORMATS:
            if fmt in scores:
                scores[fmt] -= 0.6

    if needs["exact_coordinates"]:
        scores["json_cells"] += 2.2
        scores["json_rows"] -= 0.4
    if needs["row_records"]:
        scores["json_rows"] += 2.0
        scores["markdown"] += 0.4
    if needs["complex_headers"]:
        if task_mode == "qa_cot":
            scores["latex"] += 1.3
        else:
            scores["latex"] += 0.2
        scores["json_cells"] += 1.0
        scores["markdown"] -= 0.4
    if needs["sparse_or_large"]:
        scores["json_cells"] += 1.4
        scores["json_rows"] -= 0.5
        scores["markdown"] -= 0.8
        scores["html"] -= 0.8
        scores["dataframe"] -= 0.5
    if needs["multi_sheet"]:
        scores["json_cells"] += 0.9
    if needs["numeric_reasoning"]:
        if task_mode == "qa_cot":
            scores["latex"] += 0.8
        scores["json_cells"] += 0.5

    dense = float(profile.get("density") or 0.0) >= 0.35
    if needs["row_records"] and dense and not needs["exact_coordinates"] and not needs["visual_style"]:
        scores["json_rows"] += 1.0
    if task_mode == "qa_cot" and int(profile.get("max_rows") or 0) <= 20 and int(profile.get("max_cols") or 0) <= 8:
        if not needs["complex_headers"] and not needs["exact_coordinates"]:
            scores["markdown"] += 0.8
    return scores


def select_format_from_needs(
    llm_decision: Dict[str, Any],
    routing_input: Dict[str, Any],
) -> Tuple[str, Dict[str, Any]]:
    available = set(routing_input.get("available_formats") or [])
    task_mode = routing_input["task_mode"]
    dataset = routing_input["dataset"]
    priority = priority_for_task(task_mode)

    if not available:
        raise ValueError("No single table format is available for routing.")

    needs = merged_need_tags(llm_decision, routing_input)
    scores = score_formats(needs, routing_input)
    rank = {fmt: i for i, fmt in enumerate(priority)}
    candidates = [fmt for fmt in priority if fmt in available]
    fmt = max(candidates, key=lambda name: (scores.get(name, -99.0), -rank.get(name, 99)))
    reasons = []

    if fmt in IMAGE_FORMATS:
        visual = bool(needs["visual_style"] or needs["format_reference"])
        too_long = all_text_formats_too_long(routing_input, available)
        if not visual and not too_long:
            fallback = default_for_task(task_mode) if default_for_task(task_mode) in available else None
            if fallback in available and fallback not in IMAGE_FORMATS:
                fmt = fallback
                reasons.append("image score suppressed because no strong visual/style need was detected")

    if fmt == "default_image":
        replacement = first_available(image_priority(dataset), available)
        if replacement:
            fmt = replacement
            if replacement != "default_image":
                reasons.append("default_image was remapped to a higher-priority available image source")

    if will_truncate(fmt, routing_input):
        compact = compact_text_fallback(task_mode, routing_input, available)
        if compact and compact != fmt:
            fmt = compact
            reasons.append("selected text format had truncation risk, so a compact text format was used")

    if fmt not in available:
        fmt = first_available(priority, available) or next(iter(available))
        reasons.append("final availability fallback was applied")

    default = default_for_task(task_mode)
    selected_score = scores.get(fmt, 0.0)
    default_score = scores.get(default, 0.0) if default in available else selected_score
    confidence = max(0.05, min(0.95, 0.55 + 0.12 * (selected_score - default_score)))

    return fmt, {
        "postprocess_reasons": reasons,
        "fallback_format": select_distinct_fallback(fmt, task_mode, available),
        "needs": needs,
        "program_needs": program_need_tags(routing_input),
        "scores": {key: round(value, 4) for key, value in sorted(scores.items()) if key in available},
        "selected_score": round(selected_score, 4),
        "default_score": round(default_score, 4),
        "confidence": round(confidence, 4),
    }


class SingleResponseLLMRouter:
    def __init__(self, args: argparse.Namespace):
        self.args = args

    def route(self, dataset: str, item: Dict[str, Any], profile: Dict[str, Any]) -> RouteDecision:
        task_mode = task_mode_for_dataset(dataset)
        available_formats = build_available_formats(profile)
        if dataset == "realhitbench":
            available_formats = [fmt for fmt in available_formats if fmt != "official_latex"]
        routing_input = self._build_routing_input(dataset, item, profile, available_formats)
        messages = self._build_messages(routing_input)
        raw_response = ""
        parsed_decision: Dict[str, Any]
        parse_error = None

        try:
            parsed_decision, raw_response = self._call_and_parse(messages, set(available_formats))
        except Exception as exc:
            if not self.args.router_fail_open:
                raise
            parse_error = str(exc)
            parsed_decision = self._default_router_decision(task_mode, set(available_formats), parse_error)

        selected_format, post = select_format_from_needs(parsed_decision, routing_input)

        reason_items = list(parsed_decision.get("evidence") or [])
        reason_items.extend(post.get("postprocess_reasons") or [])
        if not reason_items:
            reason_items = [f"Selected {selected_format} from LLM evidence tags and calibrated format scores."]

        return RouteDecision(
            solver_mode=solver_mode_for_dataset(dataset),
            table_format=selected_format,
            fallback_formats=[],
            reason="; ".join(reason_items),
            stages={
                "task_router": "qa" if dataset == "realhitbench" else "operation",
                "representation_router": selected_format,
                "reasoning_router": solver_mode_for_dataset(dataset),
                "single_response": True,
                "uses_llm_router": True,
                "router_model": self.router_model_name,
                "confidence": post.get("confidence"),
                "llm_needs": parsed_decision.get("needs"),
                "program_needs": post.get("program_needs"),
                "query_tags": post.get("needs"),
                "default_risk": parsed_decision.get("default_risk"),
                "scores": post.get("scores"),
                "selected_score": post.get("selected_score"),
                "default_score": post.get("default_score"),
                "llm_fallback_format": None,
                "postprocessed_format": selected_format,
                "postprocess": post,
                "router_error": parse_error,
                "raw_response": raw_response if self.args.save_router_responses else None,
                "routing_input": routing_input if self.args.save_router_prompts else None,
                "router_messages": messages if self.args.save_router_prompts else None,
            },
        )

    @property
    def router_model_name(self) -> Optional[str]:
        return self.args.router_model_name or self.args.model_name

    @property
    def router_url(self) -> Optional[str]:
        return self.args.router_base_url or self.args.router_url or self.args.base_url or self.args.url

    @property
    def router_api_key(self) -> Optional[str]:
        return self.args.router_api_key or self.args.api_key

    def _build_routing_input(
        self,
        dataset: str,
        item: Dict[str, Any],
        profile: Dict[str, Any],
        available_formats: Sequence[str],
    ) -> Dict[str, Any]:
        if dataset == "realhitbench":
            query = item.get("Question", "")
            instruction_type = None
            answer_position = None
            answer_sheet = None
        else:
            query = item.get("instruction", "")
            instruction_type = item.get("instruction_type")
            answer_position = item.get("answer_position")
            answer_sheet = item.get("answer_sheet")

        return {
            "dataset": dataset,
            "task_mode": task_mode_for_dataset(dataset),
            "query": query,
            "instruction_type": instruction_type,
            "answer_position": answer_position,
            "answer_sheet": answer_sheet,
            "available_formats": list(available_formats),
            "token_budget": int(self.args.max_text_tokens or 0),
            "workbook_profile": build_workbook_profile_for_router(
                profile,
                available_formats,
                preview_rows=self.args.router_preview_rows,
                preview_cols=self.args.router_preview_cols,
            ),
        }

    def _build_messages(self, routing_input: Dict[str, Any]) -> List[Dict[str, str]]:
        user_prompt = ROUTER_USER_PROMPT_TEMPLATE.format(
            routing_input_json=json.dumps(routing_input, ensure_ascii=False, indent=2),
        )
        return [
            {"role": "system", "content": ROUTER_SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ]

    def _call_and_parse(
        self,
        messages: List[Dict[str, Any]],
        available_formats: Set[str],
    ) -> Tuple[Dict[str, Any], str]:
        raw_response = ""
        current_messages = copy.deepcopy(messages)
        last_error = None

        for attempt in range(1, self.args.router_max_retries + 1):
            raw_response = chat_completion_content(
                url=self.router_url,
                api_key=self.router_api_key,
                messages=current_messages,
                model_name=self.router_model_name,
                temperature=self.args.router_temperature,
                top_p=self.args.router_top_p,
                max_tokens=self.args.router_max_tokens,
                timeout=self.args.router_timeout,
            )
            try:
                return parse_router_json(raw_response, available_formats), raw_response
            except Exception as exc:
                last_error = exc
                if attempt >= self.args.router_max_retries:
                    break
                current_messages.append({"role": "assistant", "content": raw_response})
                current_messages.append(
                    {
                        "role": "user",
                        "content": ROUTER_REPAIR_PROMPT_TEMPLATE.format(error=str(exc)),
                    }
                )

        raise ValueError(f"Router response validation failed: {last_error}")

    def _default_router_decision(self, task_mode: str, available: Set[str], error: str) -> Dict[str, Any]:
        return {
            "needs": {
                "visual_style": False,
                "format_reference": False,
                "exact_coordinates": False,
                "row_records": False,
                "complex_headers": False,
                "sparse_or_large": False,
                "multi_sheet": False,
                "numeric_reasoning": False,
            },
            "default_risk": "none",
            "evidence": [f"LLM router failed open; using program-side tags only. Error: {error}"],
        }


def build_router_suffix(args: argparse.Namespace) -> str:
    parts = ["single_resp_llm_router"]
    router_model = args.router_model_name or args.model_name
    if router_model:
        parts.append(f"rmodel_{safe_name(router_model)}")
    if args.fill_merged:
        parts.append("fillmerged")
    if not args.include_coordinates:
        parts.append("nocoord")
    if args.max_text_tokens:
        parts.append(f"{int(args.max_text_tokens / 1000)}ktoken")
    if args.router_top_p != 1.0 or args.router_temperature != 0:
        parts.append(f"rtp{args.router_top_p}_rtemp{args.router_temperature}")
    if args.top_p != 1.0 or args.temperature != 0:
        parts.append(f"tp{args.top_p}_temp{args.temperature}")
    if args.suffix:
        parts.append(args.suffix)
    return "_".join(parts)


def build_dataset_dir(args: argparse.Namespace) -> str:
    if args.dataset == "realhitbench":
        return "realhitbench"
    return f"spreadsheetbench_{args.data_split}".replace("/", "_").replace("\\", "_")


def build_output_dir(args: argparse.Namespace) -> str:
    return os.path.join(args.output_root, build_dataset_dir(args), safe_name(args.model_name), build_router_suffix(args))


def make_route_args(args: argparse.Namespace, table_format: str) -> argparse.Namespace:
    route_args = argparse.Namespace(**vars(args))
    route_args.table_format = table_format
    route_args.dry_run = False
    return route_args


def build_solver(dataset: str, args: argparse.Namespace, out_dir: str):
    if dataset == "realhitbench":
        return RealHiTCoTSolver(**vars(args))
    return SpreadSheetPoTSolver(**vars(args), output_dir=out_dir)


def attach_router_metadata(
    result: Dict[str, Any],
    decision: RouteDecision,
    profile: Dict[str, Any],
) -> Dict[str, Any]:
    result = copy.deepcopy(result)
    result["router_policy"] = "single_resp_llm"
    result["router_decision"] = asdict(decision)
    result["router_profile"] = profile
    result["router_fallback_used"] = False
    result["router_attempts"] = [
        {
            "attempt_index": 1,
            "is_primary": True,
            "table_format": decision.table_format,
            "summary": {
                "table_format": decision.table_format,
                "format_valid": bool(result.get("format_valid")),
                "error": result.get("error"),
                "table_metadata": result.get("table_metadata"),
            },
        }
    ]
    return result


def solve_one(
    dataset: str,
    args: argparse.Namespace,
    item: Dict[str, Any],
    out_dir: str,
    profiler: SpreadsheetProfiler,
    router: SingleResponseLLMRouter,
) -> Dict[str, Any]:
    profile = profiler.profile(dataset, item)
    decision = router.route(dataset, item, profile)
    try:
        result = run_solver(dataset, args, item, decision.table_format, out_dir)
    except Exception:
        result = build_error_result(dataset, item, traceback.format_exc(), args)
    return attach_router_metadata(result, decision, profile)


def add_scores(dataset: str, score_lists: Dict[str, Any], result: Dict[str, Any]) -> None:
    if dataset == "realhitbench":
        realhit_entry.add_scores(score_lists, result)
    else:
        spreadsheet_entry.add_scores(score_lists, result)


def average_scores(dataset: str, score_lists: Dict[str, Any]) -> Dict[str, Any]:
    if dataset == "realhitbench":
        return realhit_entry.average_scores(score_lists)
    return spreadsheet_entry.average_scores(score_lists)


def save_outputs(
    dataset: str,
    out_dir: str,
    outs: List[Optional[Dict[str, Any]]],
    eval_results: List[Any],
    scores: Dict[str, Any],
    partial: bool = False,
) -> None:
    out_name, eval_name, score_name = output_filenames(dataset, partial=partial)
    save_jsonl(outs, os.path.join(out_dir, out_name))
    save_jsonl(eval_results, os.path.join(out_dir, eval_name))
    save_jsonl(scores, os.path.join(out_dir, score_name))

    if not partial:
        decisions = [
            {
                "id": result.get("id") if result else None,
                "router_decision": result.get("router_decision") if result else None,
                "router_fallback_used": result.get("router_fallback_used") if result else None,
            }
            for result in outs
        ]
        save_jsonl(decisions, os.path.join(out_dir, "router_decisions.jsonl"))


def dry_run_payload(
    dataset: str,
    prompt: Any,
    metadata: Dict[str, Any],
    profile: Dict[str, Any],
    decision: RouteDecision,
) -> Dict[str, Any]:
    if dataset == "realhitbench":
        payload = realhit_entry.dry_run_payload(prompt, metadata)
    else:
        payload = spreadsheet_entry.dry_run_payload(prompt, metadata)
    payload.update({"router_profile": profile, "router_decision": asdict(decision)})
    return payload


def get_dataset(args: argparse.Namespace, out_dir: str) -> List[Dict[str, Any]]:
    if args.dataset == "realhitbench":
        return realhit_entry.get_dataset(args, output_dir=out_dir)
    spreadsheet_out_dir = os.path.join(out_dir, "spreadsheet")
    if os.path.exists(spreadsheet_out_dir) and not args.resume and not args.dry_run:
        shutil.rmtree(spreadsheet_out_dir)
    os.makedirs(spreadsheet_out_dir, exist_ok=True)
    os.chmod(spreadsheet_out_dir, 0o777)
    return spreadsheet_entry.get_dataset(args, spreadsheet_out_dir)


def load_resume_state(dataset: str, args: argparse.Namespace, out_dir: str, data: List[Dict[str, Any]]):
    outs, eval_results = [None] * len(data), [None] * len(data)
    score_lists = defaultdict(lambda: defaultdict(list)) if dataset == "realhitbench" else defaultdict(list)
    partial_out_path = os.path.join(out_dir, output_filenames(dataset, partial=True)[0])
    partial_eval_path = os.path.join(out_dir, output_filenames(dataset, partial=True)[1])

    if not args.resume:
        return outs, eval_results, score_lists

    if not os.path.exists(partial_out_path):
        print(f"Resume enabled, but no partial result found in {out_dir}; starting from scratch.")
        return outs, eval_results, score_lists

    loaded_outs = load_jsonl(partial_out_path)
    loaded_eval_results = load_jsonl(partial_eval_path) if os.path.exists(partial_eval_path) else []
    resumed_count = 0
    for idx, result in enumerate(loaded_outs[: len(data)]):
        if result is None:
            continue
        if str(result.get("id")) != str(data[idx].get("id")):
            continue
        outs[idx] = result
        eval_results[idx] = loaded_eval_results[idx] if idx < len(loaded_eval_results) else build_eval_entry(dataset, result)
        add_scores(dataset, score_lists, result)
        resumed_count += 1

    print(f"Resume loaded {resumed_count} completed results from {partial_out_path}")
    return outs, eval_results, score_lists


def solution(args: argparse.Namespace) -> None:
    if args.base_url:
        os.environ["BASE_URL"] = args.base_url
    if args.api_key:
        os.environ["API_KEY"] = args.api_key

    out_dir = build_output_dir(args)
    os.makedirs(out_dir, exist_ok=True)
    data = get_dataset(args, out_dir)
    profiler = SpreadsheetProfiler(args)
    router = SingleResponseLLMRouter(args)

    if args.dry_run:
        if not data:
            print("No data selected.")
            return
        item = data[min(args.dry_run_index, len(data) - 1)]
        profile = profiler.profile(args.dataset, item)
        args.save_router_prompts = True
        args.save_router_responses = True
        decision = router.route(args.dataset, item, profile)
        route_args = make_route_args(args, decision.table_format)
        solver = build_solver(args.dataset, route_args, out_dir)
        prompt, metadata = solver.build_prompt(item)
        path = os.path.join(out_dir, "dry_run_prompt.json")
        save_jsonl(dry_run_payload(args.dataset, prompt, metadata, profile, decision), path)
        print(f"Dry-run prompt saved to {path}")
        return

    outs, eval_results, score_lists = load_resume_state(args.dataset, args, out_dir, data)
    pending_indices = [idx for idx, result in enumerate(outs) if result is None]

    if not pending_indices:
        scores = average_scores(args.dataset, score_lists)
        save_outputs(args.dataset, out_dir, outs, eval_results, scores, partial=False)
        cleanup_intermediate_files(args.dataset, out_dir)
        print("No pending examples. Final files have been saved.")
        return

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(solve_one, args.dataset, args, data[idx], out_dir, profiler, router): idx
            for idx in pending_indices
        }
        for solved_count, future in tqdm(
            enumerate(as_completed(futures), start=1),
            total=len(pending_indices),
            desc=f"Solving SingleRespLLMRouter-{args.dataset}",
        ):
            idx = futures[future]
            try:
                result = future.result()
            except Exception:
                result = build_error_result(args.dataset, data[idx], traceback.format_exc(), args)
            outs[idx] = result
            eval_results[idx] = build_eval_entry(args.dataset, result)
            add_scores(args.dataset, score_lists, result)

            if args.save_every and solved_count % args.save_every == 0:
                scores = average_scores(args.dataset, score_lists)
                save_outputs(args.dataset, out_dir, outs, eval_results, scores, partial=True)

            if solved_count % args.report_every == 0 or solved_count == len(pending_indices):
                report_scores(args.dataset, score_lists)

    scores = average_scores(args.dataset, score_lists)
    save_outputs(args.dataset, out_dir, outs, eval_results, scores, partial=False)
    cleanup_intermediate_files(args.dataset, out_dir)
    print(f"Saved SingleRespLLMRouter outputs to {out_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Single-response LLM table representation router for RealHiTBench and SpreadsheetBench."
    )

    parser.add_argument("--dataset", type=str, required=True, choices=["realhitbench", "spreadsheetbench"])
    parser.add_argument("--url", type=str, default=None)
    parser.add_argument("--base_url", type=str, default=None, help="Optional solver OpenAI-compatible base URL.")
    parser.add_argument("--api_key", type=str, default=None, help="Optional solver API key. Also readable from API_KEY/OPENAI_API_KEY.")
    parser.add_argument("--model_name", type=str, default=None)
    parser.add_argument("--code_exec_url", type=str, default="localhost:8081")
    parser.add_argument("--top_p", type=float, default=1.0)
    parser.add_argument("--temperature", type=float, default=0)

    parser.add_argument("--router_url", type=str, default=None, help="Optional router API URL. Defaults to solver URL/base URL.")
    parser.add_argument("--router_base_url", type=str, default=None, help="Optional router OpenAI-compatible base URL.")
    parser.add_argument("--router_api_key", type=str, default=None, help="Optional router API key.")
    parser.add_argument("--router_model_name", type=str, default=None, help="Optional model used only for routing decisions.")
    parser.add_argument("--router_temperature", type=float, default=0)
    parser.add_argument("--router_top_p", type=float, default=1.0)
    parser.add_argument("--router_max_tokens", type=int, default=1024)
    parser.add_argument("--router_max_retries", type=int, default=2)
    parser.add_argument("--router_timeout", type=int, default=120)
    parser.add_argument("--router_preview_rows", type=int, default=4)
    parser.add_argument("--router_preview_cols", type=int, default=12)
    parser.add_argument("--router_fail_open", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--save_router_prompts", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--save_router_responses", action=argparse.BooleanOptionalAction, default=True)

    parser.add_argument("--include_coordinates", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--fill_merged", action="store_true")
    parser.add_argument(
        "--max_text_tokens",
        type=int,
        default=100000,
        help="Prompt table-text token budget used both for profiling and selected solver prompts.",
    )
    parser.add_argument("--render_formulas_before_eval", action="store_true")

    parser.add_argument("--max_retries", type=int, default=3)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--ids", type=str, default=None)
    parser.add_argument("--question_types", type=str, default=None)
    parser.add_argument("--instruction_types", type=str, default=None)
    parser.add_argument("--data_split", type=str, default="all_912", choices=sorted(SPREADSHEET_DATA_SPLITS))
    parser.add_argument("--report_every", type=int, default=10)
    parser.add_argument("--save_every", type=int, default=10)
    parser.add_argument("--save_prompts", action="store_true")
    parser.add_argument("--dry_run", action="store_true")
    parser.add_argument("--dry_run_index", type=int, default=0)
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("-s", "--suffix", type=str, default=None)
    parser.add_argument("--output_root", type=str, default=os.path.join(repo_dir, "outs"))

    return parser.parse_args()


if __name__ == "__main__":
    warnings.filterwarnings("ignore")
    solution(parse_args())
