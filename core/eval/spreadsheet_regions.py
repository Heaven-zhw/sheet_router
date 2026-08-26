"""SpreadsheetBench answer-region parsing and value extraction utilities."""

import datetime
import os
import re
from collections import OrderedDict
from collections.abc import Mapping
from dataclasses import dataclass
from typing import Any, Iterator, Optional, Tuple, Union

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


Bound = Optional[Union[int, Mapping[str, int]]]
CellKey = Tuple[str, str]
_MAX_EXCEL_ROW = 1_048_576
_MAX_EXCEL_COLUMN = 16_384


class SpreadsheetRegionError(ValueError):
    """Raised when an answer region cannot be resolved safely."""


class WorkbookRegionError(SpreadsheetRegionError):
    """Raised when a workbook cannot be opened for region extraction."""


@dataclass(frozen=True)
class CellRegion:
    sheet_name: str
    min_row: int
    min_column: int
    max_row: int
    max_column: int
    source: str

    def iter_coordinates(self) -> Iterator[CellKey]:
        for row in range(self.min_row, self.max_row + 1):
            for column in range(self.min_column, self.max_column + 1):
                yield self.sheet_name, f"{get_column_letter(column)}{row}"


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """Preserve the legacy SpreadsheetBench value-normalization semantics."""
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v.strip() if isinstance(v, str) else v


def compare_cell_value(v1, v2):
    """Preserve the legacy SpreadsheetBench cell-comparison semantics."""
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    if v1 == v2:
        return True
    else:
        return False


_CELL_RANGE_RE = re.compile(
    r"^\$?([A-Za-z]+)(?:\$?(\d+))?"
    r"(?:\s*:\s*\$?([A-Za-z]+)(?:\$?(\d+))?)?$"
)
_ROW_RANGE_RE = re.compile(r"^\$?(\d+)\s*:\s*\$?(\d+)$")


def normalize_sheet_name(sheet_name: str) -> str:
    """Remove whitespace and any unmatched surrounding Excel quote marks."""
    normalized = str(sheet_name).strip().strip("'\"").strip()
    if not normalized:
        raise SpreadsheetRegionError(f"Invalid empty worksheet name: {sheet_name!r}")
    return normalized


def _normalize_range_text(range_text: str) -> str:
    return str(range_text).strip().strip("'\"").strip()


def _looks_like_range(source: str) -> bool:
    range_text = source.rsplit("!", 1)[-1]
    range_text = _normalize_range_text(range_text)
    return bool(
        _CELL_RANGE_RE.fullmatch(range_text) or _ROW_RANGE_RE.fullmatch(range_text)
    )


def split_answer_position(answer_position: str) -> list[str]:
    """Split region commas without splitting commas inside a worksheet name.

    The benchmark contains unmatched quotes, so quote-balance alone cannot
    identify delimiters. A comma terminates a region only after a syntactically
    complete range has been seen.
    """
    regions = []
    start = 0
    for index, char in enumerate(answer_position):
        if char != ",":
            continue
        candidate = answer_position[start:index].strip()
        if _looks_like_range(candidate):
            regions.append(candidate)
            start = index + 1
    regions.append(answer_position[start:].strip())
    return regions


def _answer_sheet_candidates(answer_sheet: Optional[str]) -> list[str]:
    if answer_sheet is None:
        return []
    candidates = []
    for raw_name in str(answer_sheet).split(","):
        if raw_name.strip():
            candidates.append(normalize_sheet_name(raw_name))
    return candidates


def resolve_default_sheet(workbook, answer_sheet: Optional[str] = "") -> str:
    """Resolve the first existing answer_sheet, then fall back to sheet one."""
    if answer_sheet is not None and str(answer_sheet).strip():
        whole_name = normalize_sheet_name(str(answer_sheet))
        if whole_name in workbook.sheetnames:
            return whole_name
    for sheet_name in _answer_sheet_candidates(answer_sheet):
        if sheet_name in workbook.sheetnames:
            return sheet_name
    if workbook.sheetnames:
        return workbook.sheetnames[0]
    raise SpreadsheetRegionError("Workbook contains no worksheets")


def _resolve_bound(bound: Bound, sheet_name: str, default: int, label: str) -> int:
    if isinstance(bound, Mapping):
        value = bound.get(sheet_name, default)
    elif bound is None:
        value = default
    else:
        value = bound
    maximum = _MAX_EXCEL_ROW if label == "row" else _MAX_EXCEL_COLUMN
    if (
        isinstance(value, bool)
        or not isinstance(value, int)
        or value < 1
        or value > maximum
    ):
        raise SpreadsheetRegionError(
            f"Invalid {label} boundary for worksheet {sheet_name!r}: {value!r}"
        )
    return value


def _parse_range(
    range_text: str,
    sheet_name: str,
    max_row: int,
    max_column: int,
    source: str,
) -> CellRegion:
    row_match = _ROW_RANGE_RE.fullmatch(range_text)
    if row_match:
        min_row, end_row = (int(value) for value in row_match.groups())
        if min_row < 1 or end_row > _MAX_EXCEL_ROW or min_row > end_row:
            raise SpreadsheetRegionError(f"Invalid row range: {source!r}")
        return CellRegion(sheet_name, min_row, 1, end_row, max_column, source)

    match = _CELL_RANGE_RE.fullmatch(range_text)
    if not match:
        raise SpreadsheetRegionError(f"Invalid cell range: {source!r}")

    start_column_text, start_row_text, end_column_text, end_row_text = match.groups()
    try:
        start_column = column_index_from_string(start_column_text.upper())
    except ValueError as exc:
        raise SpreadsheetRegionError(f"Invalid column in range: {source!r}") from exc
    start_row = int(start_row_text) if start_row_text is not None else None

    if end_column_text is None:
        if start_row is None:
            raise SpreadsheetRegionError(f"Single column is not a finite cell range: {source!r}")
        end_column = start_column
        end_row = start_row
    else:
        try:
            end_column = column_index_from_string(end_column_text.upper())
        except ValueError as exc:
            raise SpreadsheetRegionError(f"Invalid column in range: {source!r}") from exc
        end_row = int(end_row_text) if end_row_text is not None else None
        if (start_row is None) != (end_row is None):
            raise SpreadsheetRegionError(f"Mixed row and whole-column range: {source!r}")
        if start_row is None:
            start_row = 1
            end_row = max_row

    if (
        start_column < 1
        or end_column > _MAX_EXCEL_COLUMN
        or start_row < 1
        or end_row > _MAX_EXCEL_ROW
    ):
        raise SpreadsheetRegionError(f"Cell range is outside Excel limits: {source!r}")
    if start_column > end_column or start_row > end_row:
        raise SpreadsheetRegionError(f"Reversed cell range in {source!r}")
    return CellRegion(sheet_name, start_row, start_column, end_row, end_column, source)


def parse_answer_regions(
    workbook,
    answer_position: Optional[str],
    answer_sheet: Optional[str] = "",
    *,
    max_rows: Bound = None,
    max_columns: Bound = None,
) -> list[CellRegion]:
    """Parse all answer_position entries against a workbook.

    ``max_rows`` and ``max_columns`` may be a single positive integer or a
    worksheet-name mapping. They make open-ended column/row ranges finite.
    """
    if answer_position is None or not str(answer_position).strip():
        raise SpreadsheetRegionError("answer_position is empty")

    default_sheet = resolve_default_sheet(workbook, answer_sheet)
    regions = []
    for raw_region in split_answer_position(str(answer_position)):
        source = raw_region.strip()
        if not source:
            raise SpreadsheetRegionError(
                f"Invalid empty region in answer_position: {answer_position!r}"
            )
        if "!" in source:
            raw_sheet_name, range_text = source.rsplit("!", 1)
            sheet_name = normalize_sheet_name(raw_sheet_name)
        else:
            sheet_name = default_sheet
            range_text = source

        if sheet_name not in workbook.sheetnames:
            raise SpreadsheetRegionError(
                f"Worksheet not found for region {source!r}: {sheet_name!r}"
            )
        range_text = _normalize_range_text(range_text)
        if not range_text:
            raise SpreadsheetRegionError(f"Missing cell range in region: {source!r}")

        worksheet = workbook[sheet_name]
        row_boundary = _resolve_bound(max_rows, sheet_name, worksheet.max_row, "row")
        column_boundary = _resolve_bound(
            max_columns, sheet_name, worksheet.max_column, "column"
        )
        regions.append(
            _parse_range(range_text, sheet_name, row_boundary, column_boundary, source)
        )
    return regions


def iter_unique_region_coordinates(regions: list[CellRegion]) -> Iterator[CellKey]:
    """Yield coordinates in region order and row-major order, without overlap."""
    seen = set()
    for region in regions:
        for key in region.iter_coordinates():
            if key not in seen:
                seen.add(key)
                yield key


def extract_normalized_region_cells(
    workbook_or_path,
    answer_position: Optional[str],
    answer_sheet: Optional[str] = "",
    *,
    max_rows: Bound = None,
    max_columns: Bound = None,
    data_only: bool = True,
) -> "OrderedDict[CellKey, Any]":
    """Extract an ordered normalized cell map without requiring a golden file.

    A path or an already-open workbook may be supplied. Workbooks opened by
    this function are always closed; caller-owned workbook objects stay open.
    """
    workbook = workbook_or_path
    should_close = False
    if isinstance(workbook_or_path, (str, os.PathLike)):
        path = os.fspath(workbook_or_path)
        if not os.path.isfile(path):
            raise WorkbookRegionError(f"Workbook file not found: {path}")
        try:
            workbook = openpyxl.load_workbook(path, data_only=data_only)
        except Exception as exc:
            raise WorkbookRegionError(f"Failed to open workbook {path}: {exc}") from exc
        should_close = True

    try:
        if not hasattr(workbook, "sheetnames"):
            raise WorkbookRegionError(
                f"Expected a workbook or workbook path, got {type(workbook).__name__}"
            )
        regions = parse_answer_regions(
            workbook,
            answer_position,
            answer_sheet,
            max_rows=max_rows,
            max_columns=max_columns,
        )
        cells = OrderedDict()
        for sheet_name, coordinate in iter_unique_region_coordinates(regions):
            cells[(sheet_name, coordinate)] = transform_value(
                workbook[sheet_name][coordinate].value
            )
        return cells
    finally:
        if should_close:
            workbook.close()
