"""SpreadsheetBench workbook evaluator."""

import os

import openpyxl

from .spreadsheet_regions import (
    SpreadsheetRegionError,
    compare_cell_value,
    datetime_to_float,
    iter_unique_region_coordinates,
    parse_answer_regions,
    transform_value,
)


def compare_fill_color(fill1, fill2):
    fg_color1 = fill1.fgColor.rgb if fill1.fgColor else None
    fg_color2 = fill2.fgColor.rgb if fill2.fgColor else None
    bg_color1 = fill1.bgColor.rgb if fill1.bgColor else None
    bg_color2 = fill2.bgColor.rgb if fill2.bgColor else None
    return fg_color1 == fg_color2 and bg_color1 == bg_color2


def compare_font_color(font_gt, font_proc):
    if font_gt.color is not None and font_proc.color is not None:
        return font_gt.color.rgb == font_proc.color.rgb
    return font_gt.color is None and font_proc.color is None


def _paired_max_rows(wb_gt, wb_proc):
    return {
        sheet_name: max(
            wb_gt[sheet_name].max_row,
            wb_proc[sheet_name].max_row if sheet_name in wb_proc.sheetnames else 1,
        )
        for sheet_name in wb_gt.sheetnames
    }


def _paired_max_columns(wb_gt, wb_proc):
    return {
        sheet_name: max(
            wb_gt[sheet_name].max_column,
            wb_proc[sheet_name].max_column if sheet_name in wb_proc.sheetnames else 1,
        )
        for sheet_name in wb_gt.sheetnames
    }


def _compare_target_cells(wb_gt, wb_proc, regions, is_cf):
    compared = 0
    for sheet_name, coordinate in iter_unique_region_coordinates(regions):
        if sheet_name not in wb_proc.sheetnames:
            return False, f"Worksheet not found in generated workbook: {sheet_name!r}"

        cell_gt = wb_gt[sheet_name][coordinate]
        cell_proc = wb_proc[sheet_name][coordinate]
        if not compare_cell_value(cell_gt.value, cell_proc.value):
            return False, (
                f"Value difference at {sheet_name!r}!{coordinate}: "
                f"gold has {cell_gt.value!r}, generated has {cell_proc.value!r}"
            )

        if is_cf and not compare_fill_color(cell_gt.fill, cell_proc.fill):
            return False, f"Fill color difference at {sheet_name!r}!{coordinate}"
        if is_cf and not compare_font_color(cell_gt.font, cell_proc.font):
            return False, f"Font color difference at {sheet_name!r}!{coordinate}"
        compared += 1

    return True, f"Success ({compared} unique cells across {len(regions)} regions)"


def compare_workbooks(
    gt_file,
    proc_file,
    instruction_type,
    answer_position,
    answer_sheet="",
):
    """Compare every unique target cell in every answer_position region."""
    del instruction_type  # Retained for backward-compatible call sites.
    if not os.path.isfile(gt_file):
        return False, f"Ground truth file not found: {gt_file}"
    if not os.path.isfile(proc_file):
        return False, f"Generated file not found: {proc_file}"

    wb_gt = None
    wb_proc = None
    try:
        try:
            wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        except Exception as exc:
            return False, f"Failed to open ground truth workbook {gt_file}: {exc}"
        try:
            wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
        except Exception as exc:
            return False, f"Failed to open generated workbook {proc_file}: {exc}"

        try:
            regions = parse_answer_regions(
                wb_gt,
                answer_position,
                answer_sheet,
                max_rows=_paired_max_rows(wb_gt, wb_proc),
                max_columns=_paired_max_columns(wb_gt, wb_proc),
            )
        except SpreadsheetRegionError as exc:
            return False, f"Invalid answer region: {exc}"

        try:
            return _compare_target_cells(wb_gt, wb_proc, regions, "CF" in str(proc_file))
        except Exception as exc:
            return False, f"Failed to compare answer regions: {exc}"
    finally:
        if wb_proc is not None:
            wb_proc.close()
        if wb_gt is not None:
            wb_gt.close()


def run_solution_one_data(data, client, runs_name):
    """Legacy three-case evaluation helper."""
    data["test_case_results"] = []
    data["test_case_messages"] = []

    for idx in range(1, 4):
        try:
            input_file = f"{idx}_{data['id']}_input.xlsx"
            output_file = f"{idx}_{data['id']}_output.xlsx"
            if data["solution"]:
                local_solution = data["solution"].replace(
                    f"1_{data['id']}_input.xlsx", input_file
                )
                local_solution = local_solution.replace(
                    f"1_{data['id']}_output.xlsx", output_file
                )
                client.execute(local_solution)

            gt_file = os.path.join(
                data["real_dir"], f"{idx}_{data['id']}_answer.xlsx"
            )
            output_file = f"outs/{runs_name}/spreadsheet/{idx}_{data['id']}_output.xlsx"
            passed, message = compare_workbooks(
                gt_file,
                output_file,
                data["instruction_type"],
                data["answer_position"],
                data.get("answer_sheet", ""),
            )
        except Exception as exc:
            passed, message = False, str(exc)

        data["test_case_results"].append(int(passed))
        data["test_case_messages"].append(message)

    data["total_soft_restriction"] = sum(data["test_case_results"]) / len(
        data["test_case_results"]
    )
    data["total_hard_restriction"] = 1.0 if all(data["test_case_results"]) else 0.0
    return data
