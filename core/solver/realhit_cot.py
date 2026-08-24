import base64
import copy
import csv
import glob
import html
import io
import json
import os
import re
import threading
import traceback
import mimetypes
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

import openpyxl
from openpyxl.utils import get_column_letter
import tiktoken

from ..excel2image_linux import render_excel_range_to_png
from ..utils import model_resp, read_text_with_encoding_fallback
from .metrics.qa_metrics import QAMetric
from .realhit_cot_prompts import Answer_Prompt


FINAL_ANSWER_PREFIX = "[Final Answer]:"

SUPPORTED_TABLE_FORMATS = (
    "latex",
    "csv",
    "tsv",
    "markdown",
    "dataframe",
    "json_rows",
    "json_cells",
    "html",
    "image",
    "excel_1_image",
    "default_image",
    "html+image",
    "markdown+image",
    "latex+image",
    "html+excel_1_image",
    "latex+excel_1_image",
    "markdown+excel_1_image",
    "html+default_image",
    "latex+default_image",
    "markdown+default_image",
)

IMAGE_TABLE_FORMATS = {"image", "excel_1_image", "default_image"}
TEXT_IMAGE_TABLE_FORMATS = {
    "html+image": ("html", "image"),
    "markdown+image": ("markdown", "image"),
    "latex+image": ("latex", "image"),
    "html+excel_1_image": ("html", "excel_1_image"),
    "latex+excel_1_image": ("latex", "excel_1_image"),
    "markdown+excel_1_image": ("markdown", "excel_1_image"),
    "html+default_image": ("html", "default_image"),
    "latex+default_image": ("latex", "default_image"),
    "markdown+default_image": ("markdown", "default_image"),
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
_DEFAULT_IMAGE_RENDER_LOCK = threading.Lock()

MODEL_ENCODINGS = {
    "gpt-4": "cl100k_base",
    "gpt-4-turbo": "cl100k_base",
    "gpt-4o": "o200k_base",
    "gpt-3.5-turbo": "cl100k_base",
    "gpt-5-nano-2025-08-07": "o200k_base",
    "text-embedding-ada-002": "cl100k_base",
    "qwen": "cl100k_base",
    "qwen3": "cl100k_base",
    "qwen3-vl": "cl100k_base",
    "qwen3.5": "cl100k_base",
    "qwen3.5-vl": "cl100k_base",
    "qwen35": "cl100k_base",
    "qwen35-vl": "cl100k_base",
}


USER_PROMPT_TEMPLATE = """You are a spreadsheet expert.

You need to answer the given spreadsheet question using only the provided spreadsheet content. The content may be provided as text, images, or text together with images. When images are attached, they are screenshots of worksheets from the same workbook in worksheet order. Do not call tools, do not write Python code, and do not invent data that is not present.

# Spreadsheet Content
{spreadsheet_content_note}

{table_text}

# Question
{question}

# Output Constraints For `final_answer`
{answer_constraints}

# Required JSON Output
Return only one valid JSON object, with no Markdown fences and no extra text before or after it.

The JSON schema is:
{{
  "reasoning": ["brief step 1", "brief step 2", "..."],
  "final_answer": "{final_answer_prefix} ..."
}}

Rules:
- `reasoning` must contain a concise reasoning summary with enough steps to justify the answer.
- `final_answer` must follow the output constraints above exactly.
- The JSON must be parseable by Python `json.loads`.
"""


REPAIR_PROMPT_TEMPLATE = """Your previous response did not satisfy the required output format.

Format error:
{error}

Regenerate the answer for the same question. Return only one valid JSON object with the schema:
{{
  "reasoning": ["brief step 1", "brief step 2", "..."],
  "final_answer": "{final_answer_prefix} ..."
}}

Do not use Markdown fences. Do not add extra text outside the JSON object.
"""


@dataclass
class ParsedResponse:
    payload: Dict[str, Any]
    final_answer_line: str
    final_answer: str


class ResponseFormatError(ValueError):
    pass


def get_final_answer(response: str) -> str:
    if FINAL_ANSWER_PREFIX in response:
        return response.split(FINAL_ANSWER_PREFIX, 1)[-1].strip()
    if "Final Answer:" in response:
        return response.split("Final Answer:", 1)[-1].strip()
    return str(response).strip()


def _strip_markdown_fence(text: str) -> str:
    text = text.strip()
    fence_match = re.fullmatch(r"```(?:json)?\s*(.*?)\s*```", text, flags=re.S | re.I)
    if fence_match:
        return fence_match.group(1).strip()
    return text


def _remove_thinking_blocks(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text, flags=re.S | re.I).strip()


def parse_and_validate_response(text: str) -> ParsedResponse:
    if text is None:
        raise ResponseFormatError("The model returned no content.")

    cleaned = _strip_markdown_fence(_remove_thinking_blocks(str(text)))
    payload = None
    json_errors = []

    for candidate in _json_candidates(cleaned):
        try:
            payload = json.loads(candidate)
            break
        except Exception as exc:
            json_errors.append(str(exc))

    if payload is None:
        joined = "; ".join(json_errors[-3:]) if json_errors else "no JSON object found"
        raise ResponseFormatError(f"Response is not parseable JSON: {joined}")

    if not isinstance(payload, dict):
        raise ResponseFormatError("Top-level JSON value must be an object.")

    reasoning = payload.get("reasoning")
    if isinstance(reasoning, str):
        if not reasoning.strip():
            raise ResponseFormatError("`reasoning` must not be empty.")
    elif isinstance(reasoning, list):
        if not reasoning or not all(isinstance(item, str) and item.strip() for item in reasoning):
            raise ResponseFormatError("`reasoning` must be a non-empty list of non-empty strings.")
    else:
        raise ResponseFormatError("`reasoning` must be a string or a list of strings.")

    final_answer_line = payload.get("final_answer")
    if not isinstance(final_answer_line, str):
        raise ResponseFormatError("`final_answer` must be a string.")

    final_answer_line = final_answer_line.strip()
    if "\n" in final_answer_line or "\r" in final_answer_line:
        raise ResponseFormatError("`final_answer` must be one line.")
    if not final_answer_line.startswith(FINAL_ANSWER_PREFIX):
        raise ResponseFormatError(f"`final_answer` must start with `{FINAL_ANSWER_PREFIX}`.")

    final_answer = get_final_answer(final_answer_line)
    if not final_answer:
        raise ResponseFormatError("`final_answer` is empty after the prefix.")

    return ParsedResponse(
        payload=payload,
        final_answer_line=final_answer_line,
        final_answer=final_answer,
    )


def _json_candidates(text: str) -> Sequence[str]:
    candidates = [text]

    json_fence = re.search(r"```json\s*(.*?)\s*```", text, flags=re.S | re.I)
    if json_fence:
        candidates.append(json_fence.group(1).strip())

    generic_fence = re.search(r"```\s*(.*?)\s*```", text, flags=re.S)
    if generic_fence:
        candidates.append(generic_fence.group(1).strip())

    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        candidates.append(text[start : end + 1])

    seen = set()
    unique = []
    for candidate in candidates:
        candidate = candidate.strip()
        if candidate and candidate not in seen:
            unique.append(candidate)
            seen.add(candidate)
    return unique


def _value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _value_to_json(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _image_to_data_url(path: str) -> str:
    mime_type, _ = mimetypes.guess_type(path)
    if not mime_type:
        mime_type = "image/png"
    with open(path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def _encoding_name_for_model(model: Optional[str]) -> str:
    model_key = (model or "gpt-4").lower()
    for name, encoding_name in sorted(MODEL_ENCODINGS.items(), key=lambda item: len(item[0]), reverse=True):
        if model_key == name or model_key.startswith(f"{name}-") or model_key.startswith(f"{name}_"):
            return encoding_name
    return "cl100k_base"


def calculate_token_cost_line(text: str, model: Optional[str] = "gpt-4") -> int:
    try:
        # Get the appropriate encoding
        encoding = tiktoken.get_encoding(_encoding_name_for_model(model))
        # Encode and count tokens
        tokens = encoding.encode(text)
        return len(tokens)
    except Exception:
        # 使用估计值
        return max(1, int(len(text) / 3.5))


def _truncate_text_to_token_budget(
    text: str,
    token_budget: int,
    model: Optional[str] = "gpt-4",
    from_end: bool = False,
) -> str:
    if token_budget <= 0:
        return ""

    encoding = tiktoken.get_encoding(_encoding_name_for_model(model))
    tokens = encoding.encode(text)
    if len(tokens) <= token_budget:
        return text

    selected_tokens = tokens[-token_budget:] if from_end else tokens[:token_budget]
    return encoding.decode(selected_tokens)


def _spreadsheet_content_note(table_format: str) -> str:
    if table_format in IMAGE_TABLE_FORMATS:
        return (
            "The spreadsheet content is provided as worksheet screenshots attached to this message. "
            "Use the image list below only to understand worksheet order."
        )
    if table_format in TEXT_IMAGE_TABLE_FORMATS:
        return (
            "The spreadsheet content is provided in two complementary forms: a text serialization below "
            "and worksheet screenshots attached to this message."
        )
    if table_format in {"json_rows", "json_cells"}:
        return "The spreadsheet content is provided as JSON. Preserve cell coordinates and worksheet names carefully."
    if table_format == "html":
        return "The spreadsheet content is provided as HTML tables. Each table records the worksheet name and used range."
    if table_format == "latex":
        return "The spreadsheet content is provided as LaTeX tables. Each worksheet section includes its used range."
    return "The spreadsheet content is provided as plain text. Each worksheet starts with its sheet name and used range."


class TableInputBuilder:
    def __init__(
        self,
        real_dir: str,
        latex_dir: str,
        image_dir: Optional[str] = None,
        excel_1_image_dir: Optional[str] = None,
        table_format: str = "latex",
        include_coordinates: bool = True,
        fill_merged: bool = False,
        max_text_tokens: int = 0,
        token_model: Optional[str] = "gpt-4",
        image_cache_dir: Optional[str] = None,
    ):
        if table_format not in SUPPORTED_TABLE_FORMATS:
            raise ValueError(f"Unsupported table_format `{table_format}`. Choices: {SUPPORTED_TABLE_FORMATS}")

        self.real_dir = real_dir
        self.latex_dir = latex_dir
        self.image_dir = image_dir
        self.excel_1_image_dir = excel_1_image_dir
        self.table_format = table_format
        self.include_coordinates = include_coordinates
        self.fill_merged = fill_merged
        self.max_text_tokens = max_text_tokens
        self.token_model = token_model or "gpt-4"
        self.image_cache_dir = image_cache_dir

    def build(self, file_name: str) -> Tuple[str, Dict[str, Any]]:
        xlsx_path = os.path.join(self.real_dir, f"{file_name}.xlsx")
        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"Spreadsheet not found: {xlsx_path}")

        image_paths = []
        # 图像表示
        if self.table_format in IMAGE_TABLE_FORMATS:
            table_text, image_paths = self._build_image_text(file_name, xlsx_path=xlsx_path)
            representation = self.table_format
        # 混合表示
        elif self.table_format in TEXT_IMAGE_TABLE_FORMATS:
            text_format, image_format = TEXT_IMAGE_TABLE_FORMATS[self.table_format]
            table_text = self._build_table_text(file_name, xlsx_path, table_format=text_format)
            image_text, image_paths = self._build_image_text(file_name, image_format=image_format, xlsx_path=xlsx_path)
            table_text = f"{table_text}\n\n# Spreadsheet Image Information\n{image_text}"
            representation = self.table_format
        # 纯文本的表示
        else:
            table_text = self._build_table_text(file_name, xlsx_path)
            representation = self.table_format

        table_text, truncated, text_tokens = self._maybe_truncate(table_text)
        metadata = {
            "xlsx_path": xlsx_path,
            "table_format": representation,
            "include_coordinates": self.include_coordinates,
            "fill_merged": self.fill_merged,
            "truncated": truncated,
            "text_chars": len(table_text),
            "text_tokens": text_tokens,
            "max_text_tokens": self.max_text_tokens,
            "token_model": self.token_model,
            "image_paths": image_paths,
        }
        return table_text, metadata

    def _build_table_text(self, file_name: str, xlsx_path: str, table_format: Optional[str] = None) -> str:
        table_format = table_format or self.table_format
        # Latex格式，去读取原始文件
        if table_format == "latex":
            latex_path = os.path.join(self.latex_dir, f"{file_name}.txt")
            if not os.path.exists(latex_path):
                raise FileNotFoundError(f"LaTeX table representation not found: {latex_path}")
            return read_text_with_encoding_fallback(latex_path)
        
        # 其他格式，基于excel文件进行转换
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        parts = []
        for ws in wb.worksheets:
            if table_format == "csv":
                body = self._serialize_delimited(ws, delimiter=",")
            elif table_format == "tsv":
                body = self._serialize_delimited(ws, delimiter="\t")
            elif table_format == "markdown":
                body = self._serialize_markdown(ws)
            elif table_format == "dataframe":
                body = self._serialize_fixed_width(ws)
            elif table_format == "json_rows":
                body = self._serialize_json_rows(ws)
            elif table_format == "json_cells":
                body = self._serialize_json_cells(ws)
            elif table_format == "html":
                body = self._serialize_html(ws)
            else:
                raise ValueError(f"Unsupported table_format: {table_format}")
            parts.append(f"## Sheet: {ws.title}\n{body}")
        wb.close()
        return "\n\n".join(parts)

    def _build_image_text(
        self,
        file_name: str,
        image_format: Optional[str] = None,
        xlsx_path: Optional[str] = None,
    ) -> Tuple[str, List[str]]:
        image_format = image_format or self.table_format
        if image_format == "default_image":
            image_paths = self._render_image_paths(file_name, xlsx_path, image_format) if xlsx_path else []
            error_hint = "`default_image` requires an xlsx_path and image_cache_dir for rendering."
        else:
            image_dir = self.image_dir if image_format == "image" else self.excel_1_image_dir
            if not image_dir:
                raise ValueError(f"`{image_format}` requires an image directory.")

            image_paths = self._find_image_paths(image_dir, file_name)
            error_hint = f"No {image_format} images found for `{file_name}` in {image_dir}"

        if not image_paths:
            raise FileNotFoundError(error_hint)

        lines = [
            f"{len(image_paths)} image(s) are attached below in sheet order.",
            "If multiple images are attached, treat them as screenshots of different sheets from the same workbook.",
        ]
        for index, path in enumerate(image_paths, start=1):
            lines.append(f"- Image {index}: {os.path.basename(path)}")
        return "\n".join(lines), image_paths

    def _find_image_paths(self, image_dir: str, file_name: str) -> List[str]:
        candidates = []
        for ext in sorted(IMAGE_EXTENSIONS):
            candidates.append(os.path.join(image_dir, f"{file_name}{ext}"))

        numbered = []
        for ext in sorted(IMAGE_EXTENSIONS):
            numbered.extend(glob.glob(os.path.join(image_dir, f"{glob.escape(file_name)}___*{ext}")))

        existing = [path for path in candidates if os.path.exists(path)]
        if existing:
            return sorted(existing)

        def sheet_index(path: str) -> Tuple[int, str]:
            stem = os.path.splitext(os.path.basename(path))[0]
            match = re.search(r"___(\d+)$", stem)
            return (int(match.group(1)) if match else 0, path)

        return sorted([path for path in numbered if os.path.exists(path)], key=sheet_index)

    def _render_image_paths(self, file_name: str, xlsx_path: Optional[str], image_format: str) -> List[str]:
        if not xlsx_path or not self.image_cache_dir:
            return []

        cache_dir = self.image_cache_dir
        os.makedirs(cache_dir, exist_ok=True)

        wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
        image_paths = []
        try:
            sheet_count = len(wb.worksheets)
            for idx, ws in enumerate(wb.worksheets, start=1):
                bounds = self._used_bounds(ws)
                cell_range = self._bounds_to_range(bounds)
                out_name = f"{file_name}.png" if sheet_count == 1 else f"{file_name}___{idx}.png"
                out_path = os.path.join(cache_dir, out_name)
                if not os.path.exists(out_path):
                    with _DEFAULT_IMAGE_RENDER_LOCK:
                        if not os.path.exists(out_path):
                            render_excel_range_to_png(
                                src_path=xlsx_path,
                                sheet_name=ws.title,
                                cell_range=cell_range,
                                out_path=out_path,
                            )
                image_paths.append(out_path)
        finally:
            wb.close()
        return image_paths

    def _used_bounds(self, ws) -> Tuple[int, int, int, int]:
        min_row = min_col = None
        max_row = max_col = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    min_row = cell.row if min_row is None else min(min_row, cell.row)
                    min_col = cell.column if min_col is None else min(min_col, cell.column)
                    max_row = cell.row if max_row is None else max(max_row, cell.row)
                    max_col = cell.column if max_col is None else max(max_col, cell.column)
        if min_row is None:
            return 1, 1, 1, 1
        return min_row, min_col, max_row, max_col

    def _merged_value_map(self, ws) -> Dict[Tuple[int, int], Any]:
        if not self.fill_merged:
            return {}
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            top_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = top_value
        return merged_values

    def _matrix(self, ws, include_coordinates: Optional[bool] = None) -> Tuple[List[List[str]], Tuple[int, int, int, int]]:
        include_coordinates = self.include_coordinates if include_coordinates is None else include_coordinates
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        merged_values = self._merged_value_map(ws)
        rows = []

        if include_coordinates:
            rows.append(["row"] + [get_column_letter(col) for col in range(min_col, max_col + 1)])

        for row_idx in range(min_row, max_row + 1):
            row = []
            if include_coordinates:
                row.append(str(row_idx))
            for col_idx in range(min_col, max_col + 1):
                value = merged_values.get((row_idx, col_idx), ws.cell(row_idx, col_idx).value)
                row.append(_value_to_text(value))
            rows.append(row)
        return rows, (min_row, min_col, max_row, max_col)

    def _serialize_delimited(self, ws, delimiter: str) -> str:
        rows, bounds = self._matrix(ws)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\n")
        writer.writerows(rows)
        return f"Used Range: {self._bounds_to_range(bounds)}\n{buffer.getvalue().rstrip()}"

    def _serialize_markdown(self, ws) -> str:
        rows, bounds = self._matrix(ws)
        escaped = [[self._escape_markdown_cell(cell) for cell in row] for row in rows]
        widths = [max(len(row[idx]) for row in escaped) for idx in range(len(escaped[0]))]

        def fmt(row):
            return "| " + " | ".join(cell.ljust(widths[idx]) for idx, cell in enumerate(row)) + " |"

        header = fmt(escaped[0])
        separator = "| " + " | ".join("-" * max(3, width) for width in widths) + " |"
        body = "\n".join(fmt(row) for row in escaped[1:])
        return f"Used Range: {self._bounds_to_range(bounds)}\n{header}\n{separator}\n{body}"

    def _serialize_fixed_width(self, ws) -> str:
        rows, bounds = self._matrix(ws)
        widths = [max(len(row[idx]) for row in rows) for idx in range(len(rows[0]))]
        lines = []
        for row in rows:
            lines.append("  ".join(row[idx].ljust(widths[idx]) for idx in range(len(row))))
        return f"Used Range: {self._bounds_to_range(bounds)}\n" + "\n".join(lines)

    def _serialize_json_rows(self, ws) -> str:
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        merged_values = self._merged_value_map(ws)
        columns = [get_column_letter(col) for col in range(min_col, max_col + 1)]
        rows = []
        for row_idx in range(min_row, max_row + 1):
            values = []
            for col_idx in range(min_col, max_col + 1):
                value = merged_values.get((row_idx, col_idx), ws.cell(row_idx, col_idx).value)
                values.append(_value_to_json(value))
            rows.append({"row": row_idx, "values": values})

        payload = {
            "sheet_name": ws.title,
            "used_range": self._bounds_to_range((min_row, min_col, max_row, max_col)),
            "columns": columns,
            "rows": rows,
            "merged_ranges": self._merged_ranges(ws),
        }
        return json.dumps(payload, ensure_ascii=False)

    def _serialize_json_cells(self, ws) -> str:
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        cells = []
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                value = ws.cell(row_idx, col_idx).value
                if value in (None, ""):
                    continue
                cells.append(
                    {
                        "cell": f"{get_column_letter(col_idx)}{row_idx}",
                        "row": row_idx,
                        "column": get_column_letter(col_idx),
                        "value": _value_to_json(value),
                    }
                )

        payload = {
            "sheet_name": ws.title,
            "used_range": self._bounds_to_range((min_row, min_col, max_row, max_col)),
            "cells": cells,
            "merged_ranges": self._merged_ranges(ws),
        }
        return json.dumps(payload, ensure_ascii=False)

    def _serialize_html(self, ws) -> str:
        rows, bounds = self._matrix(ws)
        lines = [
            f"<table data-sheet={json.dumps(ws.title)} data-used-range={json.dumps(self._bounds_to_range(bounds))}>"
        ]
        for row in rows:
            lines.append("<tr>" + "".join(f"<td>{html.escape(cell)}</td>" for cell in row) + "</tr>")
        lines.append("</table>")
        return "\n".join(lines)

    def _merged_ranges(self, ws) -> List[Dict[str, Any]]:
        ranges = []
        for merged_range in ws.merged_cells.ranges:
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            ranges.append({"range": str(merged_range), "value": _value_to_json(value)})
        return ranges

    def _bounds_to_range(self, bounds: Tuple[int, int, int, int]) -> str:
        min_row, min_col, max_row, max_col = bounds
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def _escape_markdown_cell(self, cell: str) -> str:
        return str(cell).replace("\n", " ").replace("|", "\\|")

    def _maybe_truncate(self, text: str) -> Tuple[str, bool, int]:
        text_tokens = calculate_token_cost_line(text, self.token_model)
        if not self.max_text_tokens or self.max_text_tokens <= 0 or text_tokens <= self.max_text_tokens:
            return text, False, text_tokens

        marker = "\n\n...[truncated to fit token budget]...\n\n"
        marker_tokens = calculate_token_cost_line(marker, self.token_model)
        if self.max_text_tokens <= marker_tokens + 2:
            truncated = _truncate_text_to_token_budget(text, self.max_text_tokens, self.token_model)
            return truncated, True, calculate_token_cost_line(truncated, self.token_model)

        body_budget = self.max_text_tokens - marker_tokens
        head_budget = body_budget // 2
        tail_budget = body_budget - head_budget
        truncated = (
            _truncate_text_to_token_budget(text, head_budget, self.token_model)
            + marker
            + _truncate_text_to_token_budget(text, tail_budget, self.token_model, from_end=True)
        )
        return truncated, True, calculate_token_cost_line(truncated, self.token_model)


class RealHiTCoTSolver:
    def __init__(self, *args, **kwargs):
        self.url = kwargs.get("url", "localhost:8000")
        self.model_name = kwargs.get("model_name", None)
        self.model_params = {
            "top_p": kwargs.get("top_p", 1.0),
            "temperature": kwargs.get("temperature", 0),
        }
        self.table_format = kwargs.get("table_format", "latex")
        self.include_coordinates = kwargs.get("include_coordinates", True)
        self.fill_merged = kwargs.get("fill_merged", False)
        self.max_text_tokens = kwargs.get("max_text_tokens", 0)
        self.max_retries = kwargs.get("max_retries", 3)
        self.save_prompts = kwargs.get("save_prompts", False)
        self.dry_run = kwargs.get("dry_run", False)

    def _builder(self, data: Dict[str, Any]) -> TableInputBuilder:
        return TableInputBuilder(
            real_dir=data["real_dir"],
            latex_dir=data["latex_dir"],
            image_dir=data.get("image_dir"),
            excel_1_image_dir=data.get("excel_1_image_dir"),
            table_format=self.table_format,
            include_coordinates=self.include_coordinates,
            fill_merged=self.fill_merged,
            max_text_tokens=self.max_text_tokens,
            token_model=self.model_name,
            image_cache_dir=data.get("image_cache_dir"),
        )

    def build_prompt(self, data: Dict[str, Any]) -> Tuple[Union[str, List[Dict[str, Any]]], Dict[str, Any]]:
        table_text, table_metadata = self._builder(data).build(data["FileName"])
        constraint_key = data.get("SubQType") if data.get("QuestionType") == "Data Analysis" else data.get("QuestionType")
        answer_constraints = Answer_Prompt.get(constraint_key) or Answer_Prompt.get(data.get("QuestionType"), "")
        prompt = USER_PROMPT_TEMPLATE.format(
            spreadsheet_content_note=_spreadsheet_content_note(table_metadata["table_format"]),
            table_text=table_text,
            question=data.get("Question", ""),
            answer_constraints=answer_constraints,
            final_answer_prefix=FINAL_ANSWER_PREFIX,
        )
        image_paths = table_metadata.get("image_paths") or []
        if image_paths:
            content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
            for image_path in image_paths:
                content.append(
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": _image_to_data_url(image_path)
                        },
                    }
                )
            return content, table_metadata
        return prompt, table_metadata

    def get_solution(self, data: Dict[str, Any]) -> Dict[str, Any]:
        messages: List[Dict[str, Any]] = []
        attempts: List[Dict[str, Any]] = []
        parsed: Optional[ParsedResponse] = None
        solution = ""
        error = None
        table_metadata: Dict[str, Any] = {}

        try:
            prompt, table_metadata = self.build_prompt(data)
            messages = [{"role": "user", "content": prompt}]

            if self.dry_run:
                raise ResponseFormatError("dry_run enabled; prompt was built but model was not called.")

            for attempt_idx in range(1, self.max_retries + 1):
                resp = model_resp(
                    self.url,
                    messages,
                    model_params=self.model_params,
                    model_name=self.model_name,
                )
                content = ""
                if resp and resp.get("message"):
                    content = resp["message"].get("content") or ""

                solution = content
                try:
                    parsed = parse_and_validate_response(content)
                    attempts.append(
                        {
                            "attempt": attempt_idx,
                            "valid": True,
                            "response": content,
                            "format_error": None,
                        }
                    )
                    messages.append({"role": "assistant", "content": content})
                    error = None
                    break
                except ResponseFormatError as exc:
                    format_error = str(exc)
                    attempts.append(
                        {
                            "attempt": attempt_idx,
                            "valid": False,
                            "response": content,
                            "format_error": format_error,
                        }
                    )
                    messages.append({"role": "assistant", "content": content})
                    messages.append(
                        {
                            "role": "user",
                            "content": REPAIR_PROMPT_TEMPLATE.format(
                                error=format_error,
                                final_answer_prefix=FINAL_ANSWER_PREFIX,
                            ),
                        }
                    )
                    error = format_error
        except Exception:
            error = traceback.format_exc()

        out = copy.deepcopy(data)
        out.update(
            {
                "table_metadata": table_metadata,
                "attempts": attempts,
                "format_valid": parsed is not None,
                "error": error,
                "solution": solution,
                "parsed_response": parsed.payload if parsed else None,
                "model_answer": parsed.final_answer if parsed else "",
            }
        )
        if self.save_prompts:
            out["messages"] = messages
        return out

    def _score_qa(self, reference: str, response: str) -> Dict[str, Any]:
        return QAMetric().compute([reference], [response])

    def __call__(self, data: Dict[str, Any]) -> Dict[str, Any]:
        result = self.get_solution(data)
        reference = data.get("ProcessedAnswer", "")
        response = result.get("model_answer", "")
        metric_scores: Dict[str, Any] = {}

        try:
            if data.get("QuestionType") == "Structure Comprehending":
                reference = response
                swap_data = copy.deepcopy(data)
                swap_data["FileName"] = f"{data['FileName']}_swap"
                swap_result = self.get_solution(swap_data)
                response = swap_result.get("model_answer", "")
                result["structure_reference_run"] = {
                    "FileName": data["FileName"],
                    "model_answer": reference,
                    "format_valid": result.get("format_valid"),
                    "error": result.get("error"),
                    "attempts": result.get("attempts"),
                    "parsed_response": result.get("parsed_response"),
                    "table_metadata": result.get("table_metadata"),
                }
                result["structure_swap_run"] = {
                    "FileName": swap_data["FileName"],
                    "model_answer": response,
                    "format_valid": swap_result.get("format_valid"),
                    "error": swap_result.get("error"),
                    "attempts": swap_result.get("attempts"),
                    "parsed_response": swap_result.get("parsed_response"),
                    "table_metadata": swap_result.get("table_metadata"),
                }
                if reference and response:
                    metric_scores = self._score_qa(reference, response)
                result["solution"] = swap_result.get("solution", "")
                result["model_answer"] = response
                result["format_valid"] = bool(result.get("format_valid")) and bool(swap_result.get("format_valid"))
                result["error"] = result.get("error") or swap_result.get("error")
            else:
                if response:
                    metric_scores = self._score_qa(reference, response)
        except Exception:
            result["eval_error"] = traceback.format_exc()

        result["eval"] = {
            "Model_Answer": response,
            "Reference_Answer": reference,
            "F1": metric_scores.get("F1", None),
            "EM": metric_scores.get("EM", None),
            "ROUGE-L": metric_scores.get("ROUGE-L", None),
            "SacreBLEU": metric_scores.get("SacreBLEU", None),
        }
        return result
