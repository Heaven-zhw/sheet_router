import base64
import copy
import csv
import glob
import html
import io
import json
import mimetypes
import os
import re
import threading
import traceback
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
import tiktoken
from openpyxl.utils import get_column_letter

from ..client import ClientJupyterKernel, extract_code
from ..eval.spreadsheet import compare_workbooks
from ..excel2image_linux import render_excel_range_to_png
from ..excel2tex import convert_excel_to_latex
from ..utils import model_resp, read_text_with_encoding_fallback


SPREADSHEET_DATA_SPLITS = {
    "all_912": {
        "root": "dataset/spreadsheetbench/all_data_912_v0.1",
        "input_suffix": "input",
        "answer_suffix": "answer",
        "num_test_cases": 3,
    },
    "verified_400": {
        "root": "dataset/spreadsheetbench/spreadsheetbench_verified_400",
        "input_suffix": "init",
        "answer_suffix": "golden",
        "num_test_cases": 1,
    },
}

SUPPORTED_TABLE_FORMATS = (
    "latex",
    "csv",
    "tsv",
    "markdown",
    "dataframe",
    "json_rows",
    "json_cells",
    "html",
    "image",
    "excel_1_image",
    "default_image",
    "html+image",
    "markdown+image",
    "latex+image",
    "html+excel_1_image",
    "latex+excel_1_image",
    "markdown+excel_1_image",
    "html+default_image",
    "latex+default_image",
    "markdown+default_image",
)

IMAGE_TABLE_FORMATS = {"image", "excel_1_image", "default_image"}
TEXT_IMAGE_TABLE_FORMATS = {
    "html+image": ("html", "image"),
    "markdown+image": ("markdown", "image"),
    "latex+image": ("latex", "image"),
    "html+excel_1_image": ("html", "excel_1_image"),
    "latex+excel_1_image": ("latex", "excel_1_image"),
    "markdown+excel_1_image": ("markdown", "excel_1_image"),
    "html+default_image": ("html", "default_image"),
    "latex+default_image": ("latex", "default_image"),
    "markdown+default_image": ("markdown", "default_image"),
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
_DEFAULT_IMAGE_RENDER_LOCK = threading.Lock()

MODEL_ENCODINGS = {
    "gpt-4": "cl100k_base",
    "gpt-4-turbo": "cl100k_base",
    "gpt-4o": "o200k_base",
    "qwen": "cl100k_base",
    "qwen3": "cl100k_base",
    "qwen3-vl": "cl100k_base",
    "qwen3.5": "cl100k_base",
    "qwen3.5-vl": "cl100k_base",
    "qwen35": "cl100k_base",
    "qwen35-vl": "cl100k_base",
}

FINAL_CODE_CHECK = """
import os
assert os.path.exists({output_path!r}), "Expected output file was not created: {output_path}"
print("OUTPUT_FILE_CREATED", {output_path!r})
"""

REMOVE_OUTPUT_FILE = """
import os
if os.path.exists({output_path!r}):
    os.remove({output_path!r})
"""


FIRST_ROUND_PROMPT = """You are a spreadsheet expert who can manipulate spreadsheets through Python code.

You need to solve the given spreadsheet manipulation question, which contains several types of information:
- instruction: The question about spreadsheet manipulation.
- spreadsheet_path: The path of the spreadsheet file you need to manipulate.
- spreadsheet_content: The spreadsheet content. It may be provided as text, images, or text together with images. When images are attached, they are screenshots of worksheets from the same workbook in worksheet order.
- instruction_type: There are two values (Cell-Level Manipulation, Sheet-Level Manipulation) used to indicate whether the answer applies only to specific cells or to the entire worksheet.
- answer_position: The position that needs to be modified or filled. For Cell-Level Manipulation questions, this field is filled with the cell position; for Sheet-Level Manipulation, it is the maximum range of cells you need to modify. You only need to modify or fill in values within the cell range specified by answer_position unless the instruction explicitly asks you to create, delete, or clear sheets/ranges.
- answer_sheet: The worksheet or worksheets where the final answer should be written, if provided.
- output_path: You need to generate the modified spreadsheet file in this new path.

Below is the spreadsheet manipulation question you need to solve:
### instruction
{instruction}

### spreadsheet_path
{spreadsheet_path}

### spreadsheet_content
{spreadsheet_content_note}

{spreadsheet_content}

### instruction_type
{instruction_type}

### answer_position
{answer_position}

### answer_sheet
{answer_sheet}

### output_path
{output_path}

Now, generate completed Python code to solve the question. The code should be placed in one markdown python code block, and the code should be executable in a Jupyter notebook environment. The code should save the modified spreadsheet file at the specified output_path.

Note: DO NOT use Excel formulas. Please calculate in Python and write the answer to the target cell or sheet directly. Please use cell references, sheet names, headers, and workbook content robustly, since your code will be applied to other spreadsheets with the same layout pattern."""


RETRY_PROMPT = """The previous Python code failed or did not create a usable output spreadsheet.

Attempt {attempt} execution/evaluation result:
{feedback}

Please fix the code. Return only the complete corrected Python code in one markdown python code block. The code must read the input workbook at:
{input_path}

and save the modified workbook at:
{output_path}

Remember: do not use Excel formulas as final answers; calculate in Python and write final values/styles directly."""


def _value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _value_to_json(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _image_to_data_url(path: str) -> str:
    mime_type, _ = mimetypes.guess_type(path)
    if not mime_type:
        mime_type = "image/png"
    with open(path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def _encoding_name_for_model(model: Optional[str]) -> str:
    model_key = (model or "gpt-4").lower()
    for name, encoding_name in sorted(MODEL_ENCODINGS.items(), key=lambda item: len(item[0]), reverse=True):
        if model_key == name or model_key.startswith(f"{name}-") or model_key.startswith(f"{name}_"):
            return encoding_name
    return "cl100k_base"


def calculate_token_cost_line(text: str, model: Optional[str] = "gpt-4") -> int:
    try:
        encoding = tiktoken.get_encoding(_encoding_name_for_model(model))
        return len(encoding.encode(text))
    except Exception:
        return max(1, int(len(text) / 3.5))


def _truncate_text_to_token_budget(
    text: str,
    token_budget: int,
    model: Optional[str] = "gpt-4",
    from_end: bool = False,
) -> str:
    if token_budget <= 0:
        return ""

    try:
        encoding = tiktoken.get_encoding(_encoding_name_for_model(model))
        tokens = encoding.encode(text)
        if len(tokens) <= token_budget:
            return text
        selected_tokens = tokens[-token_budget:] if from_end else tokens[:token_budget]
        return encoding.decode(selected_tokens)
    except Exception:
        char_budget = max(1, int(token_budget * 3.5))
        return text[-char_budget:] if from_end else text[:char_budget]


def _spreadsheet_content_note(table_format: str) -> str:
    if table_format in IMAGE_TABLE_FORMATS:
        return (
            "The workbook content is provided as worksheet screenshots attached to this message. "
            "Use the image list below only to understand worksheet order."
        )
    if table_format in TEXT_IMAGE_TABLE_FORMATS:
        return (
            "The workbook content is provided in two complementary forms: a text serialization below "
            "and worksheet screenshots attached to this message."
        )
    if table_format in {"json_rows", "json_cells"}:
        return "The workbook content is provided as JSON. Preserve cell coordinates and worksheet names carefully."
    if table_format == "html":
        return "The workbook content is provided as HTML tables. Each table records the worksheet name and used range."
    if table_format == "latex":
        return "The workbook content is provided as LaTeX tables. Each worksheet section includes its used range."
    return "The workbook content is provided as plain text. Each worksheet starts with its sheet name and used range."


class SpreadsheetTableInputBuilder:
    def __init__(
        self,
        real_dir: str,
        latex_dir: Optional[str] = None,
        image_dir: Optional[str] = None,
        excel_1_image_dir: Optional[str] = None,
        table_format: str = "markdown",
        include_coordinates: bool = True,
        fill_merged: bool = False,
        max_text_tokens: int = 0,
        token_model: Optional[str] = "gpt-4",
        image_cache_dir: Optional[str] = None,
    ):
        if table_format not in SUPPORTED_TABLE_FORMATS:
            raise ValueError(f"Unsupported table_format `{table_format}`. Choices: {SUPPORTED_TABLE_FORMATS}")

        self.real_dir = real_dir
        self.latex_dir = latex_dir
        self.image_dir = image_dir
        self.excel_1_image_dir = excel_1_image_dir
        self.table_format = table_format
        self.include_coordinates = include_coordinates
        self.fill_merged = fill_merged
        self.max_text_tokens = max_text_tokens
        self.token_model = token_model or "gpt-4"
        self.image_cache_dir = image_cache_dir

    def build(self, input_file: str) -> Tuple[str, Dict[str, Any]]:
        xlsx_path = os.path.join(self.real_dir, input_file)
        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"Spreadsheet not found: {xlsx_path}")

        image_paths: List[str] = []
        if self.table_format in IMAGE_TABLE_FORMATS:
            table_text, image_paths = self._build_image_text(input_file, xlsx_path=xlsx_path)
            representation = self.table_format
        elif self.table_format in TEXT_IMAGE_TABLE_FORMATS:
            text_format, image_format = TEXT_IMAGE_TABLE_FORMATS[self.table_format]
            table_text = self._build_table_text(input_file, xlsx_path, table_format=text_format)
            image_text, image_paths = self._build_image_text(input_file, image_format=image_format, xlsx_path=xlsx_path)
            table_text = f"{table_text}\n\n# Spreadsheet Image Information\n{image_text}"
            representation = self.table_format
        else:
            table_text = self._build_table_text(input_file, xlsx_path)
            representation = self.table_format

        table_text, truncated, text_tokens = self._maybe_truncate(table_text)
        metadata = {
            "xlsx_path": xlsx_path,
            "input_file": input_file,
            "table_format": representation,
            "include_coordinates": self.include_coordinates,
            "fill_merged": self.fill_merged,
            "truncated": truncated,
            "text_chars": len(table_text),
            "text_tokens": text_tokens,
            "max_text_tokens": self.max_text_tokens,
            "token_model": self.token_model,
            "image_paths": image_paths,
        }
        return table_text, metadata

    def _build_table_text(self, input_file: str, xlsx_path: str, table_format: Optional[str] = None) -> str:
        table_format = table_format or self.table_format
        if table_format == "latex":
            return self._serialize_latex_workbook(input_file, xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        parts = []
        try:
            for ws in wb.worksheets:
                if table_format == "csv":
                    body = self._serialize_delimited(ws, delimiter=",")
                elif table_format == "tsv":
                    body = self._serialize_delimited(ws, delimiter="\t")
                elif table_format == "markdown":
                    body = self._serialize_markdown(ws)
                elif table_format == "dataframe":
                    body = self._serialize_fixed_width(ws)
                elif table_format == "json_rows":
                    body = self._serialize_json_rows(ws)
                elif table_format == "json_cells":
                    body = self._serialize_json_cells(ws)
                elif table_format == "html":
                    body = self._serialize_html(ws)
                else:
                    raise ValueError(f"Unsupported table_format: {table_format}")
                parts.append(f"## Sheet: {ws.title}\n{body}")
        finally:
            wb.close()
        return "\n\n".join(parts)

    def _serialize_latex_workbook(self, input_file: str, xlsx_path: str) -> str:
        basename = os.path.splitext(input_file)[0]
        if self.latex_dir:
            latex_candidates = [
                os.path.join(self.latex_dir, f"{basename}.txt"),
                os.path.join(self.latex_dir, input_file.replace(".xlsx", ".txt")),
            ]
            for latex_path in latex_candidates:
                if os.path.exists(latex_path):
                    return read_text_with_encoding_fallback(latex_path)

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
        parts = []
        try:
            for ws in wb.worksheets:
                bounds = self._used_bounds(ws)
                cell_range = self._bounds_to_range(bounds)
                body = convert_excel_to_latex(xlsx_path, ws.title, cell_range)
                parts.append(f"## Sheet: {ws.title}\nUsed Range: {cell_range}\n{body}")
        finally:
            wb.close()
        return "\n\n".join(parts)

    def _build_image_text(
        self,
        input_file: str,
        image_format: Optional[str] = None,
        xlsx_path: Optional[str] = None,
    ) -> Tuple[str, List[str]]:
        image_format = image_format or self.table_format
        if image_format == "default_image":
            # default_image 明确表示运行时从 xlsx 渲染工作表截图，不读取数据集预生成图片目录。
            image_paths = self._render_image_paths(input_file, xlsx_path, image_format) if xlsx_path else []
            error_hint = "`default_image` requires an xlsx_path and image_cache_dir for rendering."
        else:
            image_dir = self.image_dir if image_format == "image" else self.excel_1_image_dir
            if not image_dir:
                raise ValueError(f"`{image_format}` requires an image directory.")

            image_paths = self._find_image_paths(image_dir, input_file)
            error_hint = f"No {image_format} images found for `{input_file}` in {image_dir}"

        if not image_paths:
            raise FileNotFoundError(error_hint)

        lines = [
            f"{len(image_paths)} image(s) are attached below in worksheet order.",
            "If multiple images are attached, treat them as screenshots of different worksheets from the same workbook.",
        ]
        for index, path in enumerate(image_paths, start=1):
            lines.append(f"- Image {index}: {os.path.basename(path)}")
        return "\n".join(lines), image_paths

    def _find_image_paths(self, image_dir: str, input_file: str) -> List[str]:
        def find_in_dir(search_dir: str) -> List[str]:
            stem = os.path.splitext(input_file)[0]
            candidates = []
            for ext in sorted(IMAGE_EXTENSIONS):
                candidates.extend(
                    [
                        os.path.join(search_dir, f"{stem}{ext}"),
                        os.path.join(search_dir, f"{input_file}{ext}"),
                    ]
                )

            numbered = []
            for ext in sorted(IMAGE_EXTENSIONS):
                numbered.extend(glob.glob(os.path.join(search_dir, f"{glob.escape(stem)}___*{ext}")))
                numbered.extend(glob.glob(os.path.join(search_dir, f"{glob.escape(input_file)}___*{ext}")))

            existing = [path for path in candidates if os.path.exists(path)]
            if existing:
                return sorted(existing)

            def sheet_index(path: str) -> Tuple[int, str]:
                match = re.search(r"___(\d+)$", os.path.splitext(os.path.basename(path))[0])
                return (int(match.group(1)) if match else 0, path)

            return sorted([path for path in numbered if os.path.exists(path)], key=sheet_index)

        image_paths = find_in_dir(image_dir)
        if image_paths:
            return image_paths

        parent_dir = os.path.dirname(os.path.normpath(image_dir))
        if parent_dir and parent_dir != os.path.normpath(image_dir):
            # 兼容 SpreadsheetBench 的扁平图片目录：如果传入的是 root/{id}，
            # 但图片实际在 root/1_{id}_init___1.png，则回退到父目录再查一次。
            return find_in_dir(parent_dir)

        return []

    def _render_image_paths(self, input_file: str, xlsx_path: str, image_format: str) -> List[str]:
        if not self.image_cache_dir:
            return []

        stem = os.path.splitext(input_file)[0]
        cache_dir = self.image_cache_dir
        os.makedirs(cache_dir, exist_ok=True)

        wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
        image_paths = []
        try:
            sheet_count = len(wb.worksheets)
            for idx, ws in enumerate(wb.worksheets, start=1):
                bounds = self._used_bounds(ws)
                cell_range = self._bounds_to_range(bounds)
                out_name = f"{stem}.png" if sheet_count == 1 else f"{stem}___{idx}.png"
                out_path = os.path.join(cache_dir, out_name)
                if not os.path.exists(out_path):
                    with _DEFAULT_IMAGE_RENDER_LOCK:
                        if not os.path.exists(out_path):
                            render_excel_range_to_png(
                                src_path=xlsx_path,
                                sheet_name=ws.title,
                                cell_range=cell_range,
                                out_path=out_path,
                            )
                image_paths.append(out_path)
        finally:
            wb.close()
        return image_paths

    def _used_bounds(self, ws) -> Tuple[int, int, int, int]:
        min_row = min_col = None
        max_row = max_col = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    min_row = cell.row if min_row is None else min(min_row, cell.row)
                    min_col = cell.column if min_col is None else min(min_col, cell.column)
                    max_row = cell.row if max_row is None else max(max_row, cell.row)
                    max_col = cell.column if max_col is None else max(max_col, cell.column)
        if min_row is None:
            return 1, 1, 1, 1
        return min_row, min_col, max_row, max_col

    def _merged_value_map(self, ws) -> Dict[Tuple[int, int], Any]:
        if not self.fill_merged:
            return {}
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            top_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = top_value
        return merged_values

    def _matrix(self, ws, include_coordinates: Optional[bool] = None) -> Tuple[List[List[str]], Tuple[int, int, int, int]]:
        include_coordinates = self.include_coordinates if include_coordinates is None else include_coordinates
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        merged_values = self._merged_value_map(ws)
        rows = []

        if include_coordinates:
            rows.append(["row"] + [get_column_letter(col) for col in range(min_col, max_col + 1)])

        for row_idx in range(min_row, max_row + 1):
            row = []
            if include_coordinates:
                row.append(str(row_idx))
            for col_idx in range(min_col, max_col + 1):
                value = merged_values.get((row_idx, col_idx), ws.cell(row_idx, col_idx).value)
                row.append(_value_to_text(value))
            rows.append(row)
        return rows, (min_row, min_col, max_row, max_col)

    def _serialize_delimited(self, ws, delimiter: str) -> str:
        rows, bounds = self._matrix(ws)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\n")
        writer.writerows(rows)
        return f"Used Range: {self._bounds_to_range(bounds)}\n{buffer.getvalue().rstrip()}"

    def _serialize_markdown(self, ws) -> str:
        rows, bounds = self._matrix(ws, include_coordinates=True)
        escaped = [[self._escape_markdown_cell(cell) for cell in row] for row in rows]
        widths = [max(len(row[idx]) for row in escaped) for idx in range(len(escaped[0]))]

        def fmt(row):
            return "| " + " | ".join(cell.ljust(widths[idx]) for idx, cell in enumerate(row)) + " |"

        header = fmt(escaped[0])
        separator = "| " + " | ".join("-" * max(3, width) for width in widths) + " |"
        body = "\n".join(fmt(row) for row in escaped[1:])
        return f"Used Range: {self._bounds_to_range(bounds)}\n{header}\n{separator}\n{body}"

    def _serialize_fixed_width(self, ws) -> str:
        rows, bounds = self._matrix(ws)
        widths = [max(len(row[idx]) for row in rows) for idx in range(len(rows[0]))]
        lines = []
        for row in rows:
            lines.append("  ".join(row[idx].ljust(widths[idx]) for idx in range(len(row))))
        return f"Used Range: {self._bounds_to_range(bounds)}\n" + "\n".join(lines)

    def _serialize_json_rows(self, ws) -> str:
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        merged_values = self._merged_value_map(ws)
        columns = [get_column_letter(col) for col in range(min_col, max_col + 1)]
        rows = []
        for row_idx in range(min_row, max_row + 1):
            values = []
            for col_idx in range(min_col, max_col + 1):
                value = merged_values.get((row_idx, col_idx), ws.cell(row_idx, col_idx).value)
                values.append(_value_to_json(value))
            rows.append({"row": row_idx, "values": values})

        payload = {
            "sheet_name": ws.title,
            "used_range": self._bounds_to_range((min_row, min_col, max_row, max_col)),
            "columns": columns,
            "rows": rows,
            "merged_ranges": self._merged_ranges(ws),
        }
        return json.dumps(payload, ensure_ascii=False)

    def _serialize_json_cells(self, ws) -> str:
        min_row, min_col, max_row, max_col = self._used_bounds(ws)
        cells = []
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                value = ws.cell(row_idx, col_idx).value
                if value in (None, ""):
                    continue
                cells.append(
                    {
                        "cell": f"{get_column_letter(col_idx)}{row_idx}",
                        "row": row_idx,
                        "column": get_column_letter(col_idx),
                        "value": _value_to_json(value),
                    }
                )

        payload = {
            "sheet_name": ws.title,
            "used_range": self._bounds_to_range((min_row, min_col, max_row, max_col)),
            "cells": cells,
            "merged_ranges": self._merged_ranges(ws),
        }
        return json.dumps(payload, ensure_ascii=False)

    def _serialize_html(self, ws) -> str:
        rows, bounds = self._matrix(ws, include_coordinates=True)
        lines = [
            f"<table data-sheet={json.dumps(ws.title)} data-used-range={json.dumps(self._bounds_to_range(bounds))}>"
        ]
        for row in rows:
            lines.append("<tr>" + "".join(f"<td>{html.escape(cell)}</td>" for cell in row) + "</tr>")
        lines.append("</table>")
        return "\n".join(lines)

    def _merged_ranges(self, ws) -> List[Dict[str, Any]]:
        ranges = []
        for merged_range in ws.merged_cells.ranges:
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            ranges.append({"range": str(merged_range), "value": _value_to_json(value)})
        return ranges

    def _bounds_to_range(self, bounds: Tuple[int, int, int, int]) -> str:
        min_row, min_col, max_row, max_col = bounds
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def _escape_markdown_cell(self, cell: str) -> str:
        return str(cell).replace("\n", " ").replace("|", "\\|")

    def _maybe_truncate(self, text: str) -> Tuple[str, bool, int]:
        text_tokens = calculate_token_cost_line(text, self.token_model)
        if not self.max_text_tokens or self.max_text_tokens <= 0 or text_tokens <= self.max_text_tokens:
            return text, False, text_tokens

        marker = "\n\n...[truncated to fit token budget]...\n\n"
        marker_tokens = calculate_token_cost_line(marker, self.token_model)
        if self.max_text_tokens <= marker_tokens + 2:
            truncated = _truncate_text_to_token_budget(text, self.max_text_tokens, self.token_model)
            return truncated, True, calculate_token_cost_line(truncated, self.token_model)

        body_budget = self.max_text_tokens - marker_tokens
        head_budget = body_budget // 2
        tail_budget = body_budget - head_budget
        truncated = (
            _truncate_text_to_token_budget(text, head_budget, self.token_model)
            + marker
            + _truncate_text_to_token_budget(text, tail_budget, self.token_model, from_end=True)
        )
        return truncated, True, calculate_token_cost_line(truncated, self.token_model)


class SpreadSheetPoTSolver:
    def __init__(self, *args, **kwargs):
        self.url = kwargs.get("url", "localhost:8000")
        self.model_name = kwargs.get("model_name", None)
        self.code_exec_url = kwargs.get("code_exec_url", "localhost:8081")
        self.model_params = {
            "top_p": kwargs.get("top_p", 1.0),
            "temperature": kwargs.get("temperature", 0),
        }
        self.table_format = kwargs.get("table_format", "markdown")
        self.include_coordinates = kwargs.get("include_coordinates", True)
        self.fill_merged = kwargs.get("fill_merged", False)
        self.max_text_tokens = kwargs.get("max_text_tokens", 0)
        self.max_retries = kwargs.get("max_retries", 3)
        self.render_formulas_before_eval = kwargs.get("render_formulas_before_eval", False)
        self.save_prompts = kwargs.get("save_prompts", False)
        self.dry_run = kwargs.get("dry_run", False)
        self.output_dir = kwargs.get("output_dir")
        self.data_split = kwargs.get("data_split", "all_912")
        self.split_config = SPREADSHEET_DATA_SPLITS.get(self.data_split, SPREADSHEET_DATA_SPLITS["all_912"])

    def _builder(self, data: Dict[str, Any]) -> SpreadsheetTableInputBuilder:
        return SpreadsheetTableInputBuilder(
            real_dir=data["real_dir"],
            latex_dir=data.get("latex_dir"),
            image_dir=data.get("image_dir"),
            excel_1_image_dir=data.get("excel_1_image_dir"),
            table_format=self.table_format,
            include_coordinates=self.include_coordinates,
            fill_merged=self.fill_merged,
            max_text_tokens=self.max_text_tokens,
            token_model=self.model_name,
            image_cache_dir=data.get("image_cache_dir"),
        )

    def build_prompt(self, data: Dict[str, Any]) -> Tuple[Union[str, List[Dict[str, Any]]], Dict[str, Any]]:
        table_text, table_metadata = self._builder(data).build(data["input_file"])
        prompt = FIRST_ROUND_PROMPT.format(
            instruction=data.get("instruction", ""),
            spreadsheet_path=data["input_path"],
            spreadsheet_content_note=_spreadsheet_content_note(table_metadata["table_format"]),
            spreadsheet_content=table_text,
            instruction_type=data.get("instruction_type", ""),
            answer_position=data.get("answer_position", ""),
            answer_sheet=data.get("answer_sheet", ""),
            output_path=data["output_path"],
        )
        image_paths = table_metadata.get("image_paths") or []
        if image_paths:
            content: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
            for image_path in image_paths:
                content.append({"type": "image_url", "image_url": {"url": _image_to_data_url(image_path)}})
            return content, table_metadata
        return prompt, table_metadata

    def get_solution(self, data: Dict[str, Any], client: ClientJupyterKernel) -> Dict[str, Any]:
        result = copy.deepcopy(data)
        messages: List[Dict[str, Any]] = []
        attempts: List[Dict[str, Any]] = []
        solution = ""
        table_metadata: Dict[str, Any] = {}
        error = None
        success = False

        try:
            prompt, table_metadata = self.build_prompt(data)
            messages = [{"role": "user", "content": prompt}]

            if self.dry_run:
                raise RuntimeError("dry_run enabled; prompt was built but model was not called.")

            for attempt_idx in range(1, self.max_retries + 1):
                content = ""
                code = ""
                execution_output = ""
                attempt_error = None
                output_created = False

                try:
                    resp = model_resp(
                        self.url,
                        messages,
                        model_params=self.model_params,
                        model_name=self.model_name,
                    )
                    if not resp or not resp.get("message"):
                        raise RuntimeError("Model API returned no message.")
                    content = resp["message"].get("content") or ""
                    messages.append({"role": "assistant", "content": content})

                    code = extract_code(content).strip()
                    if not code:
                        raise RuntimeError("No Python code was extracted from the model response.")

                    client.execute(REMOVE_OUTPUT_FILE.format(output_path=data["output_path"]))
                    execution_output = client.execute(code)
                    output_created = self._output_created(client, data["output_path"])
                    if self._execution_failed(execution_output) or not output_created:
                        details = execution_output
                        if not output_created:
                            details = f"{details}\nExpected output file was not created: {data['output_path']}"
                        raise RuntimeError(details)

                    success = True
                    error = None
                    solution = code
                except Exception as exc:
                    attempt_error = str(exc)
                    error = attempt_error
                    if code:
                        solution = code

                attempts.append(
                    {
                        "attempt": attempt_idx,
                        "success": success,
                        "response": content,
                        "code": code,
                        "execution_output": execution_output,
                        "output_created": output_created,
                        "error": attempt_error,
                    }
                )

                if success:
                    break

                if attempt_idx < self.max_retries:
                    feedback = self._feedback_for_retry(attempts[-1])
                    messages.append(
                        {
                            "role": "user",
                            "content": RETRY_PROMPT.format(
                                attempt=attempt_idx,
                                feedback=feedback,
                                input_path=data["input_path"],
                                output_path=data["output_path"],
                            ),
                        }
                    )
        except Exception:
            error = traceback.format_exc()

        result.update(
            {
                "table_metadata": table_metadata,
                "attempts": attempts,
                "execution_success": success,
                "format_valid": bool(success),
                "error": error,
                "solution": solution,
            }
        )
        if self.save_prompts:
            result["messages"] = messages
        return result

    def _output_created(self, client: ClientJupyterKernel, output_path: str) -> bool:
        check_output = client.execute(FINAL_CODE_CHECK.format(output_path=output_path))
        return "OUTPUT_FILE_CREATED" in check_output and not self._execution_failed(check_output)

    def _execution_failed(self, output: str) -> bool:
        if not output:
            return False
        failure_markers = [
            "Traceback (most recent call last)",
            "SyntaxError:",
            "NameError:",
            "TypeError:",
            "ValueError:",
            "KeyError:",
            "FileNotFoundError:",
            "PermissionError:",
            "ModuleNotFoundError:",
            "ImportError:",
            "AssertionError:",
            "[Execution timed out",
        ]
        return any(marker in output for marker in failure_markers)

    def _feedback_for_retry(self, attempt: Dict[str, Any]) -> str:
        pieces = []
        if attempt.get("error"):
            pieces.append(f"Error: {attempt['error']}")
        if attempt.get("execution_output"):
            pieces.append(f"Execution output:\n{attempt['execution_output']}")
        if attempt.get("code"):
            pieces.append(f"Previous code:\n```python\n{attempt['code']}\n```")
        if not pieces:
            pieces.append("No executable Python code was produced.")
        return "\n\n".join(pieces)

    def _render_formulas_for_eval(self, output_file: str) -> None:
        if not self.render_formulas_before_eval or not os.path.exists(output_file):
            return
        # 按文件路径直接加载evaluate/open_spreadsheet.py
        import importlib.util

        repo_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        open_spreadsheet_path = os.path.join(repo_dir, "evaluate", "open_spreadsheet.py")
        spec = importlib.util.spec_from_file_location("_local_open_spreadsheet", open_spreadsheet_path)
        if spec is None or spec.loader is None:
            raise RuntimeError(f"Cannot load open_spreadsheet.py from {open_spreadsheet_path}")
        open_spreadsheet = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(open_spreadsheet)

        soffice_path = open_spreadsheet.find_libreoffice()
        if not soffice_path:
            raise RuntimeError("LibreOffice not found; cannot render formulas before evaluation.")
        if not open_spreadsheet.just_open_libreoffice(output_file, soffice_path):
            raise RuntimeError(f"Failed to render formulas before evaluation: {output_file}")

    def evaluate(self, data: Dict[str, Any], client: ClientJupyterKernel) -> Dict[str, Any]:
        result = copy.deepcopy(data)
        result["test_case_results"] = []
        result["test_case_messages"] = []

        solution = result.get("solution") or ""
        input_suffix = self.split_config["input_suffix"]
        answer_suffix = self.split_config["answer_suffix"]
        num_test_cases = int(self.split_config["num_test_cases"])

        for idx in range(1, num_test_cases + 1):
            try:
                input_file = f"{idx}_{result['id']}_{input_suffix}.xlsx"
                output_file = f"{idx}_{result['id']}_output.xlsx"
                output_path = f"/mnt/data/output/{output_file}"
                local_output_file = os.path.join(self.output_dir, "spreadsheet", output_file)

                should_execute = bool(solution)
                if idx == 1 and os.path.exists(local_output_file):
                    should_execute = False

                if should_execute:
                    local_solution = solution.replace(result["input_file"], input_file)
                    local_solution = local_solution.replace(f"1_{result['id']}_output.xlsx", output_file)
                    local_solution = local_solution.replace(result["input_path"], f"/mnt/data/input/{input_file}")
                    local_solution = local_solution.replace(result["output_path"], output_path)
                    client.execute(REMOVE_OUTPUT_FILE.format(output_path=output_path))
                    exec_output = client.execute(local_solution)
                    if self._execution_failed(exec_output):
                        raise RuntimeError(exec_output)
                    
                # （可选）通过render_formulas_before_eval参数控制评测前是否渲染公式，提升评测的准确性，但会增加评测时间
                self._render_formulas_for_eval(local_output_file)
                gt_file = os.path.join(result["real_dir"], f"{idx}_{result['id']}_{answer_suffix}.xlsx")
                passed, message = compare_workbooks(
                    gt_file,
                    local_output_file,
                    result.get("instruction_type", ""),
                    result.get("answer_position", ""),
                )
            except Exception:
                passed, message = 0, traceback.format_exc()
            result["test_case_results"].append(int(passed))
            result["test_case_messages"].append(message)

        result["total_soft_restriction"] = sum(result["test_case_results"]) / len(result["test_case_results"])
        result["total_hard_restriction"] = 1.0 if all(result["test_case_results"]) else 0.0
        return result

    def __call__(self, data: Dict[str, Any]) -> Dict[str, Any]:
        try:
            client = ClientJupyterKernel(self.code_exec_url, data["mount_dir"])
            result = self.get_solution(data, client)
            result = self.evaluate(result, client)
            return result
        except Exception:
            err = traceback.format_exc()
            n = int(self.split_config.get("num_test_cases", 3))
            out = copy.deepcopy(data)
            out.update(
                {
                    "error": err,
                    "execution_success": False,
                    "format_valid": False,
                    "test_case_results": [0] * n,
                    "test_case_messages": [err] * n,
                    "total_soft_restriction": 0.0,
                    "total_hard_restriction": 0.0,
                }
            )
            return out


SpreadsheetPoTSolver = SpreadSheetPoTSolver
