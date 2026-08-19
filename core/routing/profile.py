import glob
import os
import re
import traceback
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter


STYLE_TERMS = (
    "color",
    "colour",
    "background",
    "fill",
    "font",
    "bold",
    "border",
    "highlight",
    "shade",
    "style",
    "format",
    "conditional formatting",
)
STRUCTURE_TERMS = (
    "header",
    "merged",
    "merge",
    "hierarchy",
    "category",
    "section",
    "layout",
    "structure",
    "under which",
    "row header",
    "column header",
)
FORMULA_TERMS = ("formula", "equation", "computed")
SORT_FILTER_TERMS = (
    "sort",
    "filter",
    "pivot",
    "rank",
    "ascending",
    "descending",
    "lowest",
    "highest",
    "order",
)
INSERT_DELETE_TERMS = (
    "insert",
    "delete",
    "remove",
    "clear",
    "create",
    "add row",
    "add column",
)
AGGREGATION_TERMS = (
    "sum",
    "total",
    "average",
    "mean",
    "median",
    "max",
    "min",
    "count",
    "calculate",
    "computed",
)
JOIN_TERMS = (
    "match",
    "lookup",
    "duplicate",
    "merge",
    "combine",
    "join",
    "cross sheet",
    "another sheet",
    "matching",
)
DATE_TERMS = ("date", "month", "year", "day", "time")


def unique_keep_order(values: List[Optional[str]]) -> List[str]:
    seen = set()
    out = []
    for value in values:
        if not value or value in seen:
            continue
        out.append(value)
        seen.add(value)
    return out


def text_contains_any(text: str, terms: Tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(term.lower() in lowered for term in terms)


def estimate_tokens(text: str) -> int:
    return max(1, int(len(text) / 3.5))


def estimate_tokens_from_chars(num_chars: int) -> int:
    return max(1, int(num_chars / 3.5))


def color_key(color: Any) -> Optional[str]:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb and str(rgb).upper() not in {"00000000", "000000", "FFFFFFFF"}:
        return f"rgb:{rgb}"
    indexed = getattr(color, "indexed", None)
    if indexed not in (None, 64):
        return f"indexed:{indexed}"
    theme = getattr(color, "theme", None)
    tint = getattr(color, "tint", 0)
    if theme is not None:
        return f"theme:{theme}:{tint}"
    return None


def used_bounds(ws) -> Tuple[int, int, int, int]:
    min_row = min_col = None
    max_row = max_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                min_row = cell.row if min_row is None else min(min_row, cell.row)
                min_col = cell.column if min_col is None else min(min_col, cell.column)
                max_row = cell.row if max_row is None else max(max_row, cell.row)
                max_col = cell.column if max_col is None else max(max_col, cell.column)
    if min_row is None:
        return 1, 1, 1, 1
    return min_row, min_col, max_row, max_col


def bounds_to_range(bounds: Tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def has_visible_border(cell) -> bool:
    border = cell.border
    return any(getattr(side, "style", None) for side in (border.left, border.right, border.top, border.bottom))


def profile_workbook(xlsx_path: str, token_hint_path: Optional[str] = None) -> Dict[str, Any]:
    profile: Dict[str, Any] = {
        "xlsx_path": xlsx_path,
        "xlsx_exists": os.path.exists(xlsx_path),
        "sheet_summaries": [],
    }
    if not os.path.exists(xlsx_path):
        profile["profile_error"] = f"Workbook not found: {xlsx_path}"
        return profile

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
    except Exception:
        profile["profile_error"] = traceback.format_exc()
        return profile

    total_nonempty = 0
    total_used_cells = 0
    total_text_chars = 0
    total_formulas = 0
    total_merged = 0
    total_bold = 0
    total_borders = 0
    total_hidden_rows = 0
    total_hidden_cols = 0
    total_charts = 0
    total_embedded_images = 0
    numeric_cells = 0
    date_like_cells = 0
    long_text_cells = 0
    distinct_fills = set()

    try:
        for ws in wb.worksheets:
            bounds = used_bounds(ws)
            min_row, min_col, max_row, max_col = bounds
            used_rows = max_row - min_row + 1
            used_cols = max_col - min_col + 1
            used_cells = used_rows * used_cols
            sheet_nonempty = 0
            sheet_formulas = 0
            sheet_text_chars = 0
            sheet_bold = 0
            sheet_borders = 0
            sheet_fills = set()

            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if value in (None, ""):
                        continue
                    sheet_nonempty += 1
                    value_text = str(value)
                    sheet_text_chars += len(value_text)
                    if isinstance(value, str) and value.startswith("="):
                        sheet_formulas += 1
                    if isinstance(value, (int, float)):
                        numeric_cells += 1
                    if hasattr(value, "isoformat") or text_contains_any(value_text, DATE_TERMS):
                        date_like_cells += 1
                    if isinstance(value, str) and len(value) > 60:
                        long_text_cells += 1
                    if getattr(cell.font, "bold", False):
                        sheet_bold += 1
                    if has_visible_border(cell):
                        sheet_borders += 1
                    fill = cell.fill
                    fill_key = color_key(getattr(fill, "fgColor", None)) if fill and fill.fill_type else None
                    if fill_key:
                        sheet_fills.add(fill_key)

            hidden_rows = sum(1 for dim in ws.row_dimensions.values() if getattr(dim, "hidden", False))
            hidden_cols = sum(1 for dim in ws.column_dimensions.values() if getattr(dim, "hidden", False))
            charts = len(getattr(ws, "_charts", []) or [])
            embedded_images = len(getattr(ws, "_images", []) or [])
            merged_ranges = len(ws.merged_cells.ranges)

            total_nonempty += sheet_nonempty
            total_used_cells += used_cells
            total_text_chars += sheet_text_chars
            total_formulas += sheet_formulas
            total_merged += merged_ranges
            total_bold += sheet_bold
            total_borders += sheet_borders
            total_hidden_rows += hidden_rows
            total_hidden_cols += hidden_cols
            total_charts += charts
            total_embedded_images += embedded_images
            distinct_fills.update(sheet_fills)

            profile["sheet_summaries"].append(
                {
                    "sheet_name": ws.title,
                    "used_range": bounds_to_range(bounds),
                    "used_rows": used_rows,
                    "used_cols": used_cols,
                    "used_cells": used_cells,
                    "nonempty_cells": sheet_nonempty,
                    "merged_ranges": merged_ranges,
                    "formula_cells": sheet_formulas,
                    "distinct_fill_colors": len(sheet_fills),
                    "bold_cells": sheet_bold,
                    "bordered_cells": sheet_borders,
                    "hidden_rows": hidden_rows,
                    "hidden_cols": hidden_cols,
                    "charts": charts,
                    "embedded_images": embedded_images,
                }
            )
    finally:
        wb.close()

    if token_hint_path and os.path.exists(token_hint_path):
        try:
            total_text_chars = max(total_text_chars, os.path.getsize(token_hint_path))
        except OSError:
            pass

    total_used_cells = total_used_cells or 0
    total_nonempty = total_nonempty or 0
    profile.update(
        {
            "num_sheets": len(profile["sheet_summaries"]),
            "sheet_names": [item["sheet_name"] for item in profile["sheet_summaries"]],
            "max_rows": max((item["used_rows"] for item in profile["sheet_summaries"]), default=0),
            "max_cols": max((item["used_cols"] for item in profile["sheet_summaries"]), default=0),
            "max_sheet_cells": max((item["used_cells"] for item in profile["sheet_summaries"]), default=0),
            "total_used_cells": total_used_cells,
            "total_nonempty_cells": total_nonempty,
            "nonempty_ratio": round(total_nonempty / total_used_cells, 4) if total_used_cells else 0.0,
            "total_text_chars": total_text_chars,
            "estimated_text_tokens": estimate_tokens_from_chars(total_text_chars),
            "num_merged_ranges": total_merged,
            "merged_cell_signal": total_merged > 0,
            "num_formulas": total_formulas,
            "formula_ratio": round(total_formulas / total_nonempty, 4) if total_nonempty else 0.0,
            "num_distinct_fill_colors": len(distinct_fills),
            "has_background_color": bool(distinct_fills),
            "num_bold_cells": total_bold,
            "num_bordered_cells": total_borders,
            "has_hidden_rows_or_cols": bool(total_hidden_rows or total_hidden_cols),
            "hidden_rows": total_hidden_rows,
            "hidden_cols": total_hidden_cols,
            "has_charts_or_images": bool(total_charts or total_embedded_images),
            "num_charts": total_charts,
            "num_embedded_images": total_embedded_images,
            "numeric_cell_ratio": round(numeric_cells / total_nonempty, 4) if total_nonempty else 0.0,
            "date_cell_ratio": round(date_like_cells / total_nonempty, 4) if total_nonempty else 0.0,
            "long_text_cell_ratio": round(long_text_cells / total_nonempty, 4) if total_nonempty else 0.0,
        }
    )
    return profile


def find_realhit_images(image_dir: Optional[str], file_name: str) -> List[str]:
    if not image_dir or not os.path.isdir(image_dir):
        return []
    candidates = []
    for ext in (".png", ".jpg", ".jpeg", ".webp"):
        candidates.append(os.path.join(image_dir, f"{file_name}{ext}"))
        candidates.extend(glob.glob(os.path.join(image_dir, f"{glob.escape(file_name)}___*{ext}")))
    return sorted(path for path in candidates if os.path.exists(path))


def find_spreadsheet_images(image_dir: Optional[str], input_file: str) -> List[str]:
    if not image_dir or not os.path.isdir(image_dir):
        return []
    stem = os.path.splitext(input_file)[0]

    def find_in_dir(search_dir: str) -> List[str]:
        candidates = []
        for ext in (".png", ".jpg", ".jpeg", ".webp"):
            candidates.extend([os.path.join(search_dir, f"{stem}{ext}"), os.path.join(search_dir, f"{input_file}{ext}")])
            candidates.extend(glob.glob(os.path.join(search_dir, f"{glob.escape(stem)}___*{ext}")))
            candidates.extend(glob.glob(os.path.join(search_dir, f"{glob.escape(input_file)}___*{ext}")))

        def sheet_index(path: str) -> Tuple[int, str]:
            match = re.search(r"___(\d+)$", os.path.splitext(os.path.basename(path))[0])
            return (int(match.group(1)) if match else 0, path)

        return sorted([path for path in candidates if os.path.exists(path)], key=sheet_index)

    found = find_in_dir(image_dir)
    if found:
        return found

    parent_dir = os.path.dirname(os.path.normpath(image_dir))
    if parent_dir and parent_dir != os.path.normpath(image_dir):
        return find_in_dir(parent_dir)
    return []


def query_features(text: str, task_type: Optional[str] = None) -> Dict[str, Any]:
    text = text or ""
    lowered = text.lower()
    exact_range = bool(re.search(r"\b[A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?\b", text))
    quoted_sheet = bool(re.search(r"['\"][^'\"]+['\"]\s*!", text))
    sheet_word = "sheet" in lowered or "worksheet" in lowered
    and_count = len(re.findall(r"\band\b|,", lowered))

    return {
        "task_type": task_type,
        "query_chars": len(text),
        "query_tokens_est": estimate_tokens(text),
        "mentions_color_or_style": text_contains_any(text, STYLE_TERMS),
        "mentions_header_or_structure": text_contains_any(text, STRUCTURE_TERMS),
        "mentions_formula": text_contains_any(text, FORMULA_TERMS),
        "mentions_sort_filter_pivot": text_contains_any(text, SORT_FILTER_TERMS),
        "mentions_insert_delete": text_contains_any(text, INSERT_DELETE_TERMS),
        "mentions_cross_sheet": sheet_word or text_contains_any(text, JOIN_TERMS),
        "mentions_total_sum_average_rank": text_contains_any(text, AGGREGATION_TERMS),
        "mentions_date_time": text_contains_any(text, DATE_TERMS),
        "mentions_exact_cell_or_range": exact_range,
        "mentions_sheet_name": quoted_sheet or sheet_word,
        "num_operations_in_instruction": and_count
        + sum(
            1
            for terms in (SORT_FILTER_TERMS, INSERT_DELETE_TERMS, AGGREGATION_TERMS, JOIN_TERMS, STYLE_TERMS)
            if text_contains_any(text, terms)
        ),
        "requires_lookup": text_contains_any(text, JOIN_TERMS),
        "requires_aggregation": text_contains_any(text, AGGREGATION_TERMS),
        "requires_comparison": any(word in lowered for word in ("highest", "lowest", "largest", "smallest", "greater", "less")),
        "requires_formatting": text_contains_any(text, STYLE_TERMS),
    }


class SpreadsheetProfiler:
    def __init__(self, args):
        self.args = args

    def profile(self, dataset: str, item: Dict[str, Any]) -> Dict[str, Any]:
        if dataset == "realhitbench":
            return self._profile_realhit(item)
        if dataset == "spreadsheetbench":
            return self._profile_spreadsheet(item)
        raise ValueError(f"Unsupported dataset: {dataset}")

    def _profile_realhit(self, item: Dict[str, Any]) -> Dict[str, Any]:
        file_name = item["FileName"]
        xlsx_path = os.path.join(item["real_dir"], f"{file_name}.xlsx")
        latex_path = os.path.join(item["latex_dir"], f"{file_name}.txt") if item.get("latex_dir") else None
        profile = profile_workbook(xlsx_path, token_hint_path=latex_path)
        raw_images = find_realhit_images(item.get("image_dir"), file_name)
        excel_images = find_realhit_images(item.get("excel_1_image_dir"), file_name)
        available_text = ["markdown", "html", "csv", "tsv", "dataframe", "json_rows", "json_cells"]
        if latex_path and os.path.exists(latex_path):
            available_text.insert(0, "latex")

        profile.update(
            {
                "dataset": "realhitbench",
                "sample_id": item.get("id"),
                "file_name": file_name,
                "question_type": item.get("QuestionType"),
                "sub_question_type": item.get("SubQType"),
                "complex_structure": item.get("CompStrucCata"),
                "question_features": query_features(item.get("Question", ""), item.get("QuestionType")),
                "available_text_formats": unique_keep_order(available_text),
                "available_image_formats": unique_keep_order(
                    [
                        "image" if raw_images else None,
                        "excel_1_image" if excel_images else None,
                        "default_image" if profile.get("xlsx_exists") else None,
                    ]
                ),
                "image_counts": {
                    "image": len(raw_images),
                    "excel_1_image": len(excel_images),
                    "default_image": int(profile.get("num_sheets") or 0),
                },
                "latex_available": bool(latex_path and os.path.exists(latex_path)),
            }
        )
        profile["truncation_risk"] = bool(
            self.args.max_text_tokens and profile.get("estimated_text_tokens", 0) > self.args.max_text_tokens
        )
        return profile

    def _profile_spreadsheet(self, item: Dict[str, Any]) -> Dict[str, Any]:
        input_file = item["input_file"]
        xlsx_path = os.path.join(item["real_dir"], input_file)
        profile = profile_workbook(xlsx_path)
        raw_images = find_spreadsheet_images(item.get("image_dir"), input_file)
        excel_images = find_spreadsheet_images(item.get("excel_1_image_dir"), input_file)

        profile.update(
            {
                "dataset": "spreadsheetbench",
                "sample_id": item.get("id"),
                "input_file": input_file,
                "instruction_type": item.get("instruction_type"),
                "answer_position": item.get("answer_position"),
                "answer_sheet": item.get("answer_sheet"),
                "question_features": query_features(item.get("instruction", ""), item.get("instruction_type")),
                "available_text_formats": ["markdown", "html", "csv", "tsv", "dataframe", "json_rows", "json_cells", "latex"],
                "available_image_formats": unique_keep_order(
                    [
                        "image" if raw_images else None,
                        "excel_1_image" if excel_images else None,
                        "default_image" if profile.get("xlsx_exists") else None,
                    ]
                ),
                "image_counts": {
                    "image": len(raw_images),
                    "excel_1_image": len(excel_images),
                    "default_image": int(profile.get("num_sheets") or 0),
                },
            }
        )
        profile["truncation_risk"] = bool(
            self.args.max_text_tokens and profile.get("estimated_text_tokens", 0) > self.args.max_text_tokens
        )
        profile["multi_image_risk"] = bool((profile.get("num_sheets") or 0) >= 3)
        return profile
