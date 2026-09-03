"""Gold-free SheetFlex-vote medoid selection for SpreadsheetBench verified_400."""

import datetime
import hashlib
import json
import os
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Dict, Mapping, Sequence

import openpyxl

from ..eval.spreadsheet import compare_workbooks
from ..eval.spreadsheet_regions import (
    SpreadsheetRegionError,
    compare_cell_value,
    extract_normalized_region_cells,
)
from .common import (
    FORMAT_ORDER,
    break_tie,
    count_distribution,
    logprob_summary,
    max_score_items,
    numeric_summary,
    safe_rate,
    validate_tie_break_logprob_field,
)


def _workbook_bounds(path: Path) -> tuple[dict, dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        max_rows = {sheet.title: sheet.max_row for sheet in workbook.worksheets}
        max_columns = {sheet.title: sheet.max_column for sheet in workbook.worksheets}
        return max_rows, max_columns
    finally:
        workbook.close()


def determine_shared_bounds(
    input_path: Path, candidate_paths: Mapping[str, Path]
) -> tuple[dict, dict, dict]:
    """Compute finite bounds from input plus every openable candidate, never gold."""
    max_rows: Dict[str, int] = {}
    max_columns: Dict[str, int] = {}
    errors = {}
    sources = {"input": input_path, **candidate_paths}
    for source_name, path in sources.items():
        if not path.is_file():
            errors[source_name] = f"workbook_file_not_found: {path}"
            continue
        try:
            source_rows, source_columns = _workbook_bounds(path)
        except Exception as exc:
            errors[source_name] = f"failed_to_open_workbook: {exc}"
            continue
        for sheet_name, value in source_rows.items():
            max_rows[sheet_name] = max(max_rows.get(sheet_name, 1), value)
        for sheet_name, value in source_columns.items():
            max_columns[sheet_name] = max(max_columns.get(sheet_name, 1), value)
    return max_rows, max_columns, errors


def _stable_value(value: Any):
    if value is None or value == "":
        return ["blank"]
    if isinstance(value, (str, int, float, bool)):
        return [type(value).__name__, value]
    if isinstance(value, (datetime.date, datetime.time)):
        return [type(value).__name__, value.isoformat()]
    return [type(value).__name__, repr(value)]


def normalized_region_hash(cells: Mapping[tuple[str, str], Any]) -> str:
    payload = [
        [sheet_name, coordinate, _stable_value(value)]
        for (sheet_name, coordinate), value in cells.items()
    ]
    serialized = json.dumps(
        payload, ensure_ascii=False, separators=(",", ":"), allow_nan=True
    ).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def cell_map_similarity(
    left: Mapping[tuple[str, str], Any],
    right: Mapping[tuple[str, str], Any],
) -> float:
    if list(left) != list(right):
        raise SpreadsheetRegionError("Candidate region coordinate mappings differ")
    if not left:
        raise SpreadsheetRegionError("Candidate region mapping is empty")
    matched = sum(
        1 for key in left if compare_cell_value(left[key], right[key])
    )
    return matched / len(left)


def select_spreadsheet_medoid(
    candidates: Sequence[Mapping[str, Any]],
    *,
    candidate_id_key: str = "format",
    selected_id_field: str = "selected_format",
    rank_getter: Callable[[Mapping[str, Any]], Any] | None = None,
    fallback_source: str = "format_order",
    logprob_field: str | None = "sequence_logprob_mean",
) -> Dict[str, Any]:
    validate_tie_break_logprob_field(logprob_field)
    candidates = [dict(candidate) for candidate in candidates]
    if rank_getter is None:
        rank_getter = lambda item: FORMAT_ORDER.index(item[candidate_id_key])
    valid = [candidate for candidate in candidates if candidate["valid"]]
    matrix = {
        left[candidate_id_key]: {
            right[candidate_id_key]: None for right in candidates
        }
        for left in candidates
    }
    pairwise_values = []
    if not valid:
        for candidate in candidates:
            candidate.pop("_cells", None)
        return {
            "format_valid": False,
            "valid_candidate_count": 0,
            "candidates": candidates,
            "similarity_matrix": matrix,
            "average_pairwise_region_agreement": None,
            "medoid_score": None,
            selected_id_field: None,
            "selected_source_file": None,
            "tie": False,
            "tied_candidate_count": 0,
            "tie_break_source": "not_applicable",
            "tie_break_reason": "all_candidates_invalid",
            "tie_break_logprob_field": logprob_field,
        }

    for left_index, left in enumerate(valid):
        for right_index, right in enumerate(valid):
            if right_index < left_index:
                similarity = matrix[right[candidate_id_key]][left[candidate_id_key]]
            else:
                similarity = cell_map_similarity(left["_cells"], right["_cells"])
            matrix[left[candidate_id_key]][right[candidate_id_key]] = similarity
            if right_index > left_index:
                pairwise_values.append(similarity)

    for candidate in valid:
        candidate["aggregation_score"] = sum(
            matrix[candidate[candidate_id_key]][other[candidate_id_key]]
            for other in valid
        )

    best_score, tied = max_score_items(valid)
    medoid_tie = len(tied) > 1
    if medoid_tie:
        decision = break_tie(
            tied,
            rank_getter=rank_getter,
            fallback_source=fallback_source,
            logprob_field=logprob_field,
        )
        selected = decision.selected
        tie_source = decision.source
        tie_reason = decision.reason
    else:
        selected = tied[0]
        tie_source = "not_needed"
        tie_reason = "unique_highest_medoid_score"
    selected["selected"] = True

    for candidate in candidates:
        candidate.pop("_cells", None)
    return {
        "format_valid": True,
        "valid_candidate_count": len(valid),
        "candidates": candidates,
        "similarity_matrix": matrix,
        "average_pairwise_region_agreement": (
            sum(pairwise_values) / len(pairwise_values) if pairwise_values else None
        ),
        "medoid_score": best_score,
        selected_id_field: selected[candidate_id_key],
        "selected_source_file": selected["output_file"],
        "tie": medoid_tie,
        "tied_candidate_count": len(tied),
        "tie_break_source": tie_source,
        "tie_break_reason": tie_reason,
        "tie_break_logprob_field": logprob_field,
    }


def candidate_output_path(
    sample_id: str,
    format_name: str,
    record: Mapping[str, Any] | None,
    run_dirs: Mapping[str, Path],
) -> Path:
    output_name = f"1_{sample_id}_output.xlsx"
    if record and record.get("output_file"):
        candidate_name = Path(str(record["output_file"])).name
        if candidate_name.lower().endswith(".xlsx"):
            output_name = candidate_name
    return run_dirs[format_name] / "spreadsheet" / output_name


def aggregate_spreadsheet_candidates(
    item: Mapping[str, Any],
    candidate_specs: Sequence[Mapping[str, Any]],
    input_path: Path,
    *,
    candidate_id_key: str = "format",
    selected_id_field: str = "selected_format",
    rank_getter: Callable[[Mapping[str, Any]], Any] | None = None,
    fallback_source: str = "format_order",
    logprob_field: str | None = "sequence_logprob_mean",
) -> Dict[str, Any]:
    """Validate and select candidate workbooks without reading a golden file."""
    sample_id = str(item["id"])
    candidate_ids = [spec[candidate_id_key] for spec in candidate_specs]
    if len(set(candidate_ids)) != len(candidate_ids):
        raise SpreadsheetRegionError(
            f"Duplicate candidate identifier for sample {sample_id}: {candidate_ids}"
        )
    candidate_paths = {
        spec[candidate_id_key]: Path(spec["output_path"])
        for spec in candidate_specs
    }
    max_rows, max_columns, open_errors = determine_shared_bounds(
        input_path, candidate_paths
    )

    input_cells = None
    input_error = open_errors.get("input")
    if input_error is None:
        try:
            input_cells = extract_normalized_region_cells(
                input_path,
                item.get("answer_position"),
                item.get("answer_sheet", ""),
                max_rows=max_rows,
                max_columns=max_columns,
            )
        except Exception as exc:
            input_error = f"failed_to_extract_input_regions: {exc}"

    candidates = []
    for spec in candidate_specs:
        candidate_id = spec[candidate_id_key]
        record = spec.get("record")
        output_path = candidate_paths[candidate_id]
        valid = True
        invalid_reason = None
        cells = None
        if record is None:
            valid = False
            invalid_reason = "sample_id_missing_from_run_results"
        elif record.get("execution_success") is not True:
            valid = False
            invalid_reason = "execution_success_is_not_true"
        elif candidate_id in open_errors:
            valid = False
            invalid_reason = open_errors[candidate_id]
        elif input_error is not None:
            valid = False
            invalid_reason = input_error
        else:
            try:
                cells = extract_normalized_region_cells(
                    output_path,
                    item.get("answer_position"),
                    item.get("answer_sheet", ""),
                    max_rows=max_rows,
                    max_columns=max_columns,
                )
                if list(cells) != list(input_cells):
                    raise SpreadsheetRegionError(
                        "Candidate region coordinates differ from input coordinates"
                    )
            except Exception as exc:
                valid = False
                invalid_reason = f"failed_to_extract_candidate_regions: {exc}"

        candidate = {
            candidate_id_key: candidate_id,
            "source_run_dir": str(spec["run_dir"]),
            "output_file": str(output_path),
            "valid": valid,
            "invalid_reason": invalid_reason,
            "region_cell_count": len(cells) if cells is not None else None,
            "region_hash": normalized_region_hash(cells) if cells is not None else None,
            "aggregation_score": None,
            "selected": False,
            "_cells": cells,
        }
        candidate.update(spec.get("trace_metadata") or {})
        candidate.update(logprob_summary(record))
        candidates.append(candidate)

    selection = select_spreadsheet_medoid(
        candidates,
        candidate_id_key=candidate_id_key,
        selected_id_field=selected_id_field,
        rank_getter=rank_getter,
        fallback_source=fallback_source,
        logprob_field=logprob_field,
    )
    return {
        "id": sample_id,
        "instruction_type": item.get("instruction_type"),
        "answer_position": item.get("answer_position"),
        "answer_sheet": item.get("answer_sheet"),
        "format_valid": selection["format_valid"],
        "execution_success": selection["format_valid"],
        selected_id_field: selection[selected_id_field],
        "selected_source_file": selection["selected_source_file"],
        "valid_candidate_count": selection["valid_candidate_count"],
        "trace": {
            "aggregation": "equal_weight_region_medoid",
            "boundary_sources": ["input", *candidate_ids],
            "max_rows_by_sheet": max_rows,
            "max_columns_by_sheet": max_columns,
            "input_region_error": input_error,
            **selection,
        },
    }


def aggregate_spreadsheet_sample(
    item: Mapping[str, Any],
    records_by_format: Mapping[str, Mapping[str, Any] | None],
    run_dirs: Mapping[str, Path],
    input_path: Path,
    format_order: Sequence[str] = FORMAT_ORDER,
    logprob_field: str | None = "sequence_logprob_mean",
) -> Dict[str, Any]:
    """Select a workbook medoid using only input/candidates and public metadata."""
    sample_id = str(item["id"])
    candidate_specs = [
        {
            "format": format_name,
            "record": records_by_format.get(format_name),
            "run_dir": run_dirs[format_name],
            "output_path": candidate_output_path(
                sample_id,
                format_name,
                records_by_format.get(format_name),
                run_dirs,
            ),
        }
        for format_name in format_order
    ]
    rank_map = {format_name: index for index, format_name in enumerate(format_order)}
    return aggregate_spreadsheet_candidates(
        item,
        candidate_specs,
        input_path,
        rank_getter=lambda candidate: rank_map[candidate["format"]],
        logprob_field=logprob_field,
    )


def copy_selected_workbooks(
    aggregated_rows: Sequence[Mapping[str, Any]], output_dir: Path
) -> int:
    spreadsheet_dir = Path(output_dir) / "spreadsheet"
    spreadsheet_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for row in aggregated_rows:
        source = row.get("selected_source_file")
        if not row.get("format_valid") or not source:
            continue
        destination = spreadsheet_dir / f"1_{row['id']}_output.xlsx"
        shutil.copy2(source, destination)
        copied += 1
    return copied


def _find_dataset_workbook(
    dataset_root: Path, item: Mapping[str, Any], suffix: str
) -> Path:
    sample_id = str(item["id"])
    sample_dir = dataset_root / item.get(
        "spreadsheet_path", os.path.join("spreadsheet", sample_id)
    )
    expected = sample_dir / f"1_{sample_id}_{suffix}.xlsx"
    if expected.exists():
        return expected
    candidates = sorted(sample_dir.glob(f"1_*_{suffix}.xls*"))
    if len(candidates) == 1:
        return candidates[0]
    return expected


def spreadsheet_input_path(
    dataset_root: Path, item: Mapping[str, Any]
) -> Path:
    return _find_dataset_workbook(dataset_root, item, "init")


def evaluate_spreadsheet_vote(
    aggregated_rows: Sequence[Mapping[str, Any]],
    dataset_by_id: Mapping[str, Mapping[str, Any]],
    dataset_root: Path,
    output_dir: Path,
    *,
    selected_id_field: str = "selected_format",
) -> tuple[list[dict], dict]:
    eval_rows = []
    score_lists = defaultdict(list)
    spreadsheet_dir = Path(output_dir) / "spreadsheet"
    for aggregate in aggregated_rows:
        sample_id = str(aggregate["id"])
        item = dataset_by_id[sample_id]
        gold_file = _find_dataset_workbook(dataset_root, item, "golden")
        generated_file = spreadsheet_dir / f"1_{sample_id}_output.xlsx"
        passed, message = compare_workbooks(
            str(gold_file),
            str(generated_file),
            item.get("instruction_type", ""),
            item.get("answer_position", ""),
            item.get("answer_sheet", ""),
        )
        result_value = int(bool(passed))
        entry = {
            "id": sample_id,
            "instruction": item.get("instruction"),
            "spreadsheet_path": item.get("spreadsheet_path"),
            "instruction_type": item.get("instruction_type"),
            "answer_position": item.get("answer_position"),
            "answer_sheet": item.get("answer_sheet"),
            "execution_success": bool(aggregate.get("execution_success")),
            "format_valid": bool(aggregate.get("format_valid")),
            "error": None if aggregate.get("format_valid") else "all_candidates_invalid",
            "test_case_results": [result_value],
            "test_case_messages": [message],
            "total_soft_restriction": float(result_value),
            "total_hard_restriction": float(result_value),
            "table_metadata": None,
            selected_id_field: aggregate.get(selected_id_field),
        }
        eval_rows.append(entry)
        is_sheet = "Sheet" in str(item.get("instruction_type", ""))
        for prefix in ("soft", "hard"):
            score_lists[f"{prefix}_all"].append(float(result_value))
            score_lists[f"{prefix}_{'sheet' if is_sheet else 'cell'}"].append(
                float(result_value)
            )

    accuracy = {
        key: round(sum(score_lists[key]) / len(score_lists[key]), 4)
        if score_lists[key]
        else 0.0
        for key in (
            "soft_all",
            "hard_all",
            "soft_cell",
            "hard_cell",
            "soft_sheet",
            "hard_sheet",
        )
    }
    return eval_rows, accuracy


def spreadsheet_diagnostics(
    aggregated_rows: Sequence[Mapping[str, Any]],
) -> Dict[str, Any]:
    ties = [row for row in aggregated_rows if row["trace"]["tie"]]
    logprob_breaks = [
        row for row in ties if row["trace"]["tie_break_source"] == "logprob"
    ]
    format_fallbacks = [
        row for row in ties if row["trace"]["tie_break_source"] == "format_order"
    ]
    selections = [
        row["selected_format"]
        for row in aggregated_rows
        if row["selected_format"] is not None
    ]
    pairwise = [
        row["trace"]["average_pairwise_region_agreement"]
        for row in aggregated_rows
        if row["trace"]["average_pairwise_region_agreement"] is not None
    ]
    medoid_scores = [
        row["trace"]["medoid_score"]
        for row in aggregated_rows
        if row["trace"]["medoid_score"] is not None
    ]
    invalid_reasons = [
        candidate["invalid_reason"]
        for row in aggregated_rows
        for candidate in row["trace"]["candidates"]
        if not candidate["valid"]
    ]
    return {
        "num_samples": len(aggregated_rows),
        "valid_candidate_count_distribution": count_distribution(
            row["valid_candidate_count"] for row in aggregated_rows
        ),
        "all_candidates_invalid_samples": sum(
            1 for row in aggregated_rows if not row["format_valid"]
        ),
        "tie_samples": len(ties),
        "tie_rate": safe_rate(len(ties), len(aggregated_rows)),
        "logprob_tie_breaks": len(logprob_breaks),
        "logprob_tie_break_coverage": safe_rate(len(logprob_breaks), len(ties)),
        "format_order_fallbacks": len(format_fallbacks),
        "selected_format_distribution": count_distribution(selections),
        "average_pairwise_region_agreement": numeric_summary(pairwise),
        "medoid_score_summary": numeric_summary(medoid_scores),
        "medoid_score_distribution_rounded_6dp": count_distribution(
            round(score, 6) for score in medoid_scores
        ),
        "invalid_candidate_reason_distribution": count_distribution(invalid_reasons),
    }
