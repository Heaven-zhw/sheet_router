import argparse
import base64
import copy
import os
import shutil
import subprocess
import tempfile
import traceback
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image


DEFAULT_HOST = "0.0.0.0"
DEFAULT_PORT = 8007
DEFAULT_DPI = 200


def _safe_basename(filename, fallback):
    name = os.path.basename(filename or fallback)
    return name or fallback


def _find_binary(env_name, default_name):
    binary = os.environ.get(env_name) or shutil.which(default_name)
    if not binary:
        raise RuntimeError(
            f"Cannot find '{default_name}'. "
            f"Set {env_name} or install the package first."
        )
    return binary


def _run_command(cmd, timeout=120):
    proc = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=timeout,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "Command failed:\n"
            f"{' '.join(cmd)}\n"
            f"stdout:\n{proc.stdout}\n"
            f"stderr:\n{proc.stderr}"
        )
    return proc


def _resolve_sheet_name(workbook, sheet_name):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
            )
        return sheet_name
    return workbook.sheetnames[0]


def used_range(file_path, sheet_name=None):
    workbook = load_workbook(file_path, data_only=False, read_only=True)
    sheet_name = _resolve_sheet_name(workbook, sheet_name)
    worksheet = workbook[sheet_name]
    dimension = worksheet.calculate_dimension()
    if dimension == "A1:A1" and worksheet["A1"].value is None:
        return "A1:A1"
    return dimension


def _copy_range_to_workbook(src_path, sheet_name, cell_range, dst_path):
    src_wb = load_workbook(src_path, data_only=False)
    sheet_name = _resolve_sheet_name(src_wb, sheet_name)
    src_ws = src_wb[sheet_name]

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = sheet_name

    dst_ws.sheet_view.showGridLines = True
    dst_ws.print_options.gridLines = True
    dst_ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    dst_ws.page_setup.fitToWidth = 1
    dst_ws.page_setup.fitToHeight = 1
    dst_ws.page_margins = PageMargins(
        left=0.1,
        right=0.1,
        top=0.1,
        bottom=0.1,
        header=0.0,
        footer=0.0,
    )

    row_count = max_row - min_row + 1
    col_count = max_col - min_col + 1
    if col_count > row_count:
        dst_ws.page_setup.orientation = "landscape"
    else:
        dst_ws.page_setup.orientation = "portrait"

    for src_col in range(min_col, max_col + 1):
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(src_col - min_col + 1)
        src_dim = src_ws.column_dimensions[src_letter]
        if src_dim.width is not None:
            dst_ws.column_dimensions[dst_letter].width = src_dim.width
        dst_ws.column_dimensions[dst_letter].hidden = src_dim.hidden

    for src_row in range(min_row, max_row + 1):
        dst_row = src_row - min_row + 1
        src_dim = src_ws.row_dimensions[src_row]
        if src_dim.height is not None:
            dst_ws.row_dimensions[dst_row].height = src_dim.height
        dst_ws.row_dimensions[dst_row].hidden = src_dim.hidden

        for src_col in range(min_col, max_col + 1):
            dst_col = src_col - min_col + 1
            src_cell = src_ws.cell(row=src_row, column=src_col)
            dst_cell = dst_ws.cell(row=dst_row, column=dst_col)

            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = copy.copy(src_cell._style)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
            if src_cell.font:
                dst_cell.font = copy.copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy.copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border = copy.copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy.copy(src_cell.alignment)
            if src_cell.protection:
                dst_cell.protection = copy.copy(src_cell.protection)
            if src_cell.hyperlink:
                dst_cell._hyperlink = copy.copy(src_cell.hyperlink)
            if src_cell.comment:
                dst_cell.comment = copy.copy(src_cell.comment)

    for merged_range in src_ws.merged_cells.ranges:
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = merged_range.bounds
        if (
            merged_min_col >= min_col
            and merged_max_col <= max_col
            and merged_min_row >= min_row
            and merged_max_row <= max_row
        ):
            dst_min_col = merged_min_col - min_col + 1
            dst_max_col = merged_max_col - min_col + 1
            dst_min_row = merged_min_row - min_row + 1
            dst_max_row = merged_max_row - min_row + 1
            dst_ws.merge_cells(
                start_row=dst_min_row,
                start_column=dst_min_col,
                end_row=dst_max_row,
                end_column=dst_max_col,
            )

    dst_ws.print_area = f"A1:{get_column_letter(col_count)}{row_count}"
    dst_wb.save(dst_path)


def _convert_workbook_to_png(xlsx_path, png_path, dpi=DEFAULT_DPI):
    soffice_bin = _find_binary("SOFFICE_BIN", "libreoffice")
    pdftoppm_bin = _find_binary("PDFTOPPM_BIN", "pdftoppm")

    work_dir = os.path.dirname(os.path.abspath(png_path))
    stem = Path(xlsx_path).stem
    pdf_path = os.path.join(work_dir, f"{stem}.pdf")
    png_prefix = os.path.join(work_dir, stem)

    profile_dir = os.path.join(work_dir, "libreoffice-profile")
    os.makedirs(profile_dir, exist_ok=True)
    profile_uri = Path(profile_dir).resolve().as_uri()

    _run_command(
        [
            soffice_bin,
            f"-env:UserInstallation={profile_uri}",
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to",
            "pdf",
            "--outdir",
            work_dir,
            xlsx_path,
        ]
    )

    if not os.path.exists(pdf_path):
        raise RuntimeError(f"LibreOffice did not produce the expected PDF: {pdf_path}")

    _run_command(
        [
            pdftoppm_bin,
            "-png",
            "-singlefile",
            "-r",
            str(dpi),
            pdf_path,
            png_prefix,
        ]
    )

    generated_png = f"{png_prefix}.png"
    if not os.path.exists(generated_png):
        raise RuntimeError(f"pdftoppm did not produce the expected PNG: {generated_png}")

    if os.path.abspath(generated_png) != os.path.abspath(png_path):
        shutil.move(generated_png, png_path)


def _trim_whitespace(in_path, out_path, padding=16, threshold=245):
    image = Image.open(in_path).convert("RGB")
    gray = image.convert("L")
    mask = gray.point(lambda pixel: 0 if pixel >= threshold else 255, mode="1")
    bbox = mask.getbbox()

    if bbox:
        left = max(0, bbox[0] - padding)
        upper = max(0, bbox[1] - padding)
        right = min(image.width, bbox[2] + padding)
        lower = min(image.height, bbox[3] + padding)
        image = image.crop((left, upper, right, lower))

    image.save(out_path)


def render_excel_range_to_png(src_path, sheet_name=None, cell_range=None, out_path=None, dpi=DEFAULT_DPI):
    if out_path is None:
        out_path = os.path.abspath("out.png")

    with tempfile.TemporaryDirectory(prefix="excel2image_linux_") as tmpdir:
        sheet_name = sheet_name or None
        cell_range = cell_range or used_range(src_path, sheet_name)

        staged_xlsx = os.path.join(tmpdir, "selected_range.xlsx")
        raw_png = os.path.join(tmpdir, "raw.png")

        _copy_range_to_workbook(src_path, sheet_name, cell_range, staged_xlsx)
        _convert_workbook_to_png(staged_xlsx, raw_png, dpi=dpi)
        _trim_whitespace(raw_png, out_path)

    return out_path


def convert_excel_to_base64(src_path, sheet_name=None, cell_range=None, dpi=DEFAULT_DPI):
    with tempfile.TemporaryDirectory(prefix="excel2image_resp_") as tmpdir:
        out_path = os.path.join(tmpdir, "out.png")
        render_excel_range_to_png(
            src_path=src_path,
            sheet_name=sheet_name,
            cell_range=cell_range,
            out_path=out_path,
            dpi=dpi,
        )
        with open(out_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode("utf-8")


def create_app():
    try:
        from flask import Flask, jsonify, request
    except ImportError as exc:
        raise RuntimeError(
            "Flask is required for server mode. Install it with `pip install flask`."
        ) from exc

    app = Flask(__name__)

    @app.route("/upload", methods=["POST"])
    def upload_file():
        if "file" not in request.files:
            return jsonify({"error": "missing file"}), 400

        uploaded_file = request.files["file"]
        if uploaded_file.filename == "":
            return jsonify({"error": "empty filename"}), 400

        filename = _safe_basename(uploaded_file.filename, "input.xlsx")
        uploaded_file.save(filename)
        return jsonify({"message": "file uploaded", "filename": filename}), 200

    @app.route("/excel2img", methods=["POST"])
    def excel_to_img():
        if "file" not in request.files:
            return jsonify({"error": "missing file"}), 400

        uploaded_file = request.files["file"]
        sheet_name = request.form.get("page")
        cell_range = request.form.get("_range")

        with tempfile.TemporaryDirectory(prefix="excel2image_api_") as tmpdir:
            in_path = os.path.join(
                tmpdir, _safe_basename(uploaded_file.filename, "input.xlsx")
            )
            uploaded_file.save(in_path)

            try:
                image_base64 = convert_excel_to_base64(
                    src_path=in_path,
                    sheet_name=sheet_name,
                    cell_range=cell_range,
                )
            except Exception:
                return jsonify({"error": traceback.format_exc()}), 500

        return jsonify({"image_base64": image_base64})

    @app.route("/healthz", methods=["GET"])
    def healthz():
        return jsonify({"status": "ok"})

    return app


def main():
    parser = argparse.ArgumentParser(
        description="Ubuntu-friendly Excel-to-image service for SpreadsheetAgent."
    )
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--dpi", type=int, default=DEFAULT_DPI)
    parser.add_argument("--input", help="Run once in CLI mode with a local Excel file.")
    parser.add_argument("--sheet", help="Sheet name for CLI mode.")
    parser.add_argument("--range", dest="cell_range", help="Cell range such as A1:F20.")
    parser.add_argument("--output", help="PNG output path for CLI mode.")
    args = parser.parse_args()

    if args.input:
        output_path = args.output or os.path.abspath("out.png")
        render_excel_range_to_png(
            src_path=args.input,
            sheet_name=args.sheet,
            cell_range=args.cell_range,
            out_path=output_path,
            dpi=args.dpi,
        )
        print(output_path)
        return

    app = create_app()
    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()
