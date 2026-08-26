import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import openpyxl

from core.eval.spreadsheet import compare_cell_value, compare_workbooks
from core.eval.spreadsheet_regions import (
    SpreadsheetRegionError,
    WorkbookRegionError,
    extract_normalized_region_cells,
    iter_unique_region_coordinates,
    parse_answer_regions,
)
from evaluate.reevaluate_spreadsheetbench import evaluate_run


REPO_ROOT = Path(__file__).resolve().parents[1]
VERIFIED_ROOT = (
    REPO_ROOT / "dataset/spreadsheetbench/spreadsheetbench_verified_400"
)


def save_workbook(path, sheets):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_name, values in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for coordinate, value in values.items():
            worksheet[coordinate] = value
    workbook.save(path)
    workbook.close()


class SpreadsheetRegionTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def paths(self, gold_sheets, generated_sheets=None):
        gold_path = self.root / "gold.xlsx"
        generated_path = self.root / "generated.xlsx"
        save_workbook(gold_path, gold_sheets)
        save_workbook(
            generated_path,
            gold_sheets if generated_sheets is None else generated_sheets,
        )
        return gold_path, generated_path

    def compare(self, gold_path, generated_path, position, answer_sheet=""):
        return compare_workbooks(
            str(gold_path),
            str(generated_path),
            "Cell-Level Manipulation",
            position,
            answer_sheet,
        )

    def test_single_cell_and_rectangle(self):
        gold, generated = self.paths(
            {"Target": {"A1": "header", "B2": 1, "C3": 2}}
        )
        ok, _ = self.compare(gold, generated, "B2", "Target")
        self.assertTrue(ok)
        cells = extract_normalized_region_cells(generated, "B2:C3", "Target")
        self.assertEqual(
            list(cells),
            [("Target", "B2"), ("Target", "C2"), ("Target", "B3"), ("Target", "C3")],
        )

        # The original four-argument evaluator API remains valid.
        ok, _ = compare_workbooks(
            str(gold), str(generated), "Cell-Level Manipulation", "A1"
        )
        self.assertTrue(ok)

    def test_all_comma_separated_regions_must_match(self):
        gold, generated = self.paths(
            {"Target": {"A1": "gold", "C1": "same"}},
            {"Target": {"A1": "wrong", "C1": "same"}},
        )
        ok, message = self.compare(gold, generated, "A1,C1", "Target")
        self.assertFalse(ok)
        self.assertIn("A1", message)

    def test_answer_sheet_uses_first_existing_candidate_then_first_sheet(self):
        gold, generated = self.paths(
            {"First": {"B2": "ignored"}, "Wanted": {"B2": "answer"}},
            {"First": {"B2": "different"}, "Wanted": {"B2": "answer"}},
        )
        ok, _ = self.compare(gold, generated, "B2", "Missing, 'Wanted'")
        self.assertTrue(ok)

        workbook = openpyxl.load_workbook(gold, data_only=True)
        try:
            regions = parse_answer_regions(workbook, "B2", "Missing")
            self.assertEqual(regions[0].sheet_name, "First")
        finally:
            workbook.close()

    def test_asymmetric_single_and_double_sheet_quotes(self):
        path = self.root / "quotes.xlsx"
        save_workbook(path, {"Odd Name": {"A1": 1, "B1": 2, "C1": 3}})
        cells = extract_normalized_region_cells(
            path,
            "Odd Name'!A1,'Odd Name!B1,\"Odd Name\"!C1",
        )
        self.assertEqual(list(cells.values()), [1.0, 2.0, 3.0])

    def test_misplaced_quotes_and_comma_inside_sheet_name(self):
        path = self.root / "messy_quotes.xlsx"
        save_workbook(
            path,
            {
                "b2b, sez, de": {"A1": 1},
                "Other": {"A1": 2},
            },
        )
        cells = extract_normalized_region_cells(
            path,
            "'b2b, sez, de'!A1,'Other!'A1",
        )
        self.assertEqual(
            list(cells),
            [("b2b, sez, de", "A1"), ("Other", "A1")],
        )
        unqualified = extract_normalized_region_cells(
            path,
            "A1",
            "b2b, sez, de",
        )
        self.assertEqual(list(unqualified), [("b2b, sez, de", "A1")])

    def test_quotes_around_entire_region_and_cell_range(self):
        path = self.root / "quoted_regions.xlsx"
        save_workbook(path, {"99250": {"A1": 1}, "Sheet1": {"A2": 2}})
        cells = extract_normalized_region_cells(
            path,
            "'99250!A1','Sheet1'!'A2'",
        )
        self.assertEqual(
            list(cells),
            [("99250", "A1"), ("Sheet1", "A2")],
        )

    def test_whole_columns_use_explicit_boundary(self):
        path = self.root / "columns.xlsx"
        save_workbook(path, {"Target": {"A1": 1, "G2": 2}})
        cells = extract_normalized_region_cells(
            path, "'Target'!A:G", max_rows={"Target": 3}
        )
        self.assertEqual(len(cells), 21)
        self.assertEqual(next(iter(cells)), ("Target", "A1"))
        self.assertEqual(next(reversed(cells)), ("Target", "G3"))

        row_cells = extract_normalized_region_cells(
            path, "'Target'!1:2", max_columns={"Target": 3}
        )
        self.assertEqual(len(row_cells), 6)
        self.assertEqual(next(reversed(row_cells)), ("Target", "C2"))

    def test_whole_columns_compare_generated_extra_rows(self):
        gold, generated = self.paths(
            {"Target": {"A1": "same", "G2": "same"}},
            {"Target": {"A1": "same", "G2": "same", "A4": "extra"}},
        )
        ok, message = self.compare(gold, generated, "Target'!A:G")
        self.assertFalse(ok)
        self.assertIn("A4", message)

    def test_overlapping_regions_are_deduplicated_in_order(self):
        path = self.root / "overlap.xlsx"
        save_workbook(path, {"Target": {"A1": 1, "A2": 2, "B2": 3}})
        workbook = openpyxl.load_workbook(path, data_only=True)
        try:
            regions = parse_answer_regions(workbook, "A1:A2,A2:B2", "Target")
            coordinates = list(iter_unique_region_coordinates(regions))
        finally:
            workbook.close()
        self.assertEqual(
            coordinates,
            [("Target", "A1"), ("Target", "A2"), ("Target", "B2")],
        )

    def test_legacy_empty_and_numeric_value_semantics(self):
        self.assertTrue(compare_cell_value(None, ""))
        self.assertTrue(compare_cell_value("", None))
        self.assertTrue(compare_cell_value("1.234", 1.234))
        self.assertFalse(compare_cell_value(None, 0))

        gold, generated = self.paths(
            {"Target": {"A1": 1.234}},
            {"Target": {"A1": "1.234"}},
        )
        ok, _ = self.compare(gold, generated, "A1", "Target")
        self.assertTrue(ok)

    def test_none_empty_and_invalid_ranges_return_clear_errors(self):
        gold, generated = self.paths({"Target": {"A1": 1}})
        for position in (None, ""):
            ok, message = self.compare(gold, generated, position, "Target")
            self.assertFalse(ok)
            self.assertIn("answer_position is empty", message)

        for position in ("not-a-range", "A1:G", "A0", "B2:A1"):
            ok, message = self.compare(gold, generated, position, "Target")
            self.assertFalse(ok)
            self.assertIn("Invalid answer region", message)

    def test_missing_sheet_and_files_are_clear(self):
        gold, generated = self.paths({"Target": {"A1": 1}})
        ok, message = self.compare(gold, generated, "Missing!A1")
        self.assertFalse(ok)
        self.assertIn("Worksheet not found", message)

        other_generated = self.root / "other_generated.xlsx"
        save_workbook(other_generated, {"Other": {"A1": 1}})
        ok, message = self.compare(gold, other_generated, "Target!A1")
        self.assertFalse(ok)
        self.assertIn("Worksheet not found in generated workbook", message)

        missing = self.root / "missing.xlsx"
        ok, message = self.compare(gold, missing, "A1", "Target")
        self.assertFalse(ok)
        self.assertIn("Generated file not found", message)
        with self.assertRaisesRegex(WorkbookRegionError, "Workbook file not found"):
            extract_normalized_region_cells(missing, "A1", "Target")

        ok, message = self.compare(missing, generated, "A1", "Target")
        self.assertFalse(ok)
        self.assertIn("Ground truth file not found", message)

    def test_damaged_workbook_is_clear(self):
        gold, _ = self.paths({"Target": {"A1": 1}})
        damaged = self.root / "damaged.xlsx"
        damaged.write_text("not an xlsx archive", encoding="utf-8")
        ok, message = self.compare(gold, damaged, "A1", "Target")
        self.assertFalse(ok)
        self.assertIn("Failed to open generated workbook", message)
        with self.assertRaisesRegex(WorkbookRegionError, "Failed to open workbook"):
            extract_normalized_region_cells(damaged, "A1", "Target")


class Verified400SmokeTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        with open(VERIFIED_ROOT / "dataset.json", "r", encoding="utf-8") as handle:
            cls.dataset = {str(item["id"]): item for item in json.load(handle)}

    def extract_sample(self, sample_id):
        item = self.dataset[sample_id]
        workbook_path = (
            VERIFIED_ROOT
            / item["spreadsheet_path"]
            / f"1_{sample_id}_golden.xlsx"
        )
        return extract_normalized_region_cells(
            workbook_path,
            item["answer_position"],
            item.get("answer_sheet", ""),
        )

    def test_regular_single_region_13_1(self):
        self.assertEqual(len(self.extract_sample("13-1")), 120)

    def test_multi_region_41_47(self):
        cells = self.extract_sample("41-47")
        self.assertEqual(len(cells), 6403)
        self.assertEqual(next(iter(cells)), ("OUT CAS", "A2"))

    def test_whole_column_regions_283_32(self):
        cells = self.extract_sample("283-32")
        self.assertEqual(len(cells), 28)
        self.assertIn(("Sheet3", "G2"), cells)
        self.assertIn(("Sheet4", "G2"), cells)


class ReevaluateOutputTest(unittest.TestCase):
    def test_existing_tagged_output_is_not_overwritten(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            run_dir = Path(temp_dir) / "pot_test"
            (run_dir / "spreadsheet").mkdir(parents=True)
            (run_dir / "spreadsheet_pot_eval_existing.json").touch()
            args = SimpleNamespace(output_tag="existing", overwrite=False)
            with self.assertRaisesRegex(FileExistsError, "Choose a new --output_tag"):
                evaluate_run(run_dir, [], Path(temp_dir), {}, args)


if __name__ == "__main__":
    unittest.main()
